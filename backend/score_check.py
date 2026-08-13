import io
import os
import re
import html
import uuid
from datetime import datetime, timedelta, time
from typing import Any

import openpyxl
from fastapi import APIRouter, UploadFile, File, Request
from fastapi.responses import HTMLResponse, FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()
_EXPORT_DIR = os.path.join("uploads", "score_check")
os.makedirs(_EXPORT_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Generic Excel helpers
# ---------------------------------------------------------------------------

def _clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _unique_headers(values) -> list[str]:
    result = []
    counts = {}
    for index, value in enumerate(values):
        base = _clean_header(value) or f"Column {index + 1}"
        counts[base] = counts.get(base, 0) + 1
        name = base if counts[base] == 1 else f"{base} [{counts[base]}]"
        result.append(name)
    return result


def _read_xlsx(raw: bytes):
    wb = openpyxl.load_workbook(
        io.BytesIO(raw),
        read_only=True,
        data_only=True,
    )
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if not raw_headers:
            raise ValueError("Fișierul nu conține antet.")

        headers = _unique_headers(raw_headers)
        records = []

        for source_row in rows:
            if not any(
                value is not None and str(value).strip() != ""
                for value in source_row
            ):
                continue

            record = {}
            for index, header in enumerate(headers):
                record[header] = (
                    source_row[index]
                    if index < len(source_row)
                    else None
                )
            records.append(record)

        return headers, records
    finally:
        wb.close()


def _num(value):
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text.casefold() in {"n/a", "na", "-", "none", "null"}:
        return None

    text = text.replace("%", "").replace(",", ".")
    text = re.sub(r"[^0-9+\-.]", "", text)

    if not text or text in {"+", "-", ".", "+.", "-."}:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _find_header(headers, include_any, exclude_any=()):
    include_any = tuple(item.casefold() for item in include_any)
    exclude_any = tuple(item.casefold() for item in exclude_any)

    for header in headers:
        key = header.casefold()
        if include_any and not any(item in key for item in include_any):
            continue
        if exclude_any and any(item in key for item in exclude_any):
            continue
        return header
    return None


def _get_first(row, candidates):
    for candidate in candidates:
        if candidate in row and row[candidate] not in (None, ""):
            return row[candidate]
    return None


# ---------------------------------------------------------------------------
# Driver identity / FICO helpers
# ---------------------------------------------------------------------------

def _encrypted_key(row):
    first = str(row.get("First Name") or "").strip()
    last = str(row.get("Last Name") or "").strip()
    return f"{first}\u241f{last}"


def _display_encrypted(row):
    first = str(row.get("First Name") or "").strip()
    last = str(row.get("Last Name") or "").strip()
    return " | ".join(part for part in (first, last) if part)


def _real_name(row):
    first = str(row.get("First Name") or "").strip()
    last = str(row.get("Last Name") or "").strip()
    return " ".join(part for part in (first, last) if part).strip()


def _current_fico_direct(row):
    if not row:
        return None

    keys = list(row.keys())

    preferred = []
    fallback = []

    for key in keys:
        low = key.casefold()
        if "fico" not in low or "score" not in low:
            continue
        if "last period" in low or "delta" in low or "Δ" in key or "change" in low:
            continue

        if any(term in low for term in ("this period", "current", "recent")):
            preferred.append(key)
        else:
            fallback.append(key)

    for key in preferred + fallback:
        value = _num(row.get(key))
        if value is not None and 300 <= value <= 850:
            return int(round(value))

    return None


def _fico_last(comp_row):
    exact = comp_row.get("FICO® Safe Driving Score Last Period")
    value = _num(exact)
    if value is not None:
        return int(round(value))

    for key, raw in comp_row.items():
        low = key.casefold()
        if "fico" in low and "score" in low and "last period" in low:
            value = _num(raw)
            if value is not None:
                return int(round(value))
    return None


def _fico_delta(comp_row):
    exact = comp_row.get("FICO® Safe Driving Score Δ")
    value = _num(exact)
    if value is not None:
        return value

    for key, raw in comp_row.items():
        low = key.casefold()
        if "fico" in low and ("delta" in low or "Δ" in key or "change" in low):
            value = _num(raw)
            if value is not None:
                return value
    return None


def _fico_recent(comp_row, driver_row):
    # Prefer a real current/recent FICO column when Amazon provides one.
    direct = _current_fico_direct(comp_row)
    if direct is not None:
        return direct

    direct = _current_fico_direct(driver_row)
    if direct is not None:
        return direct

    # Fallback used by the reports discussed here:
    # FICO Recent = FICO Last Period + FICO Δ (absolute point change).
    last = _fico_last(comp_row)
    delta = _fico_delta(comp_row)

    if last is None:
        return None
    if delta is None:
        return last

    return int(round(max(300, min(850, last + delta))))


def _driver_code(shift_row):
    if not shift_row:
        return ""

    candidates = [
        "Transporter ID",
        "Transporter-ID",
        "Associate ID",
        "Associate-ID",
        "Driver ID",
        "Driver-ID",
        "Vehicle Identifier",
    ]

    value = _get_first(shift_row, candidates)
    return "" if value is None else str(value).strip()


def _vehicle_identifier(shift_row):
    if not shift_row:
        return ""
    value = shift_row.get("Vehicle Identifier")
    return "" if value is None else str(value).strip()


# ---------------------------------------------------------------------------
# Hours + km matching
# ---------------------------------------------------------------------------

def _hours_to_minutes(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + value.second / 60

    if isinstance(value, time):
        return value.hour * 60 + value.minute + value.second / 60

    if isinstance(value, timedelta):
        return value.total_seconds() / 60

    if isinstance(value, (int, float)):
        number = float(value)

        # Excel time fraction.
        if 0 <= number < 1:
            return number * 24 * 60

        # Explicit decimal hours.
        if 0 <= number <= 24:
            return number * 60

        # Larger numeric time fields are treated as minutes.
        if 24 < number <= 24 * 60:
            return number

        return None

    text = str(value).strip()
    if not text:
        return None

    match = re.fullmatch(
        r"(?:(\d+)\s*h(?:ours?)?\s*)?(\d{1,2})?\s*(?:m|min|minutes?)?",
        text,
        flags=re.IGNORECASE,
    )
    if match and (match.group(1) or match.group(2)):
        hours = int(match.group(1) or 0)
        minutes = int(match.group(2) or 0)
        return hours * 60 + minutes

    match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        seconds = int(match.group(3) or 0)
        return hours * 60 + minutes + seconds / 60

    number = _num(text)
    if number is not None:
        if 0 <= number <= 24:
            return number * 60
        if 24 < number <= 1440:
            return number

    return None


def _find_hours_value(comp_row, driver_row):
    # Only use fields explicitly describing total driver hours / worked hours.
    for row in (driver_row or {}, comp_row or {}):
        for key, raw in row.items():
            low = key.casefold()
            if (
                ("total driver hours" in low)
                or ("hours worked" in low)
                or ("worked hours" in low)
                or ("total worked" in low)
            ):
                minutes = _hours_to_minutes(raw)
                if minutes is not None:
                    return minutes
    return None


def _shift_hours_minutes(shift_row):
    if not shift_row:
        return None
    for key in (
        "Total Driver Hours",
        "Hours Worked",
        "Worked Hours",
        "Total Worked",
    ):
        if key in shift_row:
            value = _hours_to_minutes(shift_row.get(key))
            if value is not None:
                return value
    return None


def _comp_km(comp_row):
    value = _num(comp_row.get("Total Driver km This Period"))
    if value is not None:
        return value

    for key, raw in comp_row.items():
        low = key.casefold()
        if "total driver km" in low and "this period" in low:
            value = _num(raw)
            if value is not None:
                return value
    return None


def _shift_km(shift_row):
    value = _num(shift_row.get("Total Driver km"))
    if value is not None:
        return value

    for key, raw in shift_row.items():
        low = key.casefold()
        if "total driver km" in low and "period" not in low:
            value = _num(raw)
            if value is not None:
                return value
    return None


def _candidate_score(comp_row, driver_row, shift_row):
    comp_km = _comp_km(comp_row)
    shift_km = _shift_km(shift_row)

    if comp_km is None or shift_km is None:
        return 999999.0, None, None

    km_diff = abs(comp_km - shift_km)

    encrypted_hours = _find_hours_value(comp_row, driver_row)
    shift_hours = _shift_hours_minutes(shift_row)

    hours_diff = None
    score = km_diff

    if encrypted_hours is not None and shift_hours is not None:
        hours_diff = abs(encrypted_hours - shift_hours)
        # Hours act as a tie-breaker; km remains the primary signal.
        score += (hours_diff / 60.0) * 0.25

    return score, km_diff, hours_diff


def _assign_real_names(comp_rows, driver_by_key, shift_rows):
    candidate_lists = {}
    all_pairs = []

    for comp_index, comp in enumerate(comp_rows):
        driver = driver_by_key.get(_encrypted_key(comp), {})
        candidates = []

        for shift_index, shift in enumerate(shift_rows):
            score, km_diff, hours_diff = _candidate_score(
                comp,
                driver,
                shift,
            )
            item = {
                "score": score,
                "km_diff": km_diff,
                "hours_diff": hours_diff,
                "shift_index": shift_index,
            }
            candidates.append(item)
            all_pairs.append((score, comp_index, shift_index))

        candidates.sort(
            key=lambda item: (
                item["score"],
                _real_name(shift_rows[item["shift_index"]]).casefold(),
            )
        )
        candidate_lists[comp_index] = candidates

    # Global one-to-one allocation so the same real driver is not assigned twice.
    used_comp = set()
    used_shift = set()
    chosen = {}

    for score, comp_index, shift_index in sorted(all_pairs, key=lambda x: x[0]):
        if comp_index in used_comp or shift_index in used_shift:
            continue
        chosen[comp_index] = shift_index
        used_comp.add(comp_index)
        used_shift.add(shift_index)

    results = []

    for comp_index, comp in enumerate(comp_rows):
        candidates = candidate_lists.get(comp_index, [])
        selected_shift_index = chosen.get(comp_index)

        if selected_shift_index is None and candidates:
            selected_shift_index = candidates[0]["shift_index"]

        if selected_shift_index is None:
            results.append({
                "name": "",
                "confidence": "Low",
                "km_diff": None,
                "hours_diff": None,
                "candidate_2": "",
                "candidate_3": "",
                "shift_row": {},
            })
            continue

        selected = next(
            item
            for item in candidates
            if item["shift_index"] == selected_shift_index
        )

        alternatives = [
            item for item in candidates
            if item["shift_index"] != selected_shift_index
        ]

        next_score = alternatives[0]["score"] if alternatives else None
        score_gap = (
            next_score - selected["score"]
            if next_score is not None
            else 999
        )

        km_diff = selected["km_diff"]
        hours_diff = selected["hours_diff"]

        hours_good_high = hours_diff is None or hours_diff <= 5
        hours_good_medium = hours_diff is None or hours_diff <= 20

        if (
            km_diff is not None
            and km_diff <= 0.05
            and hours_good_high
            and score_gap >= 0.10
        ):
            confidence = "High"
        elif (
            km_diff is not None
            and km_diff <= 0.50
            and hours_good_medium
            and score_gap >= 0.03
        ):
            confidence = "Medium"
        else:
            confidence = "Low"

        def alternative_name(position):
            if position >= len(alternatives):
                return ""

            item = alternatives[position]
            shift = shift_rows[item["shift_index"]]

            details = []
            if item["km_diff"] is not None:
                details.append(f"Δkm {item['km_diff']:.2f}")
            if item["hours_diff"] is not None:
                details.append(f"Δore {int(round(item['hours_diff']))}m")

            suffix = f" ({', '.join(details)})" if details else ""
            return _real_name(shift) + suffix

        shift_row = shift_rows[selected_shift_index]

        results.append({
            "name": _real_name(shift_row),
            "confidence": confidence,
            "km_diff": (
                round(km_diff, 3)
                if km_diff is not None
                else None
            ),
            "hours_diff": (
                int(round(hours_diff))
                if hours_diff is not None
                else None
            ),
            "candidate_2": alternative_name(0),
            "candidate_3": alternative_name(1),
            "shift_row": shift_row,
        })

    return results


# ---------------------------------------------------------------------------
# Incident handling + merged rows
# ---------------------------------------------------------------------------

_INCIDENT_TERMS = (
    "speeding",
    "speed",
    "distraction",
    "distracted",
    "hard braking",
    "harsh braking",
    "hard acceleration",
    "harsh acceleration",
    "hard cornering",
    "harsh cornering",
    "seatbelt",
    "following distance",
    "phone manipulation",
    "phone distraction",
)


def _is_incident_header(header):
    low = header.casefold()

    if any(
        blocked in low
        for blocked in (
            "rating",
            "risk",
            "score",
            "fico",
            "last period",
            "delta",
            "difference",
        )
    ):
        return False

    return any(term in low for term in _INCIDENT_TERMS)


def _incident_summary(comp_row, driver_row):
    incidents = []

    for source_name, row in (
        ("Comparison", comp_row or {}),
        ("Driver", driver_row or {}),
    ):
        for key, raw in row.items():
            if not _is_incident_header(key):
                continue

            value = _num(raw)
            if value is not None and value > 0:
                clean_key = re.sub(r"\s+", " ", key).strip()
                incidents.append(
                    f"{clean_key}: {raw}"
                )

    # Keep order but remove exact duplicates.
    return " | ".join(dict.fromkeys(incidents))


def _build_rows(
    comp_headers,
    comp_rows,
    driver_headers,
    driver_rows,
    shift_headers,
    shift_rows,
):
    driver_by_key = {
        _encrypted_key(row): row
        for row in driver_rows
    }

    matches = _assign_real_names(
        comp_rows,
        driver_by_key,
        shift_rows,
    )

    front_headers = [
        "Nume real estimat",
        "Confidence",
        "Diferență km",
        "Diferență ore (minute)",
        "Candidat alternativ 2",
        "Candidat alternativ 3",
        "Encrypted ID",
        "Cod șofer / Transporter ID",
        "Vehicle Identifier",
        "FICO Recent",
        "FICO Last Period",
        "FICO Δ",
        "Ore lucrate",
        "Km Shift Report",
        "Incidente detectate",
    ]

    all_headers = (
        front_headers
        + [f"Comparison · {header}" for header in comp_headers]
        + [f"Driver Report · {header}" for header in driver_headers]
        + [f"Shift Report · {header}" for header in shift_headers]
    )

    output_rows = []

    for index, comp in enumerate(comp_rows):
        match = matches[index]
        shift = match["shift_row"] or {}
        driver = driver_by_key.get(_encrypted_key(comp), {})

        shift_hours_raw = _get_first(
            shift,
            [
                "Total Driver Hours",
                "Hours Worked",
                "Worked Hours",
                "Total Worked",
            ],
        )

        row = {
            "Nume real estimat": match["name"],
            "Confidence": match["confidence"],
            "Diferență km": match["km_diff"],
            "Diferență ore (minute)": match["hours_diff"],
            "Candidat alternativ 2": match["candidate_2"],
            "Candidat alternativ 3": match["candidate_3"],
            "Encrypted ID": _display_encrypted(comp),
            "Cod șofer / Transporter ID": _driver_code(shift),
            "Vehicle Identifier": _vehicle_identifier(shift),
            "FICO Recent": _fico_recent(comp, driver),
            "FICO Last Period": _fico_last(comp),
            "FICO Δ": _fico_delta(comp),
            "Ore lucrate": shift_hours_raw,
            "Km Shift Report": _shift_km(shift),
            "Incidente detectate": _incident_summary(comp, driver),
        }

        for header in comp_headers:
            row[f"Comparison · {header}"] = comp.get(header)

        for header in driver_headers:
            row[f"Driver Report · {header}"] = driver.get(header)

        for header in shift_headers:
            row[f"Shift Report · {header}"] = shift.get(header)

        output_rows.append(row)

    return all_headers, output_rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def _write_original_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)

    for row in rows:
        ws.append([row.get(header) for header in headers])

    dark = PatternFill("solid", fgColor="17212B")
    white = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = dark
        cell.font = white
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for index, header in enumerate(headers, start=1):
        width = min(max(12, len(str(header)) + 2), 38)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = width


def _write_excel(
    headers,
    rows,
    comp_headers,
    comp_rows,
    driver_headers,
    driver_rows,
    shift_headers,
    shift_rows,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Verificare Scor"

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    dark = PatternFill("solid", fgColor="17212B")
    white_bold = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="E9F8EF")
    medium_fill = PatternFill("solid", fgColor="FFF4D6")
    low_fill = PatternFill("solid", fgColor="FDEEEE")
    incident_fill = PatternFill("solid", fgColor="FFE0E0")
    low_fico_fill = PatternFill("solid", fgColor="F4CCCC")

    for cell in ws[1]:
        cell.fill = dark
        cell.font = white_bold
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_index = {
        header: index + 1
        for index, header in enumerate(headers)
    }

    confidence_col = header_index.get("Confidence")
    fico_recent_col = header_index.get("FICO Recent")
    incident_summary_col = header_index.get("Incidente detectate")

    for row_number in range(2, ws.max_row + 1):
        if confidence_col:
            confidence = ws.cell(row_number, confidence_col).value
            cell = ws.cell(row_number, confidence_col)
            if confidence == "High":
                cell.fill = high_fill
            elif confidence == "Medium":
                cell.fill = medium_fill
            else:
                cell.fill = low_fill
            cell.font = Font(bold=True)

        if fico_recent_col:
            fico_value = _num(
                ws.cell(row_number, fico_recent_col).value
            )
            if fico_value is not None and fico_value < 800:
                cell = ws.cell(row_number, fico_recent_col)
                cell.fill = low_fico_fill
                cell.font = Font(bold=True)

        if incident_summary_col:
            incident_text = str(
                ws.cell(row_number, incident_summary_col).value or ""
            ).strip()
            if incident_text:
                cell = ws.cell(row_number, incident_summary_col)
                cell.fill = incident_fill
                cell.font = Font(bold=True)

    # Highlight every original incident cell with a positive value.
    for header, column_index in header_index.items():
        source_header = header.split(" · ", 1)[-1]

        if not _is_incident_header(source_header):
            continue

        for row_number in range(2, ws.max_row + 1):
            value = _num(
                ws.cell(row_number, column_index).value
            )
            if value is not None and value > 0:
                cell = ws.cell(row_number, column_index)
                cell.fill = incident_fill
                cell.font = Font(bold=True)

    for index, header in enumerate(headers, start=1):
        width = min(max(14, len(str(header)) + 2), 42)

        if header in {
            "Nume real estimat",
            "Encrypted ID",
            "Candidat alternativ 2",
            "Candidat alternativ 3",
            "Incidente detectate",
        }:
            width = 38

        ws.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = width

    # Preserve every original column/value in dedicated sheets too.
    _write_original_sheet(
        wb,
        "Comparison Original",
        comp_headers,
        comp_rows,
    )
    _write_original_sheet(
        wb,
        "Driver Report Original",
        driver_headers,
        driver_rows,
    )
    _write_original_sheet(
        wb,
        "Shift Report Original",
        shift_headers,
        shift_rows,
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Web UI
# ---------------------------------------------------------------------------

def _cleanup_exports():
    now = datetime.now()

    try:
        for name in os.listdir(_EXPORT_DIR):
            path = os.path.join(_EXPORT_DIR, name)

            try:
                age = now - datetime.fromtimestamp(
                    os.path.getmtime(path)
                )
                if age > timedelta(hours=24):
                    os.remove(path)
            except Exception:
                pass
    except Exception:
        pass


def _page(body: str, title="Verificare Scor"):
    return HTMLResponse(
        f"""<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html.escape(title)}</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;font-family:Arial,Helvetica,sans-serif;background:#f4f6f8;color:#17212b}}
.wrap{{width:min(96%,1380px);margin:28px auto 60px}}
.top{{display:flex;justify-content:space-between;align-items:center;gap:15px;margin-bottom:20px}}
h1{{margin:0;font-size:34px}}
.back{{text-decoration:none;color:#17212b;border:1px solid #d8dde3;background:#fff;padding:11px 14px;border-radius:10px;font-weight:800}}
.panel{{background:#fff;border-radius:16px;padding:22px;box-shadow:0 5px 18px rgba(0,0,0,.05);margin-bottom:16px}}
.grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}}
label{{display:block;font-weight:800;margin-bottom:8px}}
input[type=file]{{width:100%;padding:13px;border:1px solid #d8dde3;border-radius:10px;background:#fff}}
button,.download{{display:inline-flex;align-items:center;justify-content:center;text-decoration:none;padding:13px 16px;border:0;border-radius:10px;background:#17212b;color:#fff;font-weight:900;cursor:pointer;margin-top:16px}}
.help{{color:#667085;line-height:1.5}}
.notice{{padding:12px 14px;border-radius:10px;background:#fff4d6;color:#8a5a00;margin:12px 0}}
.table-wrap{{overflow:auto;max-height:68vh;border:1px solid #eceff2;border-radius:12px}}
table{{border-collapse:collapse;width:max-content;min-width:100%;font-size:13px}}
th,td{{padding:10px 11px;border-bottom:1px solid #eceff2;white-space:nowrap;text-align:left}}
th{{position:sticky;top:0;background:#17212b;color:#fff;z-index:1}}
.high{{background:#e9f8ef}}
.medium{{background:#fff4d6}}
.low{{background:#fdeeee}}
.incidents{{color:#b42318;font-weight:800}}
@media(max-width:850px){{
  .grid{{grid-template-columns:1fr}}
  h1{{font-size:28px}}
}}
</style>
</head>
<body>
<main class="wrap">
{body}
</main>
</body>
</html>"""
    )


@router.get("/admin/score-check", response_class=HTMLResponse)
def score_check_page(request: Request):
    body = """
<div class="top">
  <div>
    <div style="font-size:12px;color:#667085;font-weight:800;letter-spacing:1px">
      INSTRUMENTE FICO
    </div>
    <h1>Verificare Scor</h1>
  </div>
  <a class="back" href="/admin">← Admin Dashboard</a>
</div>

<section class="panel">
  <p class="help">
    Încarcă cele trei rapoarte Amazon. Sistemul păstrează toate coloanele,
    corelează ID-ul criptat cu Driver Report și estimează numele real folosind
    km și, atunci când există un câmp compatibil, orele. Raportul afișează
    FICO Recent, FICO Last Period, codurile șoferului și toate incidentele.
  </p>

  <form action="/admin/score-check/generate"
        method="post"
        enctype="multipart/form-data">
    <div class="grid">
      <div>
        <label>1. Comparison</label>
        <input type="file"
               name="comparison_file"
               accept=".xlsx"
               required>
      </div>

      <div>
        <label>2. Driver Report</label>
        <input type="file"
               name="driver_file"
               accept=".xlsx"
               required>
      </div>

      <div>
        <label>3. Shift Report</label>
        <input type="file"
               name="shift_file"
               accept=".xlsx"
               required>
      </div>
    </div>

    <button type="submit">
      Generează raportul complet
    </button>
  </form>
</section>

<section class="panel">
  <strong>Raportul final păstrează:</strong>
  <p class="help">
    toate coloanele originale din Comparison, Driver Report și Shift Report;
    Encrypted ID; nume real estimat; Confidence; FICO Recent; FICO Last Period;
    FICO Δ; km; ore; Transporter/Driver/Vehicle ID; Speeding, Distraction,
    Hard/Harsh Braking, Acceleration, Cornering și orice alt indicator existent.
  </p>
</section>
"""
    return _page(body)


@router.post(
    "/admin/score-check/generate",
    response_class=HTMLResponse,
)
async def score_check_generate(
    comparison_file: UploadFile = File(...),
    driver_file: UploadFile = File(...),
    shift_file: UploadFile = File(...),
):
    uploads = (
        comparison_file,
        driver_file,
        shift_file,
    )

    try:
        for upload in uploads:
            if not (upload.filename or "").lower().endswith(".xlsx"):
                raise ValueError(
                    "Toate cele trei fișiere trebuie să fie XLSX."
                )

        comp_raw = await comparison_file.read()
        driver_raw = await driver_file.read()
        shift_raw = await shift_file.read()

        comp_headers, comp_rows = _read_xlsx(comp_raw)
        driver_headers, driver_rows = _read_xlsx(driver_raw)
        shift_headers, shift_rows = _read_xlsx(shift_raw)

        required_comp = {
            "First Name",
            "Last Name",
        }
        required_driver = {
            "First Name",
            "Last Name",
        }
        required_shift = {
            "First Name",
            "Last Name",
        }

        if not required_comp.issubset(set(comp_headers)):
            raise ValueError(
                "Comparison nu conține First Name / Last Name."
            )

        if not required_driver.issubset(set(driver_headers)):
            raise ValueError(
                "Driver Report nu conține First Name / Last Name."
            )

        if not required_shift.issubset(set(shift_headers)):
            raise ValueError(
                "Shift Report nu conține First Name / Last Name."
            )

        if not any(
            "total driver km" in header.casefold()
            and "this period" in header.casefold()
            for header in comp_headers
        ):
            raise ValueError(
                "Comparison nu conține Total Driver km This Period."
            )

        if not any(
            "total driver km" in header.casefold()
            for header in shift_headers
        ):
            raise ValueError(
                "Shift Report nu conține Total Driver km."
            )

        headers, rows = _build_rows(
            comp_headers,
            comp_rows,
            driver_headers,
            driver_rows,
            shift_headers,
            shift_rows,
        )

        excel_bytes = _write_excel(
            headers,
            rows,
            comp_headers,
            comp_rows,
            driver_headers,
            driver_rows,
            shift_headers,
            shift_rows,
        )

    except Exception as exc:
        return _page(
            '<div class="notice">'
            "Nu am putut procesa fișierele: "
            f"{html.escape(str(exc))}"
            "</div>"
            '<a class="back" href="/admin/score-check">Înapoi</a>'
        )

    finally:
        for upload in uploads:
            try:
                await upload.close()
            except Exception:
                pass

    _cleanup_exports()

    token = uuid.uuid4().hex
    filename = (
        "Verificare_Scor_"
        + datetime.now().strftime("%Y-%m-%d_%H%M%S")
        + "_"
        + token[:8]
        + ".xlsx"
    )

    path = os.path.join(_EXPORT_DIR, filename)
    with open(path, "wb") as handle:
        handle.write(excel_bytes)

    preview_headers = [
        "Nume real estimat",
        "Confidence",
        "Encrypted ID",
        "Cod șofer / Transporter ID",
        "Vehicle Identifier",
        "FICO Recent",
        "FICO Last Period",
        "Ore lucrate",
        "Km Shift Report",
        "Incidente detectate",
    ]

    preview_rows = ""

    for row in rows[:300]:
        confidence = str(
            row.get("Confidence") or "Low"
        ).lower()

        cells = []

        for header in preview_headers:
            value = row.get(header)
            display = "—" if value in (None, "") else str(value)

            css_class = ""
            if header == "Incidente detectate" and display != "—":
                css_class = ' class="incidents"'

            cells.append(
                f"<td{css_class}>{html.escape(display)}</td>"
            )

        preview_rows += (
            f'<tr class="{html.escape(confidence)}">'
            + "".join(cells)
            + "</tr>"
        )

    table = (
        '<div class="table-wrap"><table><thead><tr>'
        + "".join(
            f"<th>{html.escape(header)}</th>"
            for header in preview_headers
        )
        + "</tr></thead><tbody>"
        + preview_rows
        + "</tbody></table></div>"
    )

    body = f"""
<div class="top">
  <div>
    <div style="font-size:12px;color:#667085;font-weight:800;letter-spacing:1px">
      INSTRUMENTE FICO
    </div>
    <h1>Verificare Scor</h1>
  </div>
  <a class="back" href="/admin">← Admin Dashboard</a>
</div>

<section class="panel">
  <strong>Raport generat: {len(rows)} șoferi</strong>
  <p class="help">
    Preview-ul arată datele principale. Excelul descărcat conține foaia
    „Verificare Scor” cu toate coloanele unite și încă trei foi cu datele
    originale complete: Comparison, Driver Report și Shift Report.
  </p>
  <a class="download"
     href="/admin/score-check/download/{html.escape(filename)}">
    Descarcă Excel complet
  </a>
</section>

<section class="panel">
  {table}
</section>
"""

    return _page(body)


@router.get("/admin/score-check/download/{filename}")
def score_check_download(filename: str):
    safe = os.path.basename(filename)
    path = os.path.join(_EXPORT_DIR, safe)

    if not os.path.isfile(path):
        return HTMLResponse(
            "Fișierul nu mai este disponibil.",
            status_code=404,
        )

    return FileResponse(
        path,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        filename=safe,
    )


def register_score_check(app):
    if getattr(
        app.state,
        "_score_check_registered",
        False,
    ):
        return

    app.state._score_check_registered = True
    app.include_router(router)
