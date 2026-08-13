from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse, Response
from pydantic import BaseModel
from datetime import date, datetime, timedelta, timezone, time
import sqlite3
import io
import csv
import html
import openpyxl
import os
import uuid
import unicodedata
from difflib import SequenceMatcher
import json
import urllib.request
import urllib.parse
import re
import base64
import secrets
import hmac
import hashlib
import threading
import calendar as pycalendar
import zipfile
from zoneinfo import ZoneInfo

DB = "fico.db"
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
DB_BACKEND = "postgresql" if DATABASE_URL else "sqlite"

R2_ACCESS_KEY_ID = os.getenv("R2_ACCESS_KEY_ID", "").strip()
R2_SECRET_ACCESS_KEY = os.getenv("R2_SECRET_ACCESS_KEY", "").strip()
R2_ENDPOINT = os.getenv("R2_ENDPOINT", "").strip().rstrip("/")
R2_BUCKET_NAME = os.getenv("R2_BUCKET_NAME", "").strip()

MR_LOGISTICS_LOGO = os.path.join(
    os.path.dirname(__file__),
    "ChatGPT Image 8. Aug. 2026, 15_20_51.png"
)
LOGIN_BACKGROUND = os.path.join(
    os.path.dirname(__file__),
    "amazon-login-background.webp"
)

R2_ENABLED = all([
    R2_ACCESS_KEY_ID,
    R2_SECRET_ACCESS_KEY,
    R2_ENDPOINT,
    R2_BUCKET_NAME
])

UPLOAD_DIR = "uploads"
MAX_IMAGE_BYTES = 10 * 1024 * 1024
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}

ADMIN_COOKIE_NAME = "fico_admin_session"
ADMIN_SESSION_DAYS = 7
OWNER_VERIFY_MINUTES = 30
OWNER_CODE_MINUTES = 10

ADMIN_INITIAL_PASSWORD = os.getenv("ADMIN_INITIAL_PASSWORD", "").strip()
OWNER_EMAIL = os.getenv("OWNER_EMAIL", "").strip()
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "").strip()
RESEND_FROM_EMAIL = os.getenv(
    "RESEND_FROM_EMAIL",
    "FICO Control <onboarding@resend.dev>"
).strip()

os.makedirs(UPLOAD_DIR, exist_ok=True)


def r2_client():
    if not R2_ENABLED:
        return None

    import boto3
    from botocore.config import Config

    return boto3.client(
        "s3",
        endpoint_url=R2_ENDPOINT,
        aws_access_key_id=R2_ACCESS_KEY_ID,
        aws_secret_access_key=R2_SECRET_ACCESS_KEY,
        region_name="auto",
        config=Config(
            signature_version="s3v4",
            retries={"max_attempts": 3, "mode": "standard"}
        )
    )


def storage_backend_name():
    return "r2" if R2_ENABLED else "local"


app = FastAPI(title="FICO Control")
try:
    from score_check import register_score_check
except ImportError:
    from backend.score_check import register_score_check

register_score_check(app)
HOURS_MAX_UPLOAD_BYTES = 15 * 1024 * 1024
HOURS_MAX_CSV_BYTES = 12 * 1024 * 1024
BERLIN_TZ = ZoneInfo("Europe/Berlin")


def _pg_translate_sql(sql: str) -> str:
    """
    Translate the small amount of SQLite-flavoured SQL used by the existing
    FICO Control code into PostgreSQL-compatible SQL.
    """
    translated = sql

    if re.search(r"\bINSERT\s+OR\s+IGNORE\s+INTO\b", translated, re.IGNORECASE):
        translated = re.sub(
            r"\bINSERT\s+OR\s+IGNORE\s+INTO\b",
            "INSERT INTO",
            translated,
            flags=re.IGNORECASE
        )
        if "ON CONFLICT" not in translated.upper():
            translated = translated.rstrip().rstrip(";") + " ON CONFLICT DO NOTHING"

    # Existing code uses SQLite '?' placeholders. Psycopg uses '%s'.
    translated = translated.replace("?", "%s")
    return translated


class PostgresCompatConnection:
    """
    Thin compatibility wrapper so the existing application code can keep using:
        conn.execute(...).fetchone()
        conn.execute(...).fetchall()
        conn.commit()
        conn.close()

    Rows are returned as dictionaries, matching the application's existing
    row["column"] access pattern.
    """

    def __init__(self):
        import psycopg
        from psycopg.rows import dict_row

        self._conn = psycopg.connect(
            DATABASE_URL,
            row_factory=dict_row,
            connect_timeout=10
        )

    def execute(self, sql, params=()):
        return self._conn.execute(_pg_translate_sql(sql), params or ())

    def executescript(self, script: str):
        # Schema scripts used here contain simple semicolon-delimited statements.
        for statement in script.split(";"):
            statement = statement.strip()
            if statement:
                self.execute(statement)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()


def db():
    if DB_BACKEND == "postgresql":
        return PostgresCompatConnection()

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_column(conn, table, column, definition):
    if DB_BACKEND == "postgresql":
        row = conn.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema='public'
              AND table_name=?
              AND column_name=?
            """,
            (table, column)
        ).fetchone()

        if not row:
            safe_table = re.sub(r"[^a-zA-Z0-9_]", "", table)
            safe_column = re.sub(r"[^a-zA-Z0-9_]", "", column)
            conn.execute(
                f"ALTER TABLE {safe_table} ADD COLUMN {safe_column} {definition}"
            )
        return

    columns = {
        row["name"]
        for row in conn.execute(f"PRAGMA table_info({table})").fetchall()
    }
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init_db():
    conn = db()

    if DB_BACKEND == "postgresql":
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS drivers(
            id BIGSERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            active INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS daily_required(
            id BIGSERIAL PRIMARY KEY,
            work_date TEXT NOT NULL,
            driver_id BIGINT NOT NULL,
            UNIQUE(work_date, driver_id),
            FOREIGN KEY(driver_id) REFERENCES drivers(id)
        );

        CREATE TABLE IF NOT EXISTS submissions(
            id BIGSERIAL PRIMARY KEY,
            work_date TEXT NOT NULL,
            driver_id BIGINT NOT NULL,
            fico_score INTEGER NOT NULL,
            submitted_at TEXT NOT NULL,
            UNIQUE(work_date, driver_id),
            FOREIGN KEY(driver_id) REFERENCES drivers(id)
        );

        CREATE TABLE IF NOT EXISTS unresolved_submissions(
            id BIGSERIAL PRIMARY KEY,
            work_date TEXT NOT NULL,
            entered_full_name TEXT NOT NULL,
            fico_score INTEGER NOT NULL,
            submitted_at TEXT NOT NULL,
            proof_filename TEXT,
            proof_original_name TEXT,
            detected_fico_score INTEGER,
            verification_status TEXT,
            best_match_name TEXT,
            best_match_score DOUBLE PRECISION,
            match_reason TEXT
        );

        CREATE TABLE IF NOT EXISTS admin_settings(
            setting_key TEXT PRIMARY KEY,
            setting_value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS admin_sessions(
            session_token TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            last_seen TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            revoked INTEGER DEFAULT 0,
            owner_verified_until TEXT
        );

        CREATE TABLE IF NOT EXISTS blocked_admin_names(
            normalized_name TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            blocked_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS owner_email_codes(
            id BIGSERIAL PRIMARY KEY,
            code_hash TEXT NOT NULL,
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            used INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS daily_list_imports(
            work_date TEXT PRIMARY KEY,
            imported_at TEXT NOT NULL,
            source_filename TEXT,
            driver_count INTEGER NOT NULL DEFAULT 0
        );


        CREATE TABLE IF NOT EXISTS mentor_required(
            work_date TEXT NOT NULL,
            driver_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            source_filename TEXT,
            PRIMARY KEY(work_date, normalized_name)
        );

        CREATE TABLE IF NOT EXISTS mentor_connected(
            work_date TEXT NOT NULL,
            driver_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            source_filename TEXT,
            PRIMARY KEY(work_date, normalized_name)
        );

        CREATE INDEX IF NOT EXISTS idx_daily_required_work_date
            ON daily_required(work_date);

        CREATE INDEX IF NOT EXISTS idx_submissions_work_date
            ON submissions(work_date);

        CREATE INDEX IF NOT EXISTS idx_unresolved_work_date
            ON unresolved_submissions(work_date);

        CREATE INDEX IF NOT EXISTS idx_admin_sessions_expires
            ON admin_sessions(expires_at);
        """)
    else:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS drivers(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            active INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS daily_required(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_date TEXT NOT NULL,
            driver_id INTEGER NOT NULL,
            UNIQUE(work_date, driver_id),
            FOREIGN KEY(driver_id) REFERENCES drivers(id)
        );

        CREATE TABLE IF NOT EXISTS submissions(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_date TEXT NOT NULL,
            driver_id INTEGER NOT NULL,
            fico_score INTEGER NOT NULL,
            submitted_at TEXT NOT NULL,
            UNIQUE(work_date, driver_id),
            FOREIGN KEY(driver_id) REFERENCES drivers(id)
        );

        CREATE TABLE IF NOT EXISTS unresolved_submissions(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_date TEXT NOT NULL,
            entered_full_name TEXT NOT NULL,
            fico_score INTEGER NOT NULL,
            submitted_at TEXT NOT NULL,
            proof_filename TEXT,
            proof_original_name TEXT,
            detected_fico_score INTEGER,
            verification_status TEXT,
            best_match_name TEXT,
            best_match_score REAL,
            match_reason TEXT
        );

        CREATE TABLE IF NOT EXISTS admin_settings(
            setting_key TEXT PRIMARY KEY,
            setting_value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS admin_sessions(
            session_token TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            last_seen TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            revoked INTEGER DEFAULT 0,
            owner_verified_until TEXT
        );

        CREATE TABLE IF NOT EXISTS blocked_admin_names(
            normalized_name TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            blocked_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS owner_email_codes(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code_hash TEXT NOT NULL,
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            used INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS daily_list_imports(
            work_date TEXT PRIMARY KEY,
            imported_at TEXT NOT NULL,
            source_filename TEXT,
            driver_count INTEGER NOT NULL DEFAULT 0
        );


        CREATE TABLE IF NOT EXISTS mentor_required(
            work_date TEXT NOT NULL,
            driver_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            source_filename TEXT,
            PRIMARY KEY(work_date, normalized_name)
        );

        CREATE TABLE IF NOT EXISTS mentor_connected(
            work_date TEXT NOT NULL,
            driver_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            source_filename TEXT,
            PRIMARY KEY(work_date, normalized_name)
        );
        """)

    ensure_column(conn, "submissions", "entered_full_name", "TEXT")
    ensure_column(conn, "submissions", "proof_filename", "TEXT")
    ensure_column(conn, "submissions", "proof_original_name", "TEXT")
    ensure_column(conn, "submissions", "detected_fico_score", "INTEGER")
    ensure_column(conn, "submissions", "verification_status", "TEXT")
    ensure_column(conn, "mentor_connected", "first_connection_time", "TEXT")
    ensure_column(
        conn,
        "submissions",
        "name_match_score",
        "DOUBLE PRECISION" if DB_BACKEND == "postgresql" else "REAL"
    )

    existing_password = conn.execute(
        "SELECT setting_value FROM admin_settings WHERE setting_key='shared_password_hash'"
    ).fetchone()

    if not existing_password and ADMIN_INITIAL_PASSWORD:
        salt = secrets.token_bytes(16)
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            ADMIN_INITIAL_PASSWORD.encode("utf-8"),
            salt,
            250000
        )
        encoded = (
            "pbkdf2_sha256$250000$"
            + base64.b64encode(salt).decode("ascii")
            + "$"
            + base64.b64encode(digest).decode("ascii")
        )
        conn.execute(
            "INSERT INTO admin_settings(setting_key, setting_value) VALUES(?,?)",
            ("shared_password_hash", encoded)
        )

    conn.commit()
    conn.close()


def db_healthcheck():
    conn = db()
    try:
        row = conn.execute("SELECT 1 AS ok").fetchone()
        return bool(row and row["ok"] == 1)
    finally:
        conn.close()


init_db()


@app.get("/brand-logo")
def brand_logo():
    if not os.path.isfile(MR_LOGISTICS_LOGO):
        raise HTTPException(status_code=404, detail="brand_logo_not_found")
    return FileResponse(
        MR_LOGISTICS_LOGO,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=3600"}
    )


@app.get("/login-background")
def login_background():
    if not os.path.isfile(LOGIN_BACKGROUND):
        raise HTTPException(status_code=404, detail="login_background_not_found")
    return FileResponse(
        LOGIN_BACKGROUND,
        media_type="image/webp",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"}
    )


@app.get("/health/db")
def health_db():
    try:
        ok = db_healthcheck()
        return {
            "ok": bool(ok),
            "database": DB_BACKEND
        }
    except Exception as exc:
        return {
            "ok": False,
            "database": DB_BACKEND,
            "error": type(exc).__name__
        }


@app.get("/health/daily-list")
def health_daily_list(d: str | None = None):
    selected = d or date.today().isoformat()
    conn = db()
    try:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM daily_required WHERE work_date=?",
            (selected,)
        ).fetchone()
        meta = conn.execute(
            """
            SELECT imported_at, source_filename, driver_count
            FROM daily_list_imports
            WHERE work_date=?
            """,
            (selected,)
        ).fetchone()

        return {
            "ok": True,
            "date": selected,
            "driver_count": int(row["c"] if row else 0),
            "saved": bool(row and int(row["c"]) > 0),
            "imported_at": meta["imported_at"] if meta else None
        }
    finally:
        conn.close()


@app.get("/health/storage")
def health_storage():
    if not R2_ENABLED:
        return {
            "ok": False,
            "storage": "local",
            "error": "r2_not_configured"
        }

    try:
        client = r2_client()
        client.head_bucket(Bucket=R2_BUCKET_NAME)
        return {
            "ok": True,
            "storage": "r2",
            "bucket": R2_BUCKET_NAME
        }
    except Exception as exc:
        print(
            "R2_HEALTH_ERROR:",
            type(exc).__name__,
            str(exc)[:500],
            flush=True
        )
        return {
            "ok": False,
            "storage": "r2",
            "bucket": R2_BUCKET_NAME,
            "error": type(exc).__name__
        }


def utc_now():
    return datetime.now(timezone.utc)


def iso_utc(dt):
    return dt.astimezone(timezone.utc).isoformat(timespec="seconds")


def parse_iso_utc(value):
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except Exception:
        return None


def normalize_admin_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = " ".join(value.strip().split())
    return value.casefold()


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    iterations = 250000
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        iterations
    )
    return (
        f"pbkdf2_sha256${iterations}$"
        f"{base64.b64encode(salt).decode('ascii')}$"
        f"{base64.b64encode(digest).decode('ascii')}"
    )


def verify_password(password: str, encoded: str) -> bool:
    try:
        algorithm, iterations_s, salt_b64, digest_b64 = encoded.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False

        iterations = int(iterations_s)
        salt = base64.b64decode(salt_b64)
        expected = base64.b64decode(digest_b64)

        actual = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt,
            iterations
        )
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False


def get_shared_password_hash(conn):
    row = conn.execute(
        "SELECT setting_value FROM admin_settings WHERE setting_key='shared_password_hash'"
    ).fetchone()
    return row["setting_value"] if row else None


def create_admin_session(conn, display_name: str) -> str:
    token = secrets.token_urlsafe(36)
    now = utc_now()
    expires = now + timedelta(days=ADMIN_SESSION_DAYS)

    conn.execute("""
        INSERT INTO admin_sessions(
            session_token,
            display_name,
            normalized_name,
            created_at,
            last_seen,
            expires_at,
            revoked
        ) VALUES(?,?,?,?,?,?,0)
    """, (
        token,
        display_name,
        normalize_admin_name(display_name),
        iso_utc(now),
        iso_utc(now),
        iso_utc(expires)
    ))
    conn.commit()
    return token


def get_valid_admin_session(token: str | None):
    if not token:
        return None

    conn = db()
    row = conn.execute("""
        SELECT session_token,
               display_name,
               normalized_name,
               created_at,
               last_seen,
               expires_at,
               revoked,
               owner_verified_until
        FROM admin_sessions
        WHERE session_token=?
    """, (token,)).fetchone()

    if not row:
        conn.close()
        return None

    blocked = conn.execute(
        "SELECT 1 FROM blocked_admin_names WHERE normalized_name=?",
        (row["normalized_name"],)
    ).fetchone()

    expires = parse_iso_utc(row["expires_at"])
    invalid = (
        bool(row["revoked"])
        or bool(blocked)
        or not expires
        or expires <= utc_now()
    )

    if invalid:
        conn.execute(
            "UPDATE admin_sessions SET revoked=1 WHERE session_token=?",
            (token,)
        )
        conn.commit()
        conn.close()
        return None

    conn.execute(
        "UPDATE admin_sessions SET last_seen=? WHERE session_token=?",
        (iso_utc(utc_now()), token)
    )
    conn.commit()
    conn.close()
    return dict(row)


def is_owner_verified(session):
    if not session:
        return False
    until = parse_iso_utc(session.get("owner_verified_until"))
    return bool(until and until > utc_now())


def mask_email(email: str) -> str:
    if not email or "@" not in email:
        return "neconfigurat"
    local, domain = email.split("@", 1)
    if len(local) <= 2:
        masked_local = local[:1] + "*"
    else:
        masked_local = local[:2] + "*" * max(2, len(local) - 2)
    return masked_local + "@" + domain


def hash_owner_code(code: str) -> str:
    return hashlib.sha256(code.encode("utf-8")).hexdigest()


def send_owner_verification_code(code: str):
    if not OWNER_EMAIL:
        raise RuntimeError("OWNER_EMAIL is not configured")
    if not RESEND_API_KEY:
        raise RuntimeError("RESEND_API_KEY is not configured")

    payload = json.dumps({
        "from": RESEND_FROM_EMAIL,
        "to": [OWNER_EMAIL],
        "subject": "FICO Control - Owner verification code",
        "html": (
            "<div style='font-family:Arial,sans-serif'>"
            "<h2>FICO Control</h2>"
            "<p>Codul pentru accesul Owner este:</p>"
            f"<div style='font-size:32px;font-weight:800;letter-spacing:5px'>{html.escape(code)}</div>"
            f"<p>Codul expiră în {OWNER_CODE_MINUTES} minute.</p>"
            "<p>Dacă nu ai cerut acest cod, îl poți ignora.</p>"
            "</div>"
        )
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=payload,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "FICO-Control/1.0"
        },
        method="POST"
    )

    with urllib.request.urlopen(req, timeout=20) as response:
        response.read()


def owner_gate_html(message: str = "", error: bool = False):
    message_html = ""
    if message:
        cls = "error" if error else "ok"
        message_html = f'<div class="{cls}">{html.escape(message)}</div>'

    return f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FICO Control Owner</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#f4f6f8;font-family:Arial,sans-serif;color:#17212b}}
.wrap{{width:min(92%,520px);margin:55px auto}}
.card{{background:#fff;border-radius:20px;padding:28px;box-shadow:0 10px 30px rgba(0,0,0,.08)}}
.brand{{font-size:13px;font-weight:900;letter-spacing:2px}}
h1{{margin:18px 0 8px;font-size:30px}}
p{{color:#667085;line-height:1.5}}
input,button{{width:100%;padding:14px;border-radius:11px;font-size:16px;margin-top:10px}}
input{{border:1px solid #d8dde3}}
button{{border:0;background:#17212b;color:#fff;font-weight:800;cursor:pointer}}
.secondary{{background:#fff;color:#17212b;border:1px solid #d8dde3}}
.ok,.error{{padding:12px;border-radius:10px;margin:15px 0;font-weight:700}}
.ok{{background:#e9f8ef;color:#14804a}}
.error{{background:#fdeeee;color:#b42318}}
</style>
</head>
<body>
<div class="wrap">
<div class="card">
<div class="brand">FICO CONTROL · OWNER</div>
<h1>Confirmare prin email</h1>
<p>Controlul Owner este protejat separat. Codul este trimis numai la <strong>{html.escape(mask_email(OWNER_EMAIL))}</strong>.</p>
{message_html}
<form method="post" action="/admin/owner/send-code">
<button type="submit">Trimite codul pe email</button>
</form>
<form method="post" action="/admin/owner/verify-code">
<input name="code" inputmode="numeric" maxlength="6" placeholder="Cod din 6 cifre" required>
<button type="submit">Confirmă codul</button>
</form>
<a href="/admin" style="display:block;text-align:center;margin-top:16px;color:#17212b;font-weight:700">Înapoi la Admin</a>
</div>
</div>
</body>
</html>
"""


SITE_LANGUAGE_SCRIPT = r'''
<script>
(() => {
  const translations = {
    de: {
      "Admin Dashboard":"Admin-Dashboard","Conectat ca:":"Angemeldet als:","Bază date:":"Datenbank:",
      "Programați":"Eingeplant","Au trimis":"Gesendet","Lipsesc":"Fehlen","FICO sub 800":"FICO unter 800","Necesită verificare":"Prüfung erforderlich",
      "Istoric FICO":"FICO-Verlauf","Astăzi":"Heute","Zi cu listă salvată":"Tag mit gespeicherter Liste","Zi afișată:":"Angezeigter Tag:",
      "Afișează":"Anzeigen","Reset":"Zurücksetzen","Lista zilnică din Cortex":"Tagesliste aus Cortex","Încarcă Excel Cortex":"Cortex-Excel hochladen",
      "Aplicația extrage automat șoferii din coloana „Name des Fahrers”.":"Die Anwendung liest die Fahrer automatisch aus der Spalte „Name des Fahrers“ aus.",
      "Șofer":"Fahrer","Status":"Status","FICO introdus":"Eingegebener FICO","FICO detectat":"Erkannter FICO","Verificare":"Prüfung","Ora":"Uhrzeit","Dovadă":"Nachweis",
      "Mentor Check":"Mentor-Prüfung","Control ore":"Arbeitszeitkontrolle","POD & CCC":"POD & CCC","Concesii":"Konzessionen","A doua reîncercare":"Zweiter Zustellversuch",
      "Owner":"Inhaber","Ieșire":"Abmelden","Export Excel":"Excel exportieren","Copiază șoferii lipsă":"Fehlende Fahrer kopieren","Copiat":"Kopiert",
      "INSTRUMENTE FICO":"FICO-WERKZEUGE","Control ore șoferi":"Fahrer-Arbeitszeiten","Data verificării":"Prüfdatum","Caută șofer":"Fahrer suchen",
      "Șoferi verificați":"Geprüfte Fahrer","Peste 10 ore":"Über 10 Stunden","10h 30m sau mai mult":"10 Std. 30 Min. oder mehr","cu ore reale":"mit echten Arbeitszeiten",
      "necesită atenție":"Aufmerksamkeit erforderlich","limită critică":"kritische Grenze","Ore lucrate":"Arbeitsstunden","Interval real":"Tatsächlicher Zeitraum",
      "Blocuri":"Blöcke","Total lucrat":"Gesamtarbeitszeit","Încarcă raportul săptămânal Amazon":"Wöchentlichen Amazon-Bericht hochladen","Încarcă Anfahrlisten Amazon":"Amazon-Anfahrlisten hochladen","Șoferii sunt grupați după Transporter-ID. Toate rutele sunt adunate, iar intervalele suprapuse nu sunt numărate de două ori.":"Die Fahrer werden nach Transporter-ID gruppiert. Alle Routen werden addiert, überlappende Zeiträume jedoch nur einmal gezählt.","Anfahrlisten Amazon XLSX":"Amazon-Anfahrlisten XLSX","Verifică orele":"Arbeitszeiten prüfen",
      "Descarcă CSV":"CSV herunterladen","Încarcă raportul pentru a vedea orele lucrate.":"Lade den Bericht hoch, um die Arbeitszeiten zu sehen.",
      "Încarcă cele trei fișiere Amazon":"Die drei Amazon-Dateien hochladen","Generează rapoartele":"Berichte erstellen","Descarcă Excel POD":"POD-Excel herunterladen","Descarcă Excel CCC":"CCC-Excel herunterladen",
      "Înlocuire automată Transporter ID cu numele real și rapoarte profesionale":"Transporter-ID automatisch durch den echten Namen ersetzen und professionelle Berichte erstellen",
      "Planul săptămânal furnizează numele reale. POD și CCC sunt procesate separat.":"Der Wochenplan liefert die echten Namen. POD und CCC werden getrennt verarbeitet.",
      "1. Plan săptămânal ID–nume":"1. Wochenplan ID–Name","2. Raport POD":"2. POD-Bericht","3. Raport CCC":"3. CCC-Bericht","Raport":"Bericht","Cazuri":"Fälle","Atenție":"Achtung",
      "Încarcă raportul Amazon Concessions":"Amazon-Concessions-Bericht hochladen","Generează Excel":"Excel erstellen","Descarcă Excel Concesii":"Concessions-Excel herunterladen",
      "Raport săptămânal DNR ordonat și formatat automat":"Wöchentlicher DNR-Bericht, automatisch sortiert und formatiert",
      "Șoferii cu cele mai multe concesii apar primii. Raportul final respectă modelul KW.":"Fahrer mit den meisten Konzessionen stehen zuerst. Der Abschlussbericht folgt dem KW-Modell.",
      "Trebuie conectați":"Müssen verbunden sein","Persoane conectate":"Verbundene Personen","Nu s-au conectat":"Nicht verbunden","Verifică Mentor":"Mentor prüfen",
      "1. Cortex":"1. Cortex","2. Mentor Shift Report":"2. Mentor-Schichtbericht","Compară lista Cortex cu Mentor Shift Report":"Cortex-Liste mit dem Mentor-Schichtbericht vergleichen",
      "ȘOFER CORTEX":"CORTEX-FAHRER","STATUS MENTOR":"MENTOR-STATUS","PRIMA CONECTARE":"ERSTE VERBINDUNG","POTRIVIRE MENTOR":"MENTOR-ZUORDNUNG",
      "Control acces Admin":"Admin-Zugriffskontrolle","Sesiuni":"Sitzungen","Nume blocate":"Gesperrte Namen","Ultima activitate":"Letzte Aktivität","Acțiune":"Aktion",
      "Blochează":"Sperren","Deblochează":"Entsperren","Deconectează":"Abmelden","Nicio sesiune activă.":"Keine aktive Sitzung.","Niciun nume blocat.":"Keine gesperrten Namen.",
      "Confirmare prin email":"Bestätigung per E-Mail","Trimite codul pe email":"Code per E-Mail senden","Confirmă codul":"Code bestätigen","Înapoi la Admin":"Zurück zum Admin",
      "Schimbă parola":"Passwort ändern","Schimbă parola comună":"Gemeinsames Passwort ändern","Setează parola nouă":"Neues Passwort festlegen",
      "Admin Login":"Admin-Anmeldung","Introdu numele tău și parola comună pentru a intra în Admin Dashboard.":"Gib deinen Namen und das gemeinsame Passwort ein, um das Admin-Dashboard zu öffnen.",
      "Numele tău":"Dein Name","Parola":"Passwort","Ai uitat parola?":"Passwort vergessen?","Intră în Admin":"Admin öffnen","Înapoi la login":"Zurück zur Anmeldung"
    },
    en: {
      "Admin Dashboard":"Admin Dashboard","Conectat ca:":"Signed in as:","Bază date:":"Database:",
      "Programați":"Scheduled","Au trimis":"Submitted","Lipsesc":"Missing","FICO sub 800":"FICO below 800","Necesită verificare":"Needs review",
      "Istoric FICO":"FICO history","Astăzi":"Today","Zi cu listă salvată":"Day with saved list","Zi afișată:":"Displayed day:",
      "Afișează":"Show","Reset":"Reset","Lista zilnică din Cortex":"Daily Cortex list","Încarcă Excel Cortex":"Upload Cortex Excel",
      "Aplicația extrage automat șoferii din coloana „Name des Fahrers”.":"The app automatically extracts drivers from the “Name des Fahrers” column.",
      "Șofer":"Driver","Status":"Status","FICO introdus":"Entered FICO","FICO detectat":"Detected FICO","Verificare":"Verification","Ora":"Time","Dovadă":"Proof",
      "Mentor Check":"Mentor Check","Control ore":"Hours control","POD & CCC":"POD & CCC","Concesii":"Concessions","A doua reîncercare":"Second reattempt",
      "Owner":"Owner","Ieșire":"Log out","Export Excel":"Export Excel","Copiază șoferii lipsă":"Copy missing drivers","Copiat":"Copied",
      "INSTRUMENTE FICO":"FICO TOOLS","Control ore șoferi":"Driver hours control","Data verificării":"Check date","Caută șofer":"Search driver",
      "Șoferi verificați":"Checked drivers","Peste 10 ore":"Over 10 hours","10h 30m sau mai mult":"10h 30m or more","cu ore reale":"with actual hours",
      "necesită atenție":"needs attention","limită critică":"critical limit","Ore lucrate":"Hours worked","Interval real":"Actual interval","Blocuri":"Blocks","Total lucrat":"Total worked",
      "Încarcă raportul săptămânal Amazon":"Upload the weekly Amazon report","Încarcă Anfahrlisten Amazon":"Upload Amazon Anfahrlisten","Șoferii sunt grupați după Transporter-ID. Toate rutele sunt adunate, iar intervalele suprapuse nu sunt numărate de două ori.":"Drivers are grouped by Transporter ID. All routes are added, while overlapping intervals are counted only once.","Anfahrlisten Amazon XLSX":"Amazon Anfahrlisten XLSX","Verifică orele":"Check hours","Descarcă CSV":"Download CSV","Încarcă raportul pentru a vedea orele lucrate.":"Upload the report to view worked hours.",
      "Încarcă cele trei fișiere Amazon":"Upload the three Amazon files","Generează rapoartele":"Generate reports","Descarcă Excel POD":"Download POD Excel","Descarcă Excel CCC":"Download CCC Excel",
      "Înlocuire automată Transporter ID cu numele real și rapoarte profesionale":"Automatically replace Transporter ID with the real name and create professional reports",
      "Planul săptămânal furnizează numele reale. POD și CCC sunt procesate separat.":"The weekly plan provides real names. POD and CCC are processed separately.",
      "1. Plan săptămânal ID–nume":"1. Weekly ID–name plan","2. Raport POD":"2. POD report","3. Raport CCC":"3. CCC report","Raport":"Report","Cazuri":"Cases","Atenție":"Attention",
      "Încarcă raportul Amazon Concessions":"Upload Amazon Concessions report","Generează Excel":"Generate Excel","Descarcă Excel Concesii":"Download Concessions Excel",
      "Raport săptămânal DNR ordonat și formatat automat":"Weekly DNR report, automatically sorted and formatted","Șoferii cu cele mai multe concesii apar primii. Raportul final respectă modelul KW.":"Drivers with the most concessions appear first. The final report follows the KW model.",
      "Trebuie conectați":"Must be connected","Persoane conectate":"Connected people","Nu s-au conectat":"Not connected","Verifică Mentor":"Check Mentor",
      "1. Cortex":"1. Cortex","2. Mentor Shift Report":"2. Mentor Shift Report","Compară lista Cortex cu Mentor Shift Report":"Compare the Cortex list with the Mentor Shift Report",
      "ȘOFER CORTEX":"CORTEX DRIVER","STATUS MENTOR":"MENTOR STATUS","PRIMA CONECTARE":"FIRST CONNECTION","POTRIVIRE MENTOR":"MENTOR MATCH",
      "Control acces Admin":"Admin access control","Sesiuni":"Sessions","Nume blocate":"Blocked names","Ultima activitate":"Last activity","Acțiune":"Action","Blochează":"Block","Deblochează":"Unblock","Deconectează":"Disconnect",
      "Nicio sesiune activă.":"No active sessions.","Niciun nume blocat.":"No blocked names.","Confirmare prin email":"Email confirmation","Trimite codul pe email":"Send code by email","Confirmă codul":"Confirm code","Înapoi la Admin":"Back to Admin",
      "Schimbă parola":"Change password","Schimbă parola comună":"Change shared password","Setează parola nouă":"Set new password",
      "Admin Login":"Admin Login","Introdu numele tău și parola comună pentru a intra în Admin Dashboard.":"Enter your name and the shared password to open the Admin Dashboard.",
      "Numele tău":"Your name","Parola":"Password","Ai uitat parola?":"Forgot password?","Intră în Admin":"Open Admin","Înapoi la login":"Back to login"
    }
  };
  const originals = new WeakMap();
  const normalizeLang = value => ['ro','de','en'].includes(value) ? value : 'ro';
  let lang = normalizeLang(localStorage.getItem('fico_lang') || 'ro');
  function translateTextNode(node) {
    if (!originals.has(node)) originals.set(node, node.nodeValue);
    const original = originals.get(node);
    const clean = original.trim();
    if (!clean) return;
    const translated = lang === 'ro' ? clean : (translations[lang][clean] || clean);
    node.nodeValue = original.replace(clean, translated);
  }
  function translate(root=document.body) {
    document.documentElement.lang = lang;
    const walker = document.createTreeWalker(root, NodeFilter.SHOW_TEXT);
    let node; while ((node = walker.nextNode())) {
      if (!node.parentElement || ['SCRIPT','STYLE','TEXTAREA'].includes(node.parentElement.tagName)) continue;
      translateTextNode(node);
    }
    document.querySelectorAll('input[placeholder]').forEach(el => {
      if (!el.dataset.ficoOriginalPlaceholder) el.dataset.ficoOriginalPlaceholder = el.placeholder;
      const original = el.dataset.ficoOriginalPlaceholder;
      el.placeholder = lang === 'ro' ? original : (translations[lang][original] || original);
    });
    document.querySelectorAll('[data-lang]').forEach(btn => btn.classList.toggle('active', btn.dataset.lang === lang));
    document.querySelectorAll('[data-site-lang]').forEach(btn => btn.classList.toggle('active', btn.dataset.siteLang === lang));
  }
  function selectLanguage(next) {
    lang = normalizeLang(next);
    localStorage.setItem('fico_lang', lang);
    document.cookie = 'fico_lang=' + lang + '; path=/; max-age=31536000; SameSite=Lax';
    translate();
  }
  document.querySelectorAll('[data-lang]').forEach(btn => btn.addEventListener('click', () => selectLanguage(btn.dataset.lang)));
  if (!document.querySelector('.langs')) {
    const picker = document.createElement('div');
    picker.className = 'fico-language-switcher';
    picker.innerHTML = '<button data-site-lang="ro">RO</button><button data-site-lang="de">DE</button><button data-site-lang="en">EN</button>';
    document.body.appendChild(picker);
    picker.querySelectorAll('button').forEach(btn => btn.addEventListener('click', () => selectLanguage(btn.dataset.siteLang)));
  }
  const style = document.createElement('style');
  style.textContent = '.fico-language-switcher{position:fixed;right:16px;bottom:16px;z-index:9999;display:flex;gap:4px;padding:5px;background:rgba(255,255,255,.94);border:1px solid #d8dde3;border-radius:10px;box-shadow:0 5px 18px rgba(0,0,0,.12)}.fico-language-switcher button{width:auto!important;margin:0!important;padding:7px 9px!important;border:0;border-radius:7px;background:#f2f4f7;color:#667085;font-size:12px;font-weight:800;cursor:pointer}.fico-language-switcher button.active{background:#17212b;color:#fff}';
  document.head.appendChild(style);
  translate();
  new MutationObserver(records => records.forEach(record => record.addedNodes.forEach(node => {
    if (node.nodeType === Node.TEXT_NODE) translateTextNode(node);
    else if (node.nodeType === Node.ELEMENT_NODE) translate(node);
  }))).observe(document.body, {childList:true,subtree:true});
})();
</script>
'''

SITE_DARK_BACKGROUND_STYLE = r'''
<style id="fico-dark-admin-background">
html{background:#07111d}
body{
  min-height:100vh;
  background:
    linear-gradient(rgba(5,14,25,.88),rgba(5,14,25,.93)),
    url('/login-background?v=mr-logistics-fleet-plates-4') center center/cover fixed no-repeat !important;
}
.admin>.topbar h1,.admin>.topbar .brand,
.pc-top h1,.pc-top .pc-brand,
.cn-top h1,.cn-top .cn-brand,
main>.topbar h1,main>.topbar .brand,
.wrap>.topbar h1,.wrap>.topbar .brand{
  color:#fff !important;
  text-shadow:0 2px 12px rgba(0,0,0,.35);
}
.admin>.topbar>div>div[style],
.pc-top .pc-sub,.cn-top .cn-sub,
main>.topbar .subtitle,.wrap>.topbar .subtitle{
  color:#cbd5e1 !important;
}
.side-brand{color:#cbd5e1 !important}
.app-shell,.admin,.pc-wrap,.cn-wrap,.wrap{position:relative;z-index:1}
@media(max-width:760px){
  body{background-position:center top !important;background-attachment:scroll !important}
}
</style>
'''

SITE_ENTRY_GUARD_SCRIPT = r'''
<script>
(() => {
  const path = window.location.pathname;
  const params = new URLSearchParams(window.location.search);
  const loginPage = path === '/admin/login' || path === '/admin/login/';
  const publicPage = loginPage || path.startsWith('/admin/forgot-password') ||
    path.startsWith('/admin/reset-password') || path.startsWith('/admin/setup-name');

  if (loginPage && params.get('fresh') === '1') {
    sessionStorage.removeItem('fico_authenticated_this_tab');
  }

  if (path.startsWith('/admin') && !publicPage &&
      sessionStorage.getItem('fico_authenticated_this_tab') !== '1') {
    const destination = path + window.location.search;
    window.location.replace('/admin/login?fresh=1&next=' + encodeURIComponent(destination));
    return;
  }

  const loginForm = document.querySelector('form[action="/admin/login"]');
  if (loginForm) {
    loginForm.addEventListener('submit', () => {
      sessionStorage.setItem('fico_authenticated_this_tab', '1');
    });
  }

  const logoutForm = document.querySelector('form[action="/admin/logout"]');
  if (logoutForm) {
    logoutForm.addEventListener('submit', () => {
      sessionStorage.removeItem('fico_authenticated_this_tab');
    });
  }
})();
</script>
'''


@app.middleware("http")
async def persist_site_language(request: Request, call_next):
    response = await call_next(request)
    content_type = response.headers.get("content-type", "")
    if "text/html" not in content_type.lower():
        return response
    body = b""
    async for chunk in response.body_iterator:
        body += chunk
    page = body.decode("utf-8", errors="replace")
    if "</body>" in page:
        public_light_pages = (
            "/admin/login",
            "/admin/forgot-password",
            "/admin/reset-password",
            "/admin/setup-name"
        )
        dark_style = (
            SITE_DARK_BACKGROUND_STYLE
            if request.url.path.startswith("/admin")
            and not request.url.path.startswith(public_light_pages)
            else ""
        )
        page = page.replace(
            "</body>",
            dark_style + SITE_ENTRY_GUARD_SCRIPT + SITE_LANGUAGE_SCRIPT + "</body>",
            1
        )
    headers = dict(response.headers)
    headers.pop("content-length", None)
    return Response(
        content=page,
        status_code=response.status_code,
        headers=headers,
        media_type="text/html"
    )


@app.middleware("http")
async def protect_admin_routes(request: Request, call_next):
    path = request.url.path

    public_admin_paths = {
        "/admin/login",
        "/admin/login/",
        "/admin/forgot-password",
        "/admin/forgot-password/",
        "/admin/forgot-password/send-code",
        "/admin/reset-password",
        "/admin/reset-password/",
        "/admin/setup-name",
        "/admin/setup-name/"
    }

    if path in public_admin_paths:
        return await call_next(request)

    protected = path.startswith("/admin") or path.startswith("/proof/")

    if not protected:
        return await call_next(request)

    token = request.cookies.get(ADMIN_COOKIE_NAME)
    session = get_valid_admin_session(token)

    if not session:
        next_path = path
        if request.url.query:
            next_path += "?" + request.url.query
        return RedirectResponse(
            "/admin/login?next=" + urllib.parse.quote(next_path, safe=""),
            status_code=303
        )

    request.state.admin_session = session
    return await call_next(request)


@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_page(
    request: Request,
    next: str | None = None,
    error: str | None = None,
    fresh: int | None = None
):
    existing_token = request.cookies.get(ADMIN_COOKIE_NAME)
    current = get_valid_admin_session(existing_token)
    force_login = fresh == 1
    if current and not force_login:
        return RedirectResponse(next or "/admin", status_code=303)

    if force_login and existing_token:
        conn = db()
        conn.execute(
            "UPDATE admin_sessions SET revoked=1 WHERE session_token=?",
            (existing_token,)
        )
        conn.commit()
        conn.close()

    conn = db()
    configured = bool(get_shared_password_hash(conn))
    conn.close()

    error_text = ""
    if error == "invalid":
        error_text = "Numele sau parola nu sunt corecte."
    elif error == "blocked":
        error_text = "Accesul pentru acest nume este blocat."
    elif error == "name":
        error_text = "Introdu numele tău."
    elif not configured:
        error_text = "Parola Admin nu este încă configurată în Render."

    next_value = html.escape(next or "/admin", quote=True)
    error_html = f'<div class="error">{html.escape(error_text)}</div>' if error_text else ""

    page = f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FICO Control Admin Login</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;min-height:100vh;background:#101820 url('/login-background?v=mr-logistics-fleet-plates-4') center center/cover no-repeat fixed;font-family:Arial,sans-serif;color:#17212b;position:relative}}
body::before{{content:"";position:fixed;inset:0;background:rgba(4,12,22,.20);pointer-events:none}}
.wrap{{position:relative;z-index:1;width:min(92%,470px);margin:0 auto;min-height:100vh;display:flex;align-items:center;padding:32px 0}}
.card{{width:100%;background:rgba(255,255,255,.96);border:1px solid rgba(255,255,255,.65);backdrop-filter:blur(8px);border-radius:22px;padding:30px;box-shadow:0 18px 55px rgba(0,0,0,.30)}}
.brand{{font-size:13px;font-weight:900;letter-spacing:2px}}
.login-logo-wrap{{display:flex;justify-content:center;margin:2px 0 18px}}
.login-logo{{display:block;width:min(230px,75%);height:auto;object-fit:contain}}
h1{{font-size:31px;margin:18px 0 8px}}
.subtitle{{color:#667085;line-height:1.5;margin-bottom:22px}}
label{{display:block;font-weight:800;margin:16px 0 7px}}
input,button{{width:100%;padding:14px;border-radius:11px;font-size:16px}}
input{{border:1px solid #d8dde3}}
.password-wrap{{position:relative}}
.password-wrap input{{padding-right:54px}}
.eye-btn{{position:absolute;right:8px;top:50%;transform:translateY(-50%);width:42px;height:42px;margin:0;padding:0;border:0;background:transparent;color:#667085;font-size:20px;cursor:pointer;display:flex;align-items:center;justify-content:center}}
.eye-btn:hover{{background:#f2f4f7}}
button{{margin-top:22px;border:0;background:#17212b;color:#fff;font-weight:800;cursor:pointer}}
.error{{margin:14px 0;padding:12px;border-radius:10px;background:#fdeeee;color:#b42318;font-weight:700}}
.langs{{display:flex;gap:6px;justify-content:flex-end}}
.langs button{{width:auto;margin:0;padding:7px 9px;background:#f2f4f7;color:#667085}}
.langs button.active{{background:#17212b;color:#fff}}
.forgot{{display:block;text-align:right;margin-top:11px;color:#475467;text-decoration:none;font-weight:700;font-size:14px}}
.forgot:hover{{text-decoration:underline}}
@media(max-width:760px){{
 body{{background-position:38% center;background-attachment:scroll}}
 body::before{{background:rgba(4,12,22,.42)}}
 .wrap{{margin:0 auto;padding:22px 0}}
 .card{{padding:24px}}
}}
</style>
</head>
<body>
<div class="wrap">
<div class="card">
<div class="langs">
<button type="button" data-lang="ro" class="active">RO</button>
<button type="button" data-lang="de">DE</button>
<button type="button" data-lang="en">EN</button>
</div>

<div class="login-logo-wrap">
<img class="login-logo" src="/brand-logo" alt="MR Logistics">
</div>

<h1 id="title">Admin Login</h1>
<p class="subtitle" id="subtitle">Introdu numele tău și parola comună pentru a intra în Admin Dashboard.</p>

{error_html}

<form method="post" action="/admin/login" autocomplete="off">
<input type="hidden" name="next_path" value="{next_value}">

<label id="nameLabel">Numele tău</label>
<input name="display_name" autocomplete="off" required>

<label id="passwordLabel">Parola</label>
<div class="password-wrap">
<input id="adminPassword" name="password" type="password" autocomplete="new-password" required>
<button class="eye-btn" id="togglePassword" type="button" aria-label="Arată parola" title="Arată parola">◉</button>
</div>

<a class="forgot" id="forgotLink" href="/admin/forgot-password">Ai uitat parola?</a>
<button type="submit" id="loginButton">Intră în Admin</button>
</form>
</div>
</div>

<script>
const T = {{
 ro: {{
   title:"Admin Login",
   subtitle:"Introdu numele tău și parola comună pentru a intra în Admin Dashboard.",
   name:"Numele tău",
   password:"Parola",
   forgot:"Ai uitat parola?",
   button:"Intră în Admin"
 }},
 de: {{
   title:"Admin Login",
   subtitle:"Gib deinen Namen und das gemeinsame Passwort ein, um das Admin Dashboard zu öffnen.",
   name:"Dein Name",
   password:"Passwort",
   forgot:"Passwort vergessen?",
   button:"Admin öffnen"
 }},
 en: {{
   title:"Admin Login",
   subtitle:"Enter your name and the shared password to open the Admin Dashboard.",
   name:"Your name",
   password:"Password",
   forgot:"Forgot password?",
   button:"Open Admin"
 }}
}};

document.querySelectorAll("[data-lang]").forEach(btn => {{
 btn.addEventListener("click", () => {{
   document.querySelectorAll("[data-lang]").forEach(x => x.classList.remove("active"));
   btn.classList.add("active");
   const t=T[btn.dataset.lang];
   document.getElementById("title").textContent=t.title;
   document.getElementById("subtitle").textContent=t.subtitle;
   document.getElementById("nameLabel").textContent=t.name;
   document.getElementById("passwordLabel").textContent=t.password;
   document.getElementById("forgotLink").textContent=t.forgot;
   document.getElementById("loginButton").textContent=t.button;
 }});
}});

const passwordInput = document.getElementById("adminPassword");
const togglePassword = document.getElementById("togglePassword");

togglePassword.addEventListener("click", () => {{
  const showing = passwordInput.type === "text";
  passwordInput.type = showing ? "password" : "text";
  togglePassword.textContent = showing ? "◉" : "◎";
  togglePassword.setAttribute(
    "aria-label",
    showing ? "Arată parola" : "Ascunde parola"
  );
  togglePassword.setAttribute(
    "title",
    showing ? "Arată parola" : "Ascunde parola"
  );
}});
</script>
</body>
</html>
"""
    response = HTMLResponse(page)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    if force_login:
        response.delete_cookie(ADMIN_COOKIE_NAME)
    return response


@app.post("/admin/login")
def admin_login(
    display_name: str = Form(...),
    password: str = Form(...),
    next_path: str = Form("/admin")
):
    clean_name = " ".join((display_name or "").strip().split())
    if not clean_name:
        return RedirectResponse("/admin/login?error=name", status_code=303)

    normalized = normalize_admin_name(clean_name)
    conn = db()

    blocked = conn.execute(
        "SELECT 1 FROM blocked_admin_names WHERE normalized_name=?",
        (normalized,)
    ).fetchone()

    if blocked:
        conn.close()
        return RedirectResponse("/admin/login?error=blocked", status_code=303)

    password_hash = get_shared_password_hash(conn)

    db_password_ok = bool(
        password_hash and verify_password(password, password_hash)
    )

    env_password_ok = bool(
        ADMIN_INITIAL_PASSWORD
        and hmac.compare_digest(password, ADMIN_INITIAL_PASSWORD)
    )

    if not db_password_ok and not env_password_ok:
        conn.close()
        return RedirectResponse("/admin/login?error=invalid", status_code=303)

    if env_password_ok and not db_password_ok:
        conn.execute("""
            INSERT INTO admin_settings(setting_key, setting_value)
            VALUES('shared_password_hash', ?)
            ON CONFLICT(setting_key) DO UPDATE SET
                setting_value=excluded.setting_value
        """, (hash_password(ADMIN_INITIAL_PASSWORD),))
        conn.commit()

    token = create_admin_session(conn, clean_name)
    conn.close()

    safe_next = next_path if next_path.startswith("/admin") else "/admin"
    response = RedirectResponse(safe_next, status_code=303)

    # Session-only cookie: no Max-Age / Expires.
    # It is not designed to persist as a remembered login.
    response.set_cookie(
        ADMIN_COOKIE_NAME,
        token,
        httponly=True,
        secure=True,
        samesite="lax"
    )

    # Remove any old browser-memory cookie from the previous version.
    response.delete_cookie("fico_admin_display_name")
    response.delete_cookie("fico_admin_pending")

    return response


@app.post("/admin/logout")
def admin_logout(request: Request):
    token = request.cookies.get(ADMIN_COOKIE_NAME)
    if token:
        conn = db()
        conn.execute(
            "UPDATE admin_sessions SET revoked=1 WHERE session_token=?",
            (token,)
        )
        conn.commit()
        conn.close()

    response = RedirectResponse("/admin/login", status_code=303)
    response.delete_cookie(ADMIN_COOKIE_NAME)
    response.delete_cookie("fico_admin_display_name")
    response.delete_cookie("fico_admin_pending")
    response.headers["Cache-Control"] = "no-store"
    return response


@app.get("/admin/forgot-password", response_class=HTMLResponse)
def forgot_password_page(message: str | None = None, error: str | None = None):
    message_html = ""
    if message:
        message_html = f'<div class="ok">{html.escape(message)}</div>'
    if error:
        message_html = f'<div class="error">{html.escape(error)}</div>'

    return HTMLResponse(f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FICO Control - Ai uitat parola?</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#f4f6f8;font-family:Arial,sans-serif;color:#17212b}}
.wrap{{width:min(92%,500px);margin:65px auto}}
.card{{background:#fff;border-radius:22px;padding:30px;box-shadow:0 12px 35px rgba(0,0,0,.08)}}
.brand{{font-size:13px;font-weight:900;letter-spacing:2px}}
h1{{font-size:30px;margin:22px 0 8px}}
p{{color:#667085;line-height:1.5}}
input,button{{width:100%;padding:14px;border-radius:11px;font-size:16px}}
input{{border:1px solid #d8dde3;margin-top:10px}}
button{{margin-top:16px;border:0;background:#17212b;color:#fff;font-weight:800;cursor:pointer}}
.ok,.error{{margin:14px 0;padding:12px;border-radius:10px;font-weight:700}}
.ok{{background:#e9f8ef;color:#14804a}}
.error{{background:#fdeeee;color:#b42318}}
.back{{display:block;text-align:center;margin-top:16px;color:#17212b;font-weight:700;text-decoration:none}}
</style>
</head>
<body>
<div class="wrap"><div class="card">
<div class="brand">FICO CONTROL</div>
<h1>Ai uitat parola?</h1>
<p>Parola poate fi resetată numai cu un cod trimis la emailul Owner: <strong>{html.escape(mask_email(OWNER_EMAIL))}</strong>.</p>
{message_html}
<form method="post" action="/admin/forgot-password/send-code">
<button type="submit">Trimite codul pe email</button>
</form>
<form method="post" action="/admin/reset-password">
<input name="code" inputmode="numeric" maxlength="6" placeholder="Cod din 6 cifre" required>
<input name="new_password" type="password" minlength="8" placeholder="Parola nouă (minimum 8 caractere)" required>
<button type="submit">Setează parola nouă</button>
</form>
<a class="back" href="/admin/login">Înapoi la login</a>
</div></div>
</body>
</html>
""")


@app.post("/admin/forgot-password/send-code")
def forgot_password_send_code():
    if not OWNER_EMAIL or not RESEND_API_KEY:
        return forgot_password_page(
            error="Emailul Owner sau cheia de email nu este configurată."
        )

    code = f"{secrets.randbelow(1000000):06d}"
    now = utc_now()
    expires = now + timedelta(minutes=OWNER_CODE_MINUTES)

    conn = db()
    conn.execute("UPDATE owner_email_codes SET used=1 WHERE used=0")
    conn.execute(
        """
        INSERT INTO owner_email_codes(code_hash, created_at, expires_at, used)
        VALUES(?,?,?,0)
        """,
        (hash_owner_code(code), iso_utc(now), iso_utc(expires))
    )
    conn.commit()
    conn.close()

    try:
        send_owner_verification_code(code)
    except Exception as exc:
        print(
            "RESEND_PASSWORD_RESET_ERROR:",
            type(exc).__name__,
            str(exc)[:500],
            flush=True
        )
        return forgot_password_page(
            error="Codul nu a putut fi trimis. Verifică setările de email."
        )

    return forgot_password_page(
        message=f"Codul a fost trimis la {mask_email(OWNER_EMAIL)}."
    )


@app.post("/admin/reset-password")
def reset_password_with_email_code(
    code: str = Form(...),
    new_password: str = Form(...)
):
    clean_code = re.sub(r"\D", "", code or "")
    if len(clean_code) != 6:
        return forgot_password_page(error="Cod invalid.")

    if len(new_password) < 8:
        return forgot_password_page(
            error="Parola trebuie să aibă minimum 8 caractere."
        )

    conn = db()
    row = conn.execute(
        """
        SELECT id, code_hash, expires_at
        FROM owner_email_codes
        WHERE used=0
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()

    if not row:
        conn.close()
        return forgot_password_page(error="Nu există un cod activ.")

    expires = parse_iso_utc(row["expires_at"])
    valid = (
        expires
        and expires > utc_now()
        and hmac.compare_digest(
            row["code_hash"],
            hash_owner_code(clean_code)
        )
    )

    if not valid:
        conn.close()
        return forgot_password_page(
            error="Codul este greșit sau a expirat."
        )

    conn.execute(
        "UPDATE owner_email_codes SET used=1 WHERE id=?",
        (row["id"],)
    )
    conn.execute(
        """
        INSERT INTO admin_settings(setting_key, setting_value)
        VALUES('shared_password_hash', ?)
        ON CONFLICT(setting_key) DO UPDATE SET
            setting_value=excluded.setting_value
        """,
        (hash_password(new_password),)
    )
    # Force everyone to use the new password.
    conn.execute("UPDATE admin_sessions SET revoked=1")
    conn.commit()
    conn.close()

    return RedirectResponse("/admin/login", status_code=303)


@app.post("/admin/owner/send-code")
def owner_send_code(request: Request):
    if not OWNER_EMAIL or not RESEND_API_KEY:
        return HTMLResponse(
            owner_gate_html(
                "OWNER_EMAIL sau RESEND_API_KEY nu este configurat în Render.",
                error=True
            ),
            status_code=503
        )

    code = f"{secrets.randbelow(1000000):06d}"
    now = utc_now()
    expires = now + timedelta(minutes=OWNER_CODE_MINUTES)

    conn = db()
    conn.execute("UPDATE owner_email_codes SET used=1 WHERE used=0")
    conn.execute("""
        INSERT INTO owner_email_codes(
            code_hash, created_at, expires_at, used
        ) VALUES(?,?,?,0)
    """, (
        hash_owner_code(code),
        iso_utc(now),
        iso_utc(expires)
    ))
    conn.commit()
    conn.close()

    try:
        send_owner_verification_code(code)
    except Exception as exc:
        print(
            "RESEND_OWNER_EMAIL_ERROR:",
            type(exc).__name__,
            str(exc)[:500],
            flush=True
        )
        return HTMLResponse(
            owner_gate_html(
                "Codul nu a putut fi trimis. Verifică setările de email din Render.",
                error=True
            ),
            status_code=502
        )

    return HTMLResponse(
        owner_gate_html(
            f"Codul a fost trimis la {mask_email(OWNER_EMAIL)}."
        )
    )


@app.post("/admin/owner/verify-code")
def owner_verify_code(request: Request, code: str = Form(...)):
    clean_code = re.sub(r"\\D", "", code or "")
    if len(clean_code) != 6:
        return HTMLResponse(
            owner_gate_html("Cod invalid.", error=True),
            status_code=400
        )

    conn = db()
    row = conn.execute("""
        SELECT id, code_hash, expires_at
        FROM owner_email_codes
        WHERE used=0
        ORDER BY id DESC
        LIMIT 1
    """).fetchone()

    if not row:
        conn.close()
        return HTMLResponse(
            owner_gate_html("Nu există un cod activ.", error=True),
            status_code=400
        )

    expires = parse_iso_utc(row["expires_at"])
    valid = (
        expires
        and expires > utc_now()
        and hmac.compare_digest(
            row["code_hash"],
            hash_owner_code(clean_code)
        )
    )

    if not valid:
        conn.close()
        return HTMLResponse(
            owner_gate_html("Codul este greșit sau a expirat.", error=True),
            status_code=400
        )

    token = request.cookies.get(ADMIN_COOKIE_NAME)
    owner_until = utc_now() + timedelta(minutes=OWNER_VERIFY_MINUTES)

    conn.execute(
        "UPDATE owner_email_codes SET used=1 WHERE id=?",
        (row["id"],)
    )
    conn.execute(
        "UPDATE admin_sessions SET owner_verified_until=? WHERE session_token=?",
        (iso_utc(owner_until), token)
    )
    conn.commit()
    conn.close()

    return RedirectResponse("/admin/owner", status_code=303)


@app.get("/admin/owner", response_class=HTMLResponse)
def admin_owner_page(request: Request):
    session = getattr(request.state, "admin_session", None)
    if not is_owner_verified(session):
        return HTMLResponse(owner_gate_html())

    conn = db()
    active = conn.execute("""
        SELECT display_name,
               normalized_name,
               MIN(created_at) AS created_at,
               MAX(last_seen) AS last_seen,
               MAX(expires_at) AS expires_at,
               COUNT(*) AS session_count
        FROM admin_sessions
        WHERE revoked=0
          AND expires_at > ?
        GROUP BY normalized_name, display_name
        ORDER BY MAX(last_seen) DESC
    """, (iso_utc(utc_now()),)).fetchall()

    blocked = conn.execute("""
        SELECT display_name, normalized_name, blocked_at
        FROM blocked_admin_names
        ORDER BY blocked_at DESC
    """).fetchall()
    conn.close()

    rows = ""
    for r in active:
        rows += f"""
        <tr>
          <td><strong>{html.escape(r["display_name"])}</strong></td>
          <td>{html.escape(r["last_seen"][0:16].replace("T"," "))}</td>
          <td>{r["session_count"]}</td>
          <td>
            <form method="post" action="/admin/owner/revoke" style="display:inline">
              <input type="hidden" name="display_name" value="{html.escape(r["display_name"], quote=True)}">
              <button class="small" type="submit">Deconectează</button>
            </form>
            <form method="post" action="/admin/owner/block" style="display:inline">
              <input type="hidden" name="display_name" value="{html.escape(r["display_name"], quote=True)}">
              <button class="small danger" type="submit">Blochează</button>
            </form>
          </td>
        </tr>
        """

    blocked_rows = ""
    for r in blocked:
        blocked_rows += f"""
        <tr>
          <td><strong>{html.escape(r["display_name"])}</strong></td>
          <td>{html.escape(r["blocked_at"][0:16].replace("T"," "))}</td>
          <td>
            <form method="post" action="/admin/owner/unblock">
              <input type="hidden" name="display_name" value="{html.escape(r["display_name"], quote=True)}">
              <button class="small" type="submit">Deblochează</button>
            </form>
          </td>
        </tr>
        """

    page = f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FICO Control Owner</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#f4f6f8;font-family:Arial,sans-serif;color:#17212b}}
.wrap{{width:min(95%,1100px);margin:35px auto 60px}}
.top{{display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap}}
.brand{{font-size:13px;font-weight:900;letter-spacing:2px}}
h1{{font-size:36px;margin:14px 0 25px}}
.card{{background:#fff;border-radius:16px;padding:20px;margin-bottom:16px;box-shadow:0 5px 18px rgba(0,0,0,.05)}}
table{{width:100%;border-collapse:collapse}}
th,td{{text-align:left;padding:13px;border-bottom:1px solid #eceff2}}
input,button{{padding:11px 13px;border-radius:9px;border:1px solid #d8dde3}}
button{{cursor:pointer;font-weight:800}}
.small{{background:#fff;color:#17212b}}
.danger{{background:#fdeeee;color:#b42318;border-color:#f4b8b3}}
.primary{{background:#17212b;color:#fff;border:0}}
a{{color:#17212b;font-weight:800;text-decoration:none}}
.password-row{{display:flex;gap:10px;flex-wrap:wrap}}
.password-row input{{min-width:260px;flex:1}}
</style>
</head>
<body>
<main class="wrap">
<div class="top">
<div>
<div class="brand">FICO CONTROL · OWNER</div>
<h1>Control acces Admin</h1>
</div>
<a href="/admin">Înapoi la Dashboard</a>
</div>

<section class="card">
<h2>Schimbă parola comună</h2>
<p>Confirmarea Owner este valabilă temporar după codul primit pe email. Schimbarea parolei va deconecta toate celelalte sesiuni.</p>
<form class="password-row" method="post" action="/admin/owner/change-password">
<input type="password" name="new_password" minlength="8" placeholder="Parola nouă (minimum 8 caractere)" required>
<button class="primary" type="submit">Schimbă parola</button>
</form>
</section>

<section class="card">
<h2>Persoane conectate</h2>
<table>
<thead><tr><th>Nume</th><th>Ultima activitate</th><th>Sesiuni</th><th>Acțiuni</th></tr></thead>
<tbody>{rows or '<tr><td colspan="4">Nicio sesiune activă.</td></tr>'}</tbody>
</table>
</section>

<section class="card">
<h2>Nume blocate</h2>
<table>
<thead><tr><th>Nume</th><th>Blocat la</th><th>Acțiune</th></tr></thead>
<tbody>{blocked_rows or '<tr><td colspan="3">Niciun nume blocat.</td></tr>'}</tbody>
</table>
</section>
</main>
</body>
</html>
"""
    return HTMLResponse(page)


def require_owner_verified(request: Request):
    session = getattr(request.state, "admin_session", None)
    if not is_owner_verified(session):
        raise HTTPException(status_code=403, detail="owner_verification_required")
    return session


@app.post("/admin/owner/revoke")
def owner_revoke_user(request: Request, display_name: str = Form(...)):
    require_owner_verified(request)
    normalized = normalize_admin_name(display_name)

    conn = db()
    conn.execute(
        "UPDATE admin_sessions SET revoked=1 WHERE normalized_name=?",
        (normalized,)
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/admin/owner", status_code=303)


@app.post("/admin/owner/block")
def owner_block_user(request: Request, display_name: str = Form(...)):
    require_owner_verified(request)
    clean = " ".join((display_name or "").strip().split())
    normalized = normalize_admin_name(clean)

    conn = db()
    conn.execute("""
        INSERT INTO blocked_admin_names(
            normalized_name, display_name, blocked_at
        ) VALUES(?,?,?)
        ON CONFLICT(normalized_name) DO UPDATE SET
            display_name=excluded.display_name,
            blocked_at=excluded.blocked_at
    """, (normalized, clean, iso_utc(utc_now())))
    conn.execute(
        "UPDATE admin_sessions SET revoked=1 WHERE normalized_name=?",
        (normalized,)
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/admin/owner", status_code=303)


@app.post("/admin/owner/unblock")
def owner_unblock_user(request: Request, display_name: str = Form(...)):
    require_owner_verified(request)
    normalized = normalize_admin_name(display_name)

    conn = db()
    conn.execute(
        "DELETE FROM blocked_admin_names WHERE normalized_name=?",
        (normalized,)
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/admin/owner", status_code=303)


@app.post("/admin/owner/change-password")
def owner_change_password(
    request: Request,
    new_password: str = Form(...)
):
    session = require_owner_verified(request)

    if len(new_password) < 8:
        raise HTTPException(
            status_code=400,
            detail="password_must_have_at_least_8_characters"
        )

    token = request.cookies.get(ADMIN_COOKIE_NAME)
    conn = db()
    conn.execute("""
        INSERT INTO admin_settings(setting_key, setting_value)
        VALUES('shared_password_hash', ?)
        ON CONFLICT(setting_key) DO UPDATE SET
            setting_value=excluded.setting_value
    """, (hash_password(new_password),))

    conn.execute(
        "UPDATE admin_sessions SET revoked=1 WHERE session_token<>?",
        (token,)
    )
    conn.commit()
    conn.close()

    return RedirectResponse("/admin/owner", status_code=303)




def normalize_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("-", " ")
    value = " ".join(value.strip().split())
    return value.casefold()


def token_matches(entered_token: str, full_token: str) -> bool:
    if not entered_token or not full_token:
        return False

    if entered_token == full_token:
        return True

    if len(entered_token) == 1:
        return full_token.startswith(entered_token)

    if len(entered_token) >= 2 and full_token.startswith(entered_token):
        return True

    if len(entered_token) >= 4:
        return SequenceMatcher(None, entered_token, full_token).ratio() >= 0.84

    return False


def token_subsequence_score(entered_tokens: list[str], full_tokens: list[str]) -> float:
    if not entered_tokens or not full_tokens:
        return 0.0

    matched_positions = []
    search_from = 0

    for entered_token in entered_tokens:
        found = None
        for idx in range(search_from, len(full_tokens)):
            if token_matches(entered_token, full_tokens[idx]):
                found = idx
                break

        if found is None:
            return 0.0

        matched_positions.append(found)
        search_from = found + 1

    exact = sum(
        1
        for entered_token, idx in zip(entered_tokens, matched_positions)
        if entered_token == full_tokens[idx]
    )
    exact_ratio = exact / len(entered_tokens)

    first_last_bonus = 0.0
    if len(entered_tokens) >= 2:
        if (
            token_matches(entered_tokens[0], full_tokens[0])
            and token_matches(entered_tokens[-1], full_tokens[-1])
        ):
            first_last_bonus = 0.06

    return min(1.0, 0.88 + 0.06 * exact_ratio + first_last_bonus)


def unordered_token_score(entered_tokens: list[str], full_tokens: list[str]) -> float:
    if not entered_tokens or not full_tokens:
        return 0.0

    used = set()
    exact = 0

    for entered_token in entered_tokens:
        best_idx = None
        for idx, full_token in enumerate(full_tokens):
            if idx in used:
                continue
            if token_matches(entered_token, full_token):
                best_idx = idx
                if entered_token == full_token:
                    exact += 1
                break

        if best_idx is None:
            return 0.0

        used.add(best_idx)

    exact_ratio = exact / len(entered_tokens)
    return min(0.94, 0.86 + 0.08 * exact_ratio)


def name_similarity(a: str, b: str) -> float:
    a_n = normalize_name(a)
    b_n = normalize_name(b)

    if not a_n or not b_n:
        return 0.0

    if a_n == b_n:
        return 1.0

    if b_n.startswith(a_n) or a_n.startswith(b_n):
        shorter = min(len(a_n), len(b_n))
        longer = max(len(a_n), len(b_n))
        return 0.92 + (0.08 * shorter / max(longer, 1))

    entered_tokens = a_n.split()
    full_tokens = b_n.split()

    sequence_score = token_subsequence_score(entered_tokens, full_tokens)
    unordered_score = unordered_token_score(entered_tokens, full_tokens)
    ratio = SequenceMatcher(None, a_n, b_n).ratio()

    return max(sequence_score, unordered_score, ratio * 0.90)


def find_required_driver(conn, work_date: str, entered_name: str):
    entered = " ".join((entered_name or "").strip().split())
    target = normalize_name(entered)

    rows = conn.execute("""
        SELECT d.id, d.name
        FROM daily_required r
        JOIN drivers d ON d.id = r.driver_id
        WHERE r.work_date = ?
        ORDER BY d.name
    """, (work_date,)).fetchall()

    # Exact match first.
    for row in rows:
        if normalize_name(row["name"]) == target:
            return row, 1.0, False

    scored = []
    for row in rows:
        score = name_similarity(entered, row["name"])
        scored.append((score, row))

    scored.sort(key=lambda item: item[0], reverse=True)

    if not scored:
        return None, 0.0, False

    best_score, best_row = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0

    # Accept clear partial names such as:
    # "Alain Sery" -> "Alain Gnebehi Kamin Sery"
    # "Elvis V"    -> "Elvis Velcu"
    #
    # If two drivers match almost equally well, do not guess.
    if best_score >= 0.84:
        ambiguous = second_score >= 0.84 and (best_score - second_score) < 0.06
        if ambiguous:
            return None, best_score, True
        return best_row, best_score, False

    return None, best_score, False


def extract_fico_candidates_from_text(text_value: str) -> list[int]:
    if not text_value:
        return []

    candidates = []
    for match in re.findall(r"(?<!\d)(\d{3})(?!\d)", text_value):
        try:
            value = int(match)
        except ValueError:
            continue

        # FICO in this workflow is capped at 850.
        if 300 <= value <= 850:
            candidates.append(value)

    return candidates


def detect_fico_from_image_bytes(raw: bytes, content_type: str) -> tuple[int | None, str]:
    """
    Optional OCR integration.

    If environment variable OCRSPACE_API_KEY is configured on Render,
    the image is sent to OCR.Space and the returned text is searched for
    a plausible FICO score. Without an OCR key, the function returns
    'ocr_not_configured' and Admin will request manual verification.
    """
    api_key = os.getenv("OCRSPACE_API_KEY", "").strip()

    if not api_key:
        return None, "ocr_not_configured"

    boundary = "----FICOControlBoundary7MA4YWxkTrZu0gW"
    ext = "jpg"
    if content_type == "image/png":
        ext = "png"
    elif content_type == "image/webp":
        ext = "webp"

    parts = []

    def add_field(name, value):
        parts.append(
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
            f"{value}\r\n".encode("utf-8")
        )

    add_field("apikey", api_key)
    add_field("language", "eng")
    add_field("isOverlayRequired", "false")
    add_field("OCREngine", "2")

    file_header = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="fico.{ext}"\r\n'
        f"Content-Type: {content_type}\r\n\r\n"
    ).encode("utf-8")

    parts.append(file_header + raw + b"\r\n")
    parts.append(f"--{boundary}--\r\n".encode("utf-8"))
    body = b"".join(parts)

    req = urllib.request.Request(
        "https://api.ocr.space/parse/image",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST"
    )

    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8", errors="replace"))
    except Exception:
        return None, "ocr_error"

    parsed = payload.get("ParsedResults") or []
    combined = "\n".join(
        str(item.get("ParsedText") or "")
        for item in parsed
    )

    candidates = extract_fico_candidates_from_text(combined)

    if not candidates:
        return None, "ocr_unreadable"

    # Usually the FICO score is the highest relevant 3-digit number in
    # the screenshot. Prefer 850 when present, otherwise the highest.
    detected = max(candidates)
    return detected, "ocr_ok"


async def save_proof_image(proof: UploadFile) -> tuple[str, str, bytes]:
    if proof.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=400, detail="invalid_image_type")

    raw = await proof.read()

    if not raw:
        raise HTTPException(status_code=400, detail="empty_image")

    if len(raw) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="image_too_large")

    extension = {
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp"
    }[proof.content_type]

    filename = f"{uuid.uuid4().hex}{extension}"
    original_name = proof.filename or "proof"

    if R2_ENABLED:
        try:
            client = r2_client()
            client.put_object(
                Bucket=R2_BUCKET_NAME,
                Key=filename,
                Body=raw,
                ContentType=proof.content_type,
                Metadata={
                    "original-name": original_name[:500]
                }
            )
        except Exception as exc:
            print(
                "R2_UPLOAD_ERROR:",
                type(exc).__name__,
                str(exc)[:500],
                flush=True
            )
            raise HTTPException(
                status_code=502,
                detail="proof_storage_error"
            )
    else:
        # Local fallback for development only.
        path = os.path.join(UPLOAD_DIR, filename)
        with open(path, "wb") as f:
            f.write(raw)

    return filename, original_name, raw


class SubmissionIn(BaseModel):
    driver_id: int
    fico_score: int


@app.get("/api/health")
def health():
    return {"ok": True, "service": "FICO Control"}


@app.get("/api/drivers/today")
def drivers_today():
    today = date.today().isoformat()
    conn = db()

    rows = conn.execute("""
        SELECT d.id, d.name
        FROM daily_required r
        JOIN drivers d ON d.id = r.driver_id
        WHERE r.work_date = ?
        ORDER BY d.name
    """, (today,)).fetchall()

    conn.close()

    return {
        "date": today,
        "drivers": [{"id": r["id"], "name": r["name"]} for r in rows]
    }


# Kept temporarily for compatibility with the existing mobile prototype.
@app.post("/api/submissions")
def submit_mobile_legacy(payload: SubmissionIn):
    if payload.fico_score < 0 or payload.fico_score > 850:
        raise HTTPException(status_code=400, detail="invalid_score")

    today = date.today().isoformat()
    conn = db()

    required = conn.execute(
        "SELECT 1 FROM daily_required WHERE work_date=? AND driver_id=?",
        (today, payload.driver_id)
    ).fetchone()

    if not required:
        conn.close()
        raise HTTPException(status_code=400, detail="not_required")

    existing = conn.execute(
        "SELECT 1 FROM submissions WHERE work_date=? AND driver_id=?",
        (today, payload.driver_id)
    ).fetchone()

    if existing:
        conn.close()
        raise HTTPException(status_code=409, detail="already_sent")

    conn.execute("""
        INSERT INTO submissions(
            work_date, driver_id, fico_score, submitted_at
        ) VALUES(?,?,?,?)
    """, (
        today,
        payload.driver_id,
        payload.fico_score,
        datetime.now().isoformat(timespec="seconds")
    ))

    conn.commit()
    conn.close()
    return {"ok": True}



@app.post("/api/submissions/photo")
async def submit_mobile_photo(
    full_name: str = Form(...),
    fico_score: int = Form(...),
    proof: UploadFile = File(...)
):
    if fico_score < 300 or fico_score > 850:
        raise HTTPException(status_code=400, detail="invalid_score")

    clean_entered_name = " ".join((full_name or "").strip().split())
    if not clean_entered_name:
        raise HTTPException(status_code=400, detail="missing_name")

    today = date.today().isoformat()
    conn = db()

    driver, match_score, ambiguous = find_required_driver(
        conn,
        today,
        clean_entered_name
    )

    # Always receive and save the proof first. Even if the name cannot be
    # identified safely, the submission must not be lost.
    try:
        filename, original_name, raw = await save_proof_image(proof)
    except Exception:
        conn.close()
        raise

    detected_score, ocr_state = detect_fico_from_image_bytes(
        raw,
        proof.content_type or "image/jpeg"
    )

    if detected_score is None:
        verification_status = "manual_review"
    elif detected_score == fico_score:
        verification_status = "verified"
    else:
        verification_status = "mismatch"

    # If we cannot safely identify the driver, keep the submission in a
    # separate review queue instead of rejecting it.
    if not driver:
        best_match_name = None

        # Find the highest candidate only as a hint for the admin.
        candidates = conn.execute("""
            SELECT d.name
            FROM daily_required r
            JOIN drivers d ON d.id = r.driver_id
            WHERE r.work_date = ?
            ORDER BY d.name
        """, (today,)).fetchall()

        best_hint_score = 0.0
        for candidate in candidates:
            candidate_score = name_similarity(
                clean_entered_name,
                candidate["name"]
            )
            if candidate_score > best_hint_score:
                best_hint_score = candidate_score
                best_match_name = candidate["name"]

        reason = "ambiguous_name" if ambiguous else "unrecognized_name"

        conn.execute("""
            INSERT INTO unresolved_submissions(
                work_date,
                entered_full_name,
                fico_score,
                submitted_at,
                proof_filename,
                proof_original_name,
                detected_fico_score,
                verification_status,
                best_match_name,
                best_match_score,
                match_reason
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """, (
            today,
            clean_entered_name,
            fico_score,
            datetime.now().isoformat(timespec="seconds"),
            filename,
            original_name,
            detected_score,
            verification_status,
            best_match_name,
            best_hint_score,
            reason
        ))

        conn.commit()
        conn.close()

        return {
            "ok": True,
            "needs_name_review": True,
            "entered_name": clean_entered_name,
            "best_match_name": best_match_name,
            "best_match_score": round(best_hint_score, 3),
            "detected_fico_score": detected_score,
            "verification_status": verification_status,
            "ocr_state": ocr_state
        }

    existing = conn.execute(
        "SELECT 1 FROM submissions WHERE work_date=? AND driver_id=?",
        (today, driver["id"])
    ).fetchone()

    if existing:
        conn.close()
        raise HTTPException(status_code=409, detail="already_sent")

    conn.execute("""
        INSERT INTO submissions(
            work_date,
            driver_id,
            fico_score,
            submitted_at,
            entered_full_name,
            proof_filename,
            proof_original_name,
            detected_fico_score,
            verification_status,
            name_match_score
        ) VALUES(?,?,?,?,?,?,?,?,?,?)
    """, (
        today,
        driver["id"],
        fico_score,
        datetime.now().isoformat(timespec="seconds"),
        clean_entered_name,
        filename,
        original_name,
        detected_score,
        verification_status,
        match_score
    ))

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "driver": driver["name"],
        "name_match_score": round(match_score, 3),
        "detected_fico_score": detected_score,
        "verification_status": verification_status,
        "ocr_state": ocr_state
    }


def extract_driver_names_from_xlsx(raw: bytes) -> list[str]:
    workbook = openpyxl.load_workbook(
        io.BytesIO(raw),
        read_only=True,
        data_only=True
    )

    ws = workbook["Strecken"] if "Strecken" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)

    if not headers:
        return []

    normalized = [
        str(h).strip().lower() if h is not None else ""
        for h in headers
    ]

    candidates = [
        "name des fahrers",
        "fahrername",
        "fahrer",
        "driver name",
        "driver",
        "name",
        "nume șofer",
        "nume sofer",
        "nume"
    ]

    name_index = None
    for candidate in candidates:
        if candidate in normalized:
            name_index = normalized.index(candidate)
            break

    if name_index is None:
        raise ValueError("driver_column_not_found")

    names, seen = [], set()

    for row in rows:
        if name_index >= len(row) or row[name_index] is None:
            continue

        raw_name = str(row[name_index]).strip()
        if not raw_name:
            continue

        # Cortex can show the assigned driver followed by rescue/helper drivers
        # in the same cell, usually separated by "|" or line breaks.
        # For FICO Control, ONLY the first driver belongs to that route.
        parts = re.split(r"[|\n\r]+", raw_name)
        name = next((part.strip() for part in parts if part.strip()), "")

        if not name:
            continue

        key = normalize_name(name)
        if key not in seen:
            seen.add(key)
            names.append(name)

    return names


def mentor_name_key(value: str) -> str:
    """Normalize a person's name independent of first/last-name order."""
    cleaned = normalize_name(value or "")
    tokens = [token for token in cleaned.split() if token]
    return " ".join(sorted(tokens))


def parse_mentor_time(value):
    """Return comparable minutes since midnight + HH:MM display text."""
    if value is None:
        return None, None

    if isinstance(value, datetime):
        return value.hour * 60 + value.minute, f"{value.hour:02d}:{value.minute:02d}"

    if isinstance(value, time):
        return value.hour * 60 + value.minute, f"{value.hour:02d}:{value.minute:02d}"

    raw = str(value).strip()
    if not raw:
        return None, None

    formats = [
        "%I:%M %p",
        "%I:%M:%S %p",
        "%H:%M",
        "%H:%M:%S",
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(raw.lower(), fmt)
            minutes = parsed.hour * 60 + parsed.minute
            return minutes, f"{parsed.hour:02d}:{parsed.minute:02d}"
        except Exception:
            pass

    return None, raw


def extract_mentor_names_from_xlsx(raw: bytes) -> list[dict]:
    """
    Read Mentor Shift Report and return one record per unique driver.
    If a driver appears multiple times, keep the earliest Begin Route Time.
    """
    workbook = openpyxl.load_workbook(
        io.BytesIO(raw),
        read_only=True,
        data_only=True
    )
    ws = workbook[workbook.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    header_map = {
        str(value).strip().lower(): index
        for index, value in enumerate(headers)
        if value is not None
    }

    first_index = header_map.get("first name")
    last_index = header_map.get("last name")
    begin_index = header_map.get("begin route time")

    if first_index is None or last_index is None:
        raise ValueError("mentor_name_columns_not_found")

    records = {}

    for row in rows:
        first = ""
        last = ""

        if first_index < len(row) and row[first_index] is not None:
            first = str(row[first_index]).strip()

        if last_index < len(row) and row[last_index] is not None:
            last = str(row[last_index]).strip()

        name = " ".join(part for part in (first, last) if part).strip()
        if not name:
            continue

        key = mentor_name_key(name)
        if not key:
            continue

        raw_begin = (
            row[begin_index]
            if begin_index is not None and begin_index < len(row)
            else None
        )
        minutes, display_time = parse_mentor_time(raw_begin)

        existing = records.get(key)

        if existing is None:
            records[key] = {
                "name": name,
                "first_connection_time": display_time,
                "_minutes": minutes,
            }
            continue

        old_minutes = existing.get("_minutes")

        if minutes is not None and (
            old_minutes is None or minutes < old_minutes
        ):
            existing["first_connection_time"] = display_time
            existing["_minutes"] = minutes

    result = []
    for record in records.values():
        record.pop("_minutes", None)
        result.append(record)

    return result


def mentor_compare(required_name: str, connected_names: list[str]):
    """Return connected/review/missing plus best Mentor name and confidence."""
    required_key = mentor_name_key(required_name)
    if not required_key:
        return "missing", None, 0.0

    exact = {mentor_name_key(name): name for name in connected_names}
    if required_key in exact:
        return "connected", exact[required_key], 1.0

    required_tokens = set(required_key.split())
    best_name = None
    best_score = 0.0

    for candidate in connected_names:
        candidate_key = mentor_name_key(candidate)
        candidate_tokens = set(candidate_key.split())

        sequence_score = SequenceMatcher(None, required_key, candidate_key).ratio()
        union = required_tokens | candidate_tokens
        token_score = (
            len(required_tokens & candidate_tokens) / len(union)
            if union else 0.0
        )
        score = max(sequence_score, sequence_score * 0.65 + token_score * 0.35)

        if score > best_score:
            best_score = score
            best_name = candidate

    if best_score >= 0.88:
        return "connected", best_name, best_score
    if best_score >= 0.72:
        return "review", best_name, best_score
    return "missing", best_name, best_score


@app.post("/admin/mentor/upload-cortex")
async def mentor_upload_cortex(
    work_date: str = Form(...),
    file: UploadFile = File(...)
):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return RedirectResponse(
            f"/admin/mentor?d={work_date}&error=cortex",
            status_code=303
        )

    raw = await file.read()
    try:
        # This existing function already applies the user's rescue rule:
        # for Name des Fahrers values separated with |, only the first driver counts.
        names = extract_driver_names_from_xlsx(raw)
    except Exception:
        return RedirectResponse(
            f"/admin/mentor?d={work_date}&error=cortex",
            status_code=303
        )

    conn = db()
    conn.execute("DELETE FROM mentor_required WHERE work_date=?", (work_date,))
    imported_at = datetime.now(timezone.utc).isoformat()

    for name in names:
        key = mentor_name_key(name)
        if not key:
            continue
        conn.execute(
            """
            INSERT INTO mentor_required(
                work_date, driver_name, normalized_name, imported_at, source_filename
            ) VALUES(?,?,?,?,?)
            ON CONFLICT(work_date, normalized_name) DO UPDATE SET
                driver_name=excluded.driver_name,
                imported_at=excluded.imported_at,
                source_filename=excluded.source_filename
            """,
            (work_date, name, key, imported_at, file.filename)
        )

    conn.commit()
    conn.close()
    return RedirectResponse(f"/admin/mentor?d={work_date}", status_code=303)


@app.post("/admin/mentor/upload-shift")
async def mentor_upload_shift(
    work_date: str = Form(...),
    file: UploadFile = File(...)
):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return RedirectResponse(
            f"/admin/mentor?d={work_date}&error=mentor",
            status_code=303
        )

    raw = await file.read()
    try:
        mentor_records = extract_mentor_names_from_xlsx(raw)
    except Exception:
        return RedirectResponse(
            f"/admin/mentor?d={work_date}&error=mentor",
            status_code=303
        )

    conn = db()
    conn.execute("DELETE FROM mentor_connected WHERE work_date=?", (work_date,))
    imported_at = datetime.now(timezone.utc).isoformat()

    for record in mentor_records:
        name = record["name"]
        first_connection_time = record.get("first_connection_time")
        key = mentor_name_key(name)

        if not key:
            continue

        conn.execute(
            """
            INSERT INTO mentor_connected(
                work_date,
                driver_name,
                normalized_name,
                imported_at,
                source_filename,
                first_connection_time
            ) VALUES(?,?,?,?,?,?)
            ON CONFLICT(work_date, normalized_name) DO UPDATE SET
                driver_name=excluded.driver_name,
                imported_at=excluded.imported_at,
                source_filename=excluded.source_filename,
                first_connection_time=excluded.first_connection_time
            """,
            (
                work_date,
                name,
                key,
                imported_at,
                file.filename,
                first_connection_time
            )
        )

    conn.commit()
    conn.close()
    return RedirectResponse(f"/admin/mentor?d={work_date}", status_code=303)



@app.post("/admin/mentor/upload-both")
async def mentor_upload_both(
    work_date: str = Form(...),
    cortex_file: UploadFile = File(...),
    mentor_file: UploadFile = File(...)
):
    if not (cortex_file.filename or "").lower().endswith(".xlsx"):
        return RedirectResponse(
            f"/admin/mentor?d={work_date}&error=cortex",
            status_code=303
        )

    if not (mentor_file.filename or "").lower().endswith(".xlsx"):
        return RedirectResponse(
            f"/admin/mentor?d={work_date}&error=mentor",
            status_code=303
        )

    cortex_raw = await cortex_file.read()
    mentor_raw = await mentor_file.read()

    try:
        required_names = extract_driver_names_from_xlsx(cortex_raw)
        mentor_records = extract_mentor_names_from_xlsx(mentor_raw)
    except Exception:
        return RedirectResponse(
            f"/admin/mentor?d={work_date}&error=files",
            status_code=303
        )

    conn = db()
    imported_at = datetime.now(timezone.utc).isoformat()

    # Replace both daily lists only after BOTH files were parsed successfully.
    conn.execute("DELETE FROM mentor_required WHERE work_date=?", (work_date,))
    conn.execute("DELETE FROM mentor_connected WHERE work_date=?", (work_date,))

    for name in required_names:
        key = mentor_name_key(name)
        if not key:
            continue

        conn.execute(
            """
            INSERT INTO mentor_required(
                work_date,
                driver_name,
                normalized_name,
                imported_at,
                source_filename
            ) VALUES(?,?,?,?,?)
            ON CONFLICT(work_date, normalized_name) DO UPDATE SET
                driver_name=excluded.driver_name,
                imported_at=excluded.imported_at,
                source_filename=excluded.source_filename
            """,
            (
                work_date,
                name,
                key,
                imported_at,
                cortex_file.filename
            )
        )

    for record in mentor_records:
        name = record["name"]
        key = mentor_name_key(name)
        if not key:
            continue

        conn.execute(
            """
            INSERT INTO mentor_connected(
                work_date,
                driver_name,
                normalized_name,
                imported_at,
                source_filename,
                first_connection_time
            ) VALUES(?,?,?,?,?,?)
            ON CONFLICT(work_date, normalized_name) DO UPDATE SET
                driver_name=excluded.driver_name,
                imported_at=excluded.imported_at,
                source_filename=excluded.source_filename,
                first_connection_time=excluded.first_connection_time
            """,
            (
                work_date,
                name,
                key,
                imported_at,
                mentor_file.filename,
                record.get("first_connection_time")
            )
        )

    conn.commit()
    conn.close()

    return RedirectResponse(
        f"/admin/mentor?d={work_date}&uploaded=1",
        status_code=303
    )


def parse_hours_timestamp(value: str) -> datetime:
    clean = (value or "").strip()
    if not clean:
        raise ValueError("missing timestamp")
    return datetime.fromisoformat(clean.replace("Z", "+00:00"))


def merge_work_intervals(intervals):
    merged = []
    for start_at, end_at in sorted(intervals, key=lambda item: (item[0], item[1])):
        if not merged or start_at > merged[-1][1]:
            merged.append([start_at, end_at])
        elif end_at > merged[-1][1]:
            merged[-1][1] = end_at
    return merged


def parse_service_details_csv(csv_bytes: bytes):
    text = csv_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    required = {"Datum", "Zustellmitarbeiter", "Anmelden", "Abmelden"}
    if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
        raise ValueError("Raportul nu conține coloanele necesare.")

    grouped = {}
    for row in reader:
        driver = (row.get("Zustellmitarbeiter") or "").strip()
        work_date = (row.get("Datum") or "").strip()
        if not driver or not work_date:
            continue
        try:
            start_at = parse_hours_timestamp(row.get("Anmelden") or "")
            end_at = parse_hours_timestamp(row.get("Abmelden") or "")
        except (TypeError, ValueError):
            continue
        if end_at <= start_at:
            continue

        key = (work_date, driver)
        group = grouped.setdefault(key, {"intervals": set(), "routes": set()})
        group["intervals"].add((start_at, end_at))
        route = (row.get("Route") or "").strip()
        if route:
            group["routes"].add(route)

    results = []
    for (work_date, driver), group in grouped.items():
        merged = merge_work_intervals(group["intervals"])
        total_seconds = int(sum((end - start).total_seconds() for start, end in merged))
        if not merged or total_seconds <= 0:
            continue
        total_minutes = total_seconds // 60
        first_start = min(start for start, _ in merged).astimezone(BERLIN_TZ)
        last_end = max(end for _, end in merged).astimezone(BERLIN_TZ)
        if total_minutes >= 630:
            status = "critical"
        elif total_minutes > 600:
            status = "warning"
        else:
            status = "safe"
        results.append({
            "date": work_date,
            "name": driver,
            "start": first_start.strftime("%H:%M"),
            "end": last_end.strftime("%H:%M"),
            "minutes": total_minutes,
            "seconds": total_seconds,
            "blocks": len(merged),
            "routes": sorted(group["routes"]),
            "status": status
        })
    return sorted(results, key=lambda row: (row["date"], row["name"].casefold()))


def parse_anfahrlisten_clock(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)

    if isinstance(value, time):
        return value.replace(tzinfo=None)

    if isinstance(value, (int, float)) and 0 <= value < 1:
        seconds = int(round(float(value) * 24 * 60 * 60)) % (24 * 60 * 60)
        return time(
            hour=seconds // 3600,
            minute=(seconds % 3600) // 60,
            second=seconds % 60
        )

    clean = str(value).strip()
    if not clean or clean.casefold() in {"fehlt", "missing", "none", "-"}:
        return None

    match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", clean)
    if not match:
        return None

    hour, minute, second = (int(part or 0) for part in match.groups())
    if hour > 23 or minute > 59 or second > 59:
        return None
    return time(hour=hour, minute=minute, second=second)


def parse_anfahrlisten_filename(upload_name: str):
    match = re.search(
        r"(20\d{2})[-_](\d{2})[-_](\d{2})"
        r"(?:[_ T-](\d{2})[_:\-](\d{2}))?",
        upload_name or ""
    )
    if not match:
        raise ValueError(
            "Data nu a putut fi citită din numele fișierului Anfahrlisten."
        )

    try:
        work_date = date(
            int(match.group(1)),
            int(match.group(2)),
            int(match.group(3))
        )
    except ValueError as exc:
        raise ValueError("Data din numele fișierului nu este validă.") from exc

    snapshot_at = None
    if match.group(4) is not None:
        try:
            snapshot_at = datetime.combine(
                work_date,
                time(int(match.group(4)), int(match.group(5))),
                tzinfo=BERLIN_TZ
            )
        except ValueError as exc:
            raise ValueError(
                "Ora din numele fișierului Anfahrlisten nu este validă."
            ) from exc

    return work_date, snapshot_at


def clean_anfahrlisten_name(value):
    name = re.sub(r"\s*,\s*", " ", str(value or "").strip())
    return re.sub(r"\s+", " ", name).strip()


def parse_anfahrlisten_xlsx(upload_name: str, xlsx_bytes: bytes):
    work_date, snapshot_at = parse_anfahrlisten_filename(upload_name)

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes),
            read_only=True,
            data_only=True
        )
    except (
        zipfile.BadZipFile,
        KeyError,
        OSError,
        ValueError,
        openpyxl.utils.exceptions.InvalidFileException
    ) as exc:
        raise ValueError("Fișierul Anfahrlisten XLSX nu este valid.") from exc

    required = {
        "Transporter-ID",
        "Name des Fahrers",
        "Routencode",
        "App-Anmeldung:",
        "App-Abmeldung:"
    }

    try:
        worksheet = workbook.active
        header = None
        rows = worksheet.iter_rows(values_only=True)

        for candidate in rows:
            normalized = [
                re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()
                for value in candidate
            ]
            if required.issubset(set(normalized)):
                header = normalized
                break

        if header is None:
            raise ValueError(
                "Fișierul nu conține coloanele Anfahrlisten necesare."
            )

        indexes = {column: header.index(column) for column in required}
        grouped = {}

        for row in rows:
            def cell(column):
                index = indexes[column]
                return row[index] if index < len(row) else None

            driver_id = str(cell("Transporter-ID") or "").strip()
            driver_name = clean_anfahrlisten_name(cell("Name des Fahrers"))
            if not driver_id or not driver_name:
                continue

            start_clock = parse_anfahrlisten_clock(cell("App-Anmeldung:"))
            if start_clock is None:
                continue

            start_at = datetime.combine(
                work_date,
                start_clock,
                tzinfo=BERLIN_TZ
            )
            end_clock = parse_anfahrlisten_clock(cell("App-Abmeldung:"))
            is_active = end_clock is None

            if end_clock is None:
                if snapshot_at is None:
                    if work_date == datetime.now(BERLIN_TZ).date():
                        end_at = datetime.now(BERLIN_TZ)
                    else:
                        continue
                else:
                    end_at = snapshot_at
            else:
                end_at = datetime.combine(
                    work_date,
                    end_clock,
                    tzinfo=BERLIN_TZ
                )
                if end_at <= start_at:
                    end_at += timedelta(days=1)

            if end_at <= start_at:
                continue

            key = (work_date.isoformat(), driver_id)
            group = grouped.setdefault(key, {
                "name": driver_name,
                "intervals": set(),
                "routes": set(),
                "active": False
            })
            group["intervals"].add((start_at, end_at))
            group["active"] = group["active"] or is_active

            route = str(cell("Routencode") or "").strip()
            if route:
                group["routes"].add(route)
    finally:
        workbook.close()

    results = []
    for (work_date_text, _driver_id), group in grouped.items():
        merged = merge_work_intervals(group["intervals"])
        total_seconds = int(sum(
            (end_at - start_at).total_seconds()
            for start_at, end_at in merged
        ))
        if not merged or total_seconds <= 0:
            continue

        total_minutes = total_seconds // 60
        if total_minutes >= 630:
            status = "critical"
        elif total_minutes > 600:
            status = "warning"
        else:
            status = "safe"

        results.append({
            "date": work_date_text,
            "name": group["name"],
            "start": min(start for start, _ in merged).strftime("%H:%M"),
            "end": max(end for _, end in merged).strftime("%H:%M"),
            "minutes": total_minutes,
            "seconds": total_seconds,
            "blocks": len(merged),
            "routes": sorted(group["routes"]),
            "status": status,
            "active": group["active"]
        })

    return sorted(
        results,
        key=lambda row: (row["date"], row["name"].casefold())
    )


def read_hours_report(upload_name: str, payload: bytes):
    lower_name = (upload_name or "").lower()
    if len(payload) > HOURS_MAX_UPLOAD_BYTES:
        raise ValueError("Fișierul este prea mare. Limita este 15 MB.")

    if lower_name.endswith(".xlsx"):
        return parse_anfahrlisten_xlsx(upload_name, payload)

    if lower_name.endswith(".csv"):
        if len(payload) > HOURS_MAX_CSV_BYTES:
            raise ValueError("Raportul CSV este prea mare.")
        return parse_service_details_csv(payload)

    if not lower_name.endswith(".zip"):
        raise ValueError("Încarcă fișierul Anfahrlisten XLSX.")

    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            candidates = [
                info for info in archive.infolist()
                if not info.is_dir()
                and re.search(r"service details report.*\.csv$", info.filename, re.IGNORECASE)
            ]
            if not candidates:
                raise ValueError("Arhiva nu conține Service Details Report CSV.")
            report = candidates[0]
            if report.file_size > HOURS_MAX_CSV_BYTES:
                raise ValueError("Raportul din arhivă este prea mare.")
            return parse_service_details_csv(archive.read(report))
    except zipfile.BadZipFile as exc:
        raise ValueError("Arhiva ZIP nu este validă.") from exc


def hours_control_html(results=None, filename="", error=""):
    results = results or []
    payload_json = json.dumps(results, ensure_ascii=False).replace("</", "<\\/")
    filename_safe = html.escape(filename)
    error_html = (
        f'<div class="error"><strong>Raportul nu a putut fi analizat.</strong>'
        f'<span>{html.escape(error)}</span></div>'
        if error else ""
    )
    file_status = (
        f'<strong>{filename_safe}</strong><small>Raport analizat în memorie · fișierul nu a fost salvat</small>'
        if results else
        '<strong>Niciun raport încărcat</strong><small>Anfahrlisten Amazon XLSX</small>'
    )
    return f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Control ore · FICO Control</title>
<style>
*{{box-sizing:border-box}}:root{{--ink:#17212b;--muted:#667085;--line:#d8dde3;--safe:#14804a;--warn:#d98b12;--critical:#d13b2e}}
body{{margin:0;background:#f4f6f8;color:var(--ink);font-family:Arial,Helvetica,sans-serif}}
.wrap{{width:min(96%,1280px);margin:28px auto 60px}}.topbar{{display:flex;justify-content:space-between;align-items:flex-start;gap:18px;margin-bottom:22px}}
.brand{{font-size:13px;font-weight:900;letter-spacing:2px}}h1{{font-size:38px;margin:10px 0 6px}}.subtitle{{color:var(--muted);font-size:13px}}
.actions{{display:flex;gap:9px;flex-wrap:wrap}}.btn{{display:inline-flex;align-items:center;justify-content:center;text-decoration:none;border:0;border-radius:10px;padding:12px 15px;font-weight:800;cursor:pointer}}
.btn-dark{{background:var(--ink);color:#fff}}.btn-light{{background:#fff;color:var(--ink);border:1px solid var(--line)}}
.hero{{display:grid;grid-template-columns:1fr auto;gap:22px;align-items:center;background:linear-gradient(120deg,#0d4f6b,#177e9c);color:#fff;border-radius:18px;padding:28px;margin-bottom:16px}}
.hero h2{{margin:0 0 8px;font-size:27px}}.hero p{{margin:0;color:#d5edf3;line-height:1.55}}.upload{{display:flex;align-items:center;gap:10px;flex-wrap:wrap;background:#fff;padding:12px;border-radius:12px}}
.upload input{{max-width:320px;color:var(--ink)}}.file-row{{display:flex;gap:10px;align-items:center;background:#fff;border:1px solid #e4e7ec;border-radius:14px;padding:15px 18px;margin-bottom:16px}}
.file-row i{{width:10px;height:10px;border-radius:50%;background:{'#14804a' if results else '#98a2b3'}}}.file-row div{{display:flex;flex-direction:column;gap:3px}}.file-row small{{color:var(--muted)}}
.error{{display:flex;flex-direction:column;gap:4px;background:#fff0f0;color:#9f2f27;border-left:4px solid var(--critical);padding:14px 16px;margin-bottom:15px}}
.controls{{display:grid;grid-template-columns:220px 1fr auto;gap:12px;align-items:end;margin-bottom:16px}}label{{display:flex;flex-direction:column;gap:6px;color:var(--muted);font-size:12px;font-weight:800}}
select,input[type=search]{{height:45px;border:1px solid var(--line);border-radius:9px;background:#fff;padding:0 12px;font-size:14px}}.export{{height:45px}}
.stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:16px}}.stat{{display:grid;grid-template-columns:1fr auto;align-items:center;background:#fff;border-radius:15px;padding:19px;border:1px solid #e4e7ec}}
.stat span{{color:var(--muted);font-size:13px;font-weight:800}}.stat strong{{grid-row:1/3;grid-column:2;font-size:36px}}.stat small{{color:#98a2b3;margin-top:5px}}.stat.warn{{border-top:4px solid var(--warn)}}.stat.warn strong{{color:#ae6e0a}}.stat.critical{{border-top:4px solid var(--critical)}}.stat.critical strong{{color:var(--critical)}}
.panel{{background:#fff;border:1px solid #e4e7ec;border-radius:16px;overflow:hidden}}.panel-head{{display:flex;justify-content:space-between;gap:18px;align-items:center;padding:19px;border-bottom:1px solid #e8ebef}}.panel-head h2{{margin:0 0 5px;font-size:20px}}.panel-head p{{margin:0;color:var(--muted);font-size:12px}}
.legend{{display:flex;gap:13px;color:var(--muted);font-size:11px;white-space:nowrap}}.legend span{{display:flex;gap:5px;align-items:center}}.legend i{{width:8px;height:8px;border-radius:50%}}
.dot-safe{{background:var(--safe)}}.dot-warn{{background:var(--warn)}}.dot-critical{{background:var(--critical)}}.table-wrap{{overflow-x:auto}}table{{width:100%;border-collapse:collapse;min-width:820px}}
th,td{{text-align:left;padding:14px 15px;border-bottom:1px solid #edf0f2}}th{{font-size:11px;color:var(--muted);background:#fafafa;text-transform:uppercase;letter-spacing:.5px}}td strong,td small{{display:block}}td small{{color:#98a2b3;margin-top:4px}}
tr.warning{{background:#fffaf0}}tr.critical{{background:#fff1f1}}tr.warning td:first-child{{border-left:4px solid var(--warn)}}tr.critical td:first-child{{border-left:4px solid var(--critical)}}.hours{{font-size:18px;font-weight:900}}
.badge{{display:inline-flex;align-items:center;gap:6px;border-radius:999px;padding:6px 9px;font-size:11px;font-weight:900}}.badge.safe{{background:#e9f8ef;color:var(--safe)}}.badge.warning{{background:#fff1d6;color:#9a6208}}.badge.critical{{background:#fde2e1;color:#b42318}}.empty{{padding:38px;text-align:center;color:var(--muted)}}
@media(max-width:850px){{.topbar{{display:block}}.actions{{margin-top:15px}}.hero{{grid-template-columns:1fr}}.controls{{grid-template-columns:1fr}}.stats{{grid-template-columns:1fr}}.panel-head{{align-items:flex-start;flex-direction:column}}}}
</style>
</head>
<body><main class="wrap">
<div class="topbar"><div><div class="brand">FICO CONTROL</div><h1>Control ore șoferi</h1><div class="subtitle">Verificare zilnică după Anmelden și Abmelden · limita maximă 10h 30m</div></div>
<div class="actions"><a class="btn btn-light" href="/admin">FICO Dashboard</a><a class="btn btn-light" href="/admin/mentor">Mentor Check</a><a class="btn btn-light" href="/admin/pod-ccc">POD & CCC</a><a class="btn btn-light" href="/admin/concessions">Concesii</a><a class="btn btn-light" href="/admin/owner">Owner</a></div></div>
{error_html}
<section class="hero"><div><h2>Încarcă Anfahrlisten Amazon</h2><p>Șoferii sunt grupați după Transporter-ID. Toate rutele sunt adunate, iar intervalele suprapuse nu sunt numărate de două ori.</p></div>
<form class="upload" method="post" action="/admin/hours" enctype="multipart/form-data"><input type="file" name="report_file" accept=".xlsx" required><button class="btn btn-dark" type="submit">Verifică orele</button></form></section>
<div class="file-row"><i></i><div>{file_status}</div></div>
<section class="controls"><label>Data verificării<select id="dateSelect"></select></label><label>Caută șofer<input id="searchInput" type="search" placeholder="Scrie un nume..."></label><button class="btn btn-light export" id="exportBtn" type="button">Descarcă CSV</button></section>
<section class="stats"><div class="stat"><div><span>Șoferi verificați</span><small>cu ore reale</small></div><strong id="totalCount">0</strong></div><div class="stat warn"><div><span>Peste 10 ore</span><small>necesită atenție</small></div><strong id="warningCount">0</strong></div><div class="stat critical"><div><span>10h 30m sau mai mult</span><small>limită critică</small></div><strong id="criticalCount">0</strong></div></section>
<section class="panel"><div class="panel-head"><div><h2 id="tableTitle">Ore lucrate</h2><p>Ora Germaniei · fișierele nu sunt păstrate</p></div><div class="legend"><span><i class="dot-safe"></i>≤10h</span><span><i class="dot-warn"></i>10h01–10h29</span><span><i class="dot-critical"></i>≥10h30</span></div></div><div class="table-wrap"><table><thead><tr><th>Șofer</th><th>Interval real</th><th>Blocuri</th><th>Total lucrat</th><th>Status</th></tr></thead><tbody id="hoursBody"></tbody></table></div></section>
</main><script>
const reportData={payload_json};
const dateSelect=document.getElementById('dateSelect'), searchInput=document.getElementById('searchInput'), body=document.getElementById('hoursBody');
const dates=[...new Set(reportData.map(r=>r.date))].sort();
for(const d of dates){{const o=document.createElement('option');o.value=d;o.textContent=d.split('-').reverse().join('.');dateSelect.appendChild(o)}}
if(dates.length) dateSelect.value=dates[dates.length-1]; else {{const o=document.createElement('option');o.textContent='Nicio dată';dateSelect.appendChild(o);dateSelect.disabled=true}}
function duration(m){{return Math.floor(m/60)+'h '+String(m%60).padStart(2,'0')+'m'}}
function escapeHtml(v){{return String(v??'').replace(/[&<>\"']/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#039;'}}[c]))}}
function statusLabel(s){{return s==='critical'?'‼ LIMITĂ ATINSĂ':s==='warning'?'! ATENȚIE':'✓ ÎN REGULĂ'}}
function currentRows(){{const q=searchInput.value.trim().toLocaleLowerCase('ro');return reportData.filter(r=>r.date===dateSelect.value&&r.name.toLocaleLowerCase('ro').includes(q))}}
function render(){{const rows=currentRows();document.getElementById('totalCount').textContent=rows.length;document.getElementById('warningCount').textContent=rows.filter(r=>r.status==='warning').length;document.getElementById('criticalCount').textContent=rows.filter(r=>r.status==='critical').length;document.getElementById('tableTitle').textContent='Ore lucrate · '+(dateSelect.value?dateSelect.value.split('-').reverse().join('.'):'—');
body.innerHTML=rows.length?rows.map(r=>`<tr class="${{r.status}}"><td><strong>${{escapeHtml(r.name)}}</strong><small>${{escapeHtml(r.routes.join(' · '))}}</small></td><td>${{r.start}} — ${{r.end}}</td><td>${{r.blocks}}</td><td class="hours" title="${{r.seconds}} secunde">${{duration(r.minutes)}}</td><td><span class="badge ${{r.status}}">${{statusLabel(r.status)}}</span></td></tr>`).join(''):'<tr><td class="empty" colspan="5">Încarcă raportul pentru a vedea orele lucrate.</td></tr>'}}
dateSelect.addEventListener('change',render);searchInput.addEventListener('input',render);render();
document.getElementById('exportBtn').addEventListener('click',()=>{{const rows=currentRows();if(!rows.length)return;const lines=[['Data','Șofer','Început','Sfârșit','Ore lucrate','Status','Rute'],...rows.map(r=>[r.date,r.name,r.start,r.end,duration(r.minutes),statusLabel(r.status),r.routes.join(' / ')])];const csv='\uFEFF'+lines.map(row=>row.map(v=>'"'+String(v).replaceAll('"','""')+'"').join(',')).join('\\n');const url=URL.createObjectURL(new Blob([csv],{{type:'text/csv;charset=utf-8'}}));const a=document.createElement('a');a.href=url;a.download='ore-soferi-'+dateSelect.value+'.csv';a.click();URL.revokeObjectURL(url)}});
</script></body></html>
"""


@app.get("/admin/hours", response_class=HTMLResponse)
def hours_control_page():
    return HTMLResponse(hours_control_html())


@app.post("/admin/hours", response_class=HTMLResponse)
async def hours_control_upload(report_file: UploadFile = File(...)):
    try:
        payload = await report_file.read(HOURS_MAX_UPLOAD_BYTES + 1)
        results = read_hours_report(report_file.filename or "", payload)
        if not results:
            raise ValueError("Raportul nu conține ore reale valide.")
        return HTMLResponse(hours_control_html(results, report_file.filename or "Raport"))
    except ValueError as exc:
        return HTMLResponse(
            hours_control_html(filename=report_file.filename or "", error=str(exc)),
            status_code=400
        )
    finally:
        await report_file.close()


POD_CCC_MAX_FILE_BYTES = 12 * 1024 * 1024


def clean_driver_display_name(value):
    name = re.sub(r"\s+", " ", str(value or "")).strip()
    return name.split("•", 1)[0].strip()


def extract_transporter_mapping(xlsx_bytes: bytes):
    workbook = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=True
    )
    mapping = {}
    try:
        for sheet in workbook.worksheets:
            header_row = None
            name_col = None
            id_col = None
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True),
                start=1
            ):
                headers = [re.sub(r"\s+", " ", str(value or "")).strip().casefold() for value in row]
                for index, header in enumerate(headers):
                    if header in {"name des mitarbeiters", "driver name", "name"}:
                        name_col = index
                    if header in {"transporter-id", "transporter id", "associate id"}:
                        id_col = index
                if name_col is not None and id_col is not None:
                    header_row = row_number
                    break
            if header_row is None:
                continue
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if name_col >= len(row) or id_col >= len(row):
                    continue
                driver_id = re.sub(r"\s+", "", str(row[id_col] or "")).upper()
                driver_name = clean_driver_display_name(row[name_col])
                if driver_id and driver_name and "GESAMT" not in driver_id:
                    mapping[driver_id] = driver_name
            if mapping:
                break
    finally:
        workbook.close()
    if not mapping:
        raise ValueError("Nu am găsit coloanele Name des Mitarbeiters și Transporter-ID în plan.")
    return mapping


def decode_amazon_csv(csv_bytes: bytes):
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            decoded = csv_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("Fișierul CSV nu poate fi citit.")
    try:
        dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
    rows = list(reader)
    if not reader.fieldnames or "Transporter ID" not in reader.fieldnames:
        raise ValueError("Raportul nu conține coloana Transporter ID.")
    return rows, reader.fieldnames


def parse_report_date(filename: str):
    match = re.search(r"(20\d{2})[-_](\d{2})[-_](\d{2})", filename or "")
    return "-".join(match.groups()) if match else date.today().isoformat()


def group_quality_issues(rows, mapping, report_kind):
    expected = "POD Audit" if report_kind == "POD" else "Call Duration (Seconds)"
    if rows and expected not in rows[0]:
        raise ValueError(f"Fișierul selectat nu pare să fie raport {report_kind}.")
    groups = {}
    unmapped_ids = set()
    for row in rows:
        driver_id = re.sub(r"\s+", "", str(row.get("Transporter ID") or "")).upper()
        if not driver_id:
            continue
        driver_name = mapping.get(driver_id)
        if not driver_name:
            driver_name = f"ID necunoscut: {driver_id}"
            unmapped_ids.add(driver_id)
        group = groups.setdefault(driver_id, {
            "id": driver_id,
            "name": driver_name,
            "items": []
        })
        group["items"].append({key: str(value or "").replace("<br/>", " / ").strip() for key, value in row.items()})
    ordered = sorted(groups.values(), key=lambda group: (-len(group["items"]), group["name"].casefold()))
    return ordered, sorted(unmapped_ids)


def build_quality_workbook(report_kind, report_date, groups):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = report_kind
    sheet.sheet_view.showGridLines = False
    if report_kind == "POD":
        headers = ["Date", "Tracking ID", "Driver Name", "Delivery/Attempt Reason", "POD Audit", "Total Cases"]
        detail_keys = ["Tracking ID", "Delivery/Attempt Reason", "POD Audit"]
        widths = [15, 22, 32, 46, 32, 15]
    else:
        headers = ["Date", "Tracking ID", "Driver Name", "Delivery/Attempt Reason", "CC Type", "Call Duration (Seconds)", "Total Cases"]
        detail_keys = ["Tracking ID", "Delivery/Attempt Reason", "CC Type", "Call Duration (Seconds)"]
        widths = [15, 22, 32, 54, 24, 24, 15]
    last_col = len(headers)
    title_green = "DCEED6"
    header_green = "EDF6E9"
    high_fill = "FDE8E7"
    single_fill = "FFF8E6"
    ink = "17212B"
    grid = openpyxl.styles.Side(style="thin", color="4B5563")
    report_day = date.fromisoformat(report_date)
    report_day_text = report_day.strftime("%d.%m.%Y")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = sheet.cell(1, 1, f"{report_kind} – {report_day.strftime('%d.%m.%Y')}")
    title.font = openpyxl.styles.Font(name="Arial", size=17, bold=True, color=ink)
    title.fill = openpyxl.styles.PatternFill("solid", fgColor=title_green)
    title.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 34
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(2, col, header)
        cell.font = openpyxl.styles.Font(name="Arial", size=11, bold=True, color=ink)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=header_green)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = openpyxl.styles.Border(left=grid, right=grid, top=grid, bottom=grid)
    sheet.row_dimensions[2].height = 27

    current_row = 3
    for group in groups:
        total_cases = len(group["items"])
        row_fill = high_fill if total_cases >= 2 else single_fill
        for item in group["items"]:
            details = [item.get(key, "") for key in detail_keys]
            values = [report_day_text, details[0], group["name"], *details[1:], total_cases]
            for col, value in enumerate(values, start=1):
                cell = sheet.cell(current_row, col, value)
                cell.font = openpyxl.styles.Font(name="Arial", size=10, bold=col in {3, last_col}, color=ink)
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=row_fill)
                cell.border = openpyxl.styles.Border(left=grid, right=grid, top=grid, bottom=grid)
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal="center" if col in {1, last_col} else "left",
                    vertical="center",
                    wrap_text=True
                )
            sheet.row_dimensions[current_row].height = 29
            current_row += 1
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(last_col)}{max(2, current_row - 1)}"
    sheet.print_title_rows = "1:2"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def pod_ccc_html(processed=None, error=""):
    processed = processed or {}
    error_html = f'<div class="pc-error"><strong>Eroare:</strong> {html.escape(error)}</div>' if error else ""
    result_html = ""
    if processed:
        cards = []
        downloads = []
        driver_rows = []
        for kind in ("POD", "CCC"):
            info = processed[kind]
            cases = sum(len(group["items"]) for group in info["groups"])
            alert = sum(1 for group in info["groups"] if len(group["items"]) >= 2)
            cards.append(f'<div class="pc-stat"><span>{kind}</span><strong>{cases}</strong><small>{len(info["groups"])} șoferi · {alert} cu 2+ cazuri</small></div>')
            downloads.append(f'<a class="pc-btn pc-primary" download="{kind}_{info["date"]}.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{info["xlsx"]}">Descarcă Excel {kind}</a>')
            for group in info["groups"]:
                count = len(group["items"])
                driver_rows.append(f'<tr class="{"danger" if count >= 2 else "single"}"><td>{kind}</td><td><strong>{html.escape(group["name"])}</strong><small>{html.escape(group["id"])}</small></td><td>{count}</td><td>{"ATENȚIE" if count >= 2 else "1 caz"}</td></tr>')
        unmapped = sorted(set(processed["POD"]["unmapped"] + processed["CCC"]["unmapped"]))
        unmapped_html = f'<div class="pc-warning">ID-uri fără nume: {html.escape(", ".join(unmapped))}</div>' if unmapped else ""
        result_html = f'<section class="pc-results"><div class="pc-stats">{"".join(cards)}</div><div class="pc-downloads">{"".join(downloads)}</div>{unmapped_html}<div class="pc-table"><table><thead><tr><th>Raport</th><th>Șofer</th><th>Cazuri</th><th>Status</th></tr></thead><tbody>{"".join(driver_rows)}</tbody></table></div></section>'
    return f'''<!doctype html><html lang="ro"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>POD & CCC · FICO Control</title>
<style>*{{box-sizing:border-box}}body{{margin:0;background:#f4f6f8;color:#17212b;font-family:Arial,sans-serif}}.pc-wrap{{width:min(96%,1180px);margin:28px auto 60px}}.pc-top{{display:flex;justify-content:space-between;gap:20px;align-items:flex-start}}.pc-brand{{font-size:12px;font-weight:900;letter-spacing:2px}}h1{{font-size:38px;margin:9px 0 5px}}.pc-sub{{color:#667085}}.pc-actions{{display:flex;gap:8px;flex-wrap:wrap}}.pc-btn{{display:inline-flex;text-decoration:none;border:1px solid #d8dde3;border-radius:10px;padding:12px 15px;background:#fff;color:#17212b;font-weight:800;cursor:pointer}}.pc-primary{{background:#17212b;color:#fff;border-color:#17212b}}.pc-hero{{margin-top:22px;background:linear-gradient(120deg,#0d4f6b,#177e9c);color:#fff;border-radius:18px;padding:26px}}.pc-hero h2{{margin:0 0 8px}}.pc-hero p{{color:#d5edf3;margin:0 0 20px}}.pc-form{{display:grid;grid-template-columns:repeat(3,1fr) auto;gap:12px;align-items:end}}.pc-upload{{background:#fff;color:#17212b;border-radius:12px;padding:12px}}.pc-upload label{{display:block;font-size:12px;font-weight:900;margin-bottom:7px}}.pc-upload input{{max-width:100%}}.pc-error,.pc-warning{{margin-top:16px;padding:14px 16px;background:#fff0f0;border-left:4px solid #d92d20;color:#9f2f27}}.pc-results{{margin-top:18px}}.pc-stats{{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}}.pc-stat{{background:#fff;border:1px solid #e4e7ec;border-radius:15px;padding:18px;display:grid;grid-template-columns:1fr auto}}.pc-stat span{{font-weight:900}}.pc-stat strong{{font-size:34px;grid-row:1/3;grid-column:2}}.pc-stat small{{color:#667085;margin-top:6px}}.pc-downloads{{display:flex;gap:10px;margin:15px 0}}.pc-table{{background:#fff;border:1px solid #e4e7ec;border-radius:15px;overflow:hidden}}table{{width:100%;border-collapse:collapse}}th,td{{padding:13px 15px;text-align:left;border-bottom:1px solid #edf0f2}}th{{font-size:11px;color:#667085;background:#fafafa;text-transform:uppercase}}td small{{display:block;color:#98a2b3;margin-top:3px}}tr.danger{{background:#fff0f0}}tr.danger td:first-child{{border-left:4px solid #d92d20}}tr.single{{background:#fffaf0}}@media(max-width:850px){{.pc-top{{display:block}}.pc-actions{{margin-top:14px}}.pc-form{{grid-template-columns:1fr}}.pc-stats{{grid-template-columns:1fr}}}}</style></head>
<body><main class="pc-wrap"><div class="pc-top"><div><div class="pc-brand">FICO CONTROL</div><h1>POD & CCC</h1><div class="pc-sub">Înlocuire automată Transporter ID cu numele real și rapoarte profesionale</div></div><div class="pc-actions"><a class="pc-btn" href="/admin">FICO Dashboard</a><a class="pc-btn" href="/admin/hours">Control ore</a><a class="pc-btn" href="/admin/concessions">Concesii</a></div></div>{error_html}<section class="pc-hero"><h2>Încarcă cele trei fișiere Amazon</h2><p>Planul săptămânal furnizează numele reale. POD și CCC sunt procesate separat.</p><form class="pc-form" method="post" action="/admin/pod-ccc" enctype="multipart/form-data"><div class="pc-upload"><label>1. Plan săptămânal ID–nume</label><input type="file" name="mapping_file" accept=".xlsx" required></div><div class="pc-upload"><label>2. Raport POD</label><input type="file" name="pod_file" accept=".csv" required></div><div class="pc-upload"><label>3. Raport CCC</label><input type="file" name="ccc_file" accept=".csv" required></div><button class="pc-btn pc-primary" type="submit">Generează rapoartele</button></form></section>{result_html}</main></body></html>'''


@app.get("/admin/pod-ccc", response_class=HTMLResponse)
def pod_ccc_page():
    return HTMLResponse(pod_ccc_html())


@app.post("/admin/pod-ccc", response_class=HTMLResponse)
async def pod_ccc_upload(
    mapping_file: UploadFile = File(...),
    pod_file: UploadFile = File(...),
    ccc_file: UploadFile = File(...)
):
    files = (mapping_file, pod_file, ccc_file)
    try:
        mapping_raw, pod_raw, ccc_raw = [
            await upload.read(POD_CCC_MAX_FILE_BYTES + 1) for upload in files
        ]
        if any(len(raw) > POD_CCC_MAX_FILE_BYTES for raw in (mapping_raw, pod_raw, ccc_raw)):
            raise ValueError("Unul dintre fișiere depășește limita de 12 MB.")
        if not (mapping_file.filename or "").lower().endswith(".xlsx"):
            raise ValueError("Lista ID–nume trebuie să fie un fișier XLSX.")
        if not (pod_file.filename or "").lower().endswith(".csv") or not (ccc_file.filename or "").lower().endswith(".csv"):
            raise ValueError("Rapoartele POD și CCC trebuie să fie fișiere CSV.")
        mapping = extract_transporter_mapping(mapping_raw)
        pod_rows, _ = decode_amazon_csv(pod_raw)
        ccc_rows, _ = decode_amazon_csv(ccc_raw)
        pod_groups, pod_unmapped = group_quality_issues(pod_rows, mapping, "POD")
        ccc_groups, ccc_unmapped = group_quality_issues(ccc_rows, mapping, "CCC")
        if not pod_groups and not ccc_groups:
            raise ValueError("Rapoartele nu conțin cazuri POD sau CCC.")
        processed = {}
        for kind, filename, groups, unmapped in (
            ("POD", pod_file.filename or "", pod_groups, pod_unmapped),
            ("CCC", ccc_file.filename or "", ccc_groups, ccc_unmapped)
        ):
            report_date = parse_report_date(filename)
            workbook_bytes = build_quality_workbook(kind, report_date, groups)
            processed[kind] = {
                "date": report_date,
                "groups": groups,
                "unmapped": unmapped,
                "xlsx": base64.b64encode(workbook_bytes).decode("ascii")
            }
        return HTMLResponse(pod_ccc_html(processed))
    except (ValueError, openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile) as exc:
        return HTMLResponse(pod_ccc_html(error=str(exc)), status_code=400)
    finally:
        for upload in files:
            await upload.close()


CONCESSIONS_DNR_COLUMN = "Pakete, die geliefert aber nicht empfangen wurden (DNR)"


def parse_integer_metric(value):
    clean = re.sub(r"[^0-9-]", "", str(value or ""))
    return int(clean) if clean not in {"", "-"} else 0


def parse_concessions_csv(csv_bytes: bytes):
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            decoded = csv_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("Fișierul de concesii nu poate fi citit.")
    try:
        dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
    required = {"Woche", "Name des Zustellenden", CONCESSIONS_DNR_COLUMN}
    if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
        raise ValueError("Raportul nu conține coloanele necesare pentru concesii.")
    parsed = []
    week_value = ""
    for source_index, row in enumerate(reader):
        week_value = week_value or str(row.get("Woche") or "").strip()
        driver_name = clean_driver_display_name(row.get("Name des Zustellenden"))
        dnr_count = parse_integer_metric(row.get(CONCESSIONS_DNR_COLUMN))
        dnr_dpmo = parse_integer_metric(row.get("DNR DPMO"))
        if driver_name and dnr_count > 0:
            parsed.append({
                "name": driver_name,
                "count": dnr_count,
                "dpmo": dnr_dpmo,
                "source_index": source_index
            })
    if not parsed:
        raise ValueError("Raportul nu conține nicio concesie DNR.")
    parsed.sort(key=lambda item: (-item["count"], item["dpmo"], item["source_index"]))
    match = re.search(r"(20\d{2})\D+(\d{1,2})", week_value)
    if not match:
        raise ValueError("Nu am putut identifica săptămâna raportului.")
    return parsed, int(match.group(1)), int(match.group(2))


def concessions_week_label(year, week):
    monday = date.fromisocalendar(year, week, 1)
    start = monday - timedelta(days=1)
    end = start + timedelta(days=6)
    months = ["Jan.", "Feb.", "Mar.", "Apr.", "May", "Jun.", "Jul.", "Aug.", "Sep.", "Oct.", "Nov.", "Dec."]
    if start.month == end.month:
        period = f"{months[start.month - 1]} {start.day}-{months[end.month - 1]} {end.day}"
        filename_period = f"{months[start.month - 1].replace('.', '')}{start.day}-{months[end.month - 1].replace('.', '')}{end.day}"
    else:
        period = f"{months[start.month - 1]} {start.day}-{months[end.month - 1]} {end.day}"
        filename_period = f"{months[start.month - 1].replace('.', '')}{start.day}-{months[end.month - 1].replace('.', '')}{end.day}"
    return period, filename_period


def build_concessions_workbook(rows, year, week):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"KW{week}"
    sheet.sheet_view.showGridLines = False
    period, _ = concessions_week_label(year, week)
    headers = ["NR.", "Fahrer name", f"Woche {week}, {period}"]
    cyan = "22B5E5"
    red = "FF0000"
    gold = "D6A900"
    yellow = "FFDB6E"
    white = "FFFFFF"
    black = "000000"
    grid = openpyxl.styles.Side(style="thin", color="A0A0A0")
    outer = openpyxl.styles.Side(style="thin", color="777777")
    border = openpyxl.styles.Border(left=outer, right=outer, top=grid, bottom=grid)

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col, header)
        cell.font = openpyxl.styles.Font(name="Carlito", size=11, bold=True, color=black)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=cyan)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.row_dimensions[1].height = 27

    for row_number, item in enumerate(rows, start=2):
        rank = row_number - 1
        count = item["count"]
        values = [rank, item["name"], count]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, col, value)
            cell.font = openpyxl.styles.Font(name="Carlito", size=11, color=black)
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center" if col != 2 else "left",
                vertical="center"
            )
            cell.border = border
        if count >= 2:
            sheet.cell(row_number, 2).fill = openpyxl.styles.PatternFill("solid", fgColor=red)
            sheet.cell(row_number, 2).font = openpyxl.styles.Font(name="Carlito", size=11, bold=True, color=white)
            sheet.cell(row_number, 3).font = openpyxl.styles.Font(name="Carlito", size=11, bold=True, color=red)
        else:
            sheet.cell(row_number, 2).fill = openpyxl.styles.PatternFill("solid", fgColor=yellow)
            sheet.cell(row_number, 2).font = openpyxl.styles.Font(name="Carlito", size=11, bold=True, color=gold)
            sheet.cell(row_number, 3).font = openpyxl.styles.Font(name="Carlito", size=11, bold=True, color=gold)
        sheet.row_dimensions[row_number].height = 22

    total_row = len(rows) + 2
    sheet.cell(total_row, 1, "")
    sheet.cell(total_row, 2, "TOTAL")
    sheet.cell(total_row, 3, sum(item["count"] for item in rows))
    for col in range(1, 4):
        cell = sheet.cell(total_row, col)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=cyan)
        cell.font = openpyxl.styles.Font(name="Carlito", size=11, bold=True, color=black)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.row_dimensions[total_row].height = 25
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 36
    sheet.column_dimensions["C"].width = 24
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{len(rows) + 1}"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def concessions_html(result=None, error=""):
    result = result or {}
    error_html = f'<div class="cn-error"><strong>Eroare:</strong> {html.escape(error)}</div>' if error else ""
    result_html = ""
    if result:
        result_html = f'''<section class="cn-result"><div><span>Total concesii DNR</span><strong>{result["total"]}</strong><small>{result["drivers"]} șoferi · KW {result["week"]}</small></div><a class="cn-btn cn-primary" download="{html.escape(result["filename"])}" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{result["xlsx"]}">Descarcă Excel Concesii</a></section>'''
    return f'''<!doctype html><html lang="ro"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Concesii · FICO Control</title><style>*{{box-sizing:border-box}}body{{margin:0;background:#f4f6f8;color:#17212b;font-family:Arial,sans-serif}}.cn-wrap{{width:min(96%,1050px);margin:28px auto 60px}}.cn-top{{display:flex;justify-content:space-between;gap:18px;align-items:flex-start}}.cn-brand{{font-size:12px;font-weight:900;letter-spacing:2px}}h1{{font-size:38px;margin:9px 0 5px}}.cn-sub{{color:#667085}}.cn-actions{{display:flex;gap:8px;flex-wrap:wrap}}.cn-btn{{display:inline-flex;text-decoration:none;border:1px solid #d8dde3;border-radius:10px;padding:12px 15px;background:#fff;color:#17212b;font-weight:800;cursor:pointer}}.cn-primary{{background:#17212b;color:#fff;border-color:#17212b}}.cn-hero{{margin-top:22px;background:linear-gradient(120deg,#0d4f6b,#177e9c);color:#fff;border-radius:18px;padding:28px}}.cn-hero h2{{margin:0 0 8px}}.cn-hero p{{color:#d5edf3;margin:0 0 20px}}.cn-form{{display:flex;gap:12px;align-items:center;flex-wrap:wrap;background:#fff;padding:13px;border-radius:12px}}.cn-form input{{color:#17212b;flex:1;min-width:250px}}.cn-error{{margin-top:16px;padding:14px 16px;background:#fff0f0;border-left:4px solid #d92d20;color:#9f2f27}}.cn-result{{margin-top:18px;background:#fff;border:1px solid #e4e7ec;border-radius:16px;padding:20px;display:flex;justify-content:space-between;align-items:center;gap:18px}}.cn-result div{{display:grid;grid-template-columns:1fr auto}}.cn-result span{{font-weight:900}}.cn-result strong{{font-size:36px;grid-row:1/3;grid-column:2;margin-left:28px}}.cn-result small{{color:#667085;margin-top:6px}}@media(max-width:700px){{.cn-top,.cn-result{{display:block}}.cn-actions{{margin-top:14px}}.cn-result .cn-btn{{margin-top:15px}}}}</style></head><body><main class="cn-wrap"><div class="cn-top"><div><div class="cn-brand">FICO CONTROL</div><h1>Concesii</h1><div class="cn-sub">Raport săptămânal DNR ordonat și formatat automat</div></div><div class="cn-actions"><a class="cn-btn" href="/admin">FICO Dashboard</a><a class="cn-btn" href="/admin/pod-ccc">POD & CCC</a></div></div>{error_html}<section class="cn-hero"><h2>Încarcă raportul Amazon Concessions</h2><p>Șoferii cu cele mai multe concesii apar primii. Raportul final respectă modelul KW.</p><form class="cn-form" method="post" action="/admin/concessions" enctype="multipart/form-data"><input type="file" name="concessions_file" accept=".csv" required><button class="cn-btn cn-primary" type="submit">Generează Excel</button></form></section>{result_html}</main></body></html>'''


@app.get("/admin/concessions", response_class=HTMLResponse)
def concessions_page():
    return HTMLResponse(concessions_html())


@app.post("/admin/concessions", response_class=HTMLResponse)
async def concessions_upload(concessions_file: UploadFile = File(...)):
    try:
        if not (concessions_file.filename or "").lower().endswith(".csv"):
            raise ValueError("Raportul de concesii trebuie să fie un fișier CSV.")
        raw = await concessions_file.read(POD_CCC_MAX_FILE_BYTES + 1)
        if len(raw) > POD_CCC_MAX_FILE_BYTES:
            raise ValueError("Fișierul depășește limita de 12 MB.")
        rows, year, week = parse_concessions_csv(raw)
        workbook_bytes = build_concessions_workbook(rows, year, week)
        _, filename_period = concessions_week_label(year, week)
        result = {
            "total": sum(item["count"] for item in rows),
            "drivers": len(rows),
            "week": week,
            "filename": f"KW{week}_Concessions_{filename_period}.xlsx",
            "xlsx": base64.b64encode(workbook_bytes).decode("ascii")
        }
        return HTMLResponse(concessions_html(result))
    except ValueError as exc:
        return HTMLResponse(concessions_html(error=str(exc)), status_code=400)
    finally:
        await concessions_file.close()


@app.get("/admin/mentor", response_class=HTMLResponse)
def mentor_check_page(request: Request, d: str | None = None):
    selected = d or date.today().isoformat()

    conn = db()
    required_rows = conn.execute(
        "SELECT driver_name FROM mentor_required WHERE work_date=? ORDER BY driver_name",
        (selected,)
    ).fetchall()
    mentor_rows = conn.execute(
        """
        SELECT driver_name, first_connection_time
        FROM mentor_connected
        WHERE work_date=?
        ORDER BY driver_name
        """,
        (selected,)
    ).fetchall()
    conn.close()

    required_names = [row["driver_name"] for row in required_rows]
    mentor_names = [row["driver_name"] for row in mentor_rows]
    mentor_time_by_key = {
        mentor_name_key(row["driver_name"]): row["first_connection_time"]
        for row in mentor_rows
    }

    results = []
    for required_name in required_names:
        status, matched_name, confidence = mentor_compare(required_name, mentor_names)

        first_connection_time = None
        if matched_name and status != "missing":
            first_connection_time = mentor_time_by_key.get(
                mentor_name_key(matched_name)
            )

        results.append({
            "name": required_name,
            "status": status,
            "matched_name": matched_name,
            "confidence": confidence,
            "first_connection_time": first_connection_time
        })

    connected_count = sum(row["status"] == "connected" for row in results)
    review_count = sum(row["status"] == "review" for row in results)
    missing = [row for row in results if row["status"] == "missing"]

    table_rows = ""
    for row in results:
        if row["status"] == "connected":
            badge = '<span class="badge ok">Conectat</span>'
        elif row["status"] == "review":
            badge = '<span class="badge review">Verificare</span>'
        else:
            badge = '<span class="badge missing">Nu s-a conectat</span>'

        mentor_match = "—"
        if row["matched_name"] and row["status"] != "missing":
            mentor_match = html.escape(row["matched_name"])
            if row["confidence"] < 0.999:
                mentor_match += f' <span class="confidence">{round(row["confidence"] * 100)}%</span>'

        first_connection = (
            html.escape(row["first_connection_time"])
            if row.get("first_connection_time")
            else "—"
        )

        table_rows += f"""
        <tr>
          <td><strong>{html.escape(row['name'])}</strong></td>
          <td>{badge}</td>
          <td><strong>{first_connection}</strong></td>
          <td>{mentor_match}</td>
        </tr>
        """

    if not results:
        table_rows = '<tr><td class="empty" colspan="4">Încarcă lista Cortex pentru această zi.</td></tr>'

    missing_text = "\\n".join(row["name"] for row in missing).replace("`", "\\`")

    page = f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Mentor Check · FICO Control</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#f4f6f8;color:#17212b;font-family:Arial,sans-serif}}
.wrap{{max-width:1280px;margin:0 auto;padding:28px 18px 60px}}
.topbar{{display:flex;justify-content:space-between;align-items:center;gap:16px;margin-bottom:22px}}
.brand{{font-size:13px;font-weight:900;letter-spacing:2px}}
h1{{font-size:38px;margin:5px 0 0}}
.actions{{display:flex;gap:9px;align-items:center;flex-wrap:wrap}}
.btn{{border:0;border-radius:11px;padding:12px 16px;font-weight:800;text-decoration:none;cursor:pointer;display:inline-flex;align-items:center;justify-content:center}}
.btn-dark{{background:#17212b;color:#fff}}
.btn-light{{background:#fff;color:#17212b;border:1px solid #d8dde3}}
.date-row{{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap}}
.date-form{{display:flex;gap:9px;align-items:center}}
input[type=date],input[type=file]{{border:1px solid #d8dde3;border-radius:10px;padding:11px;background:#fff;max-width:100%}}
.stats{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:18px}}
.stat{{background:#fff;border:1px solid #e4e7ec;border-radius:16px;padding:18px}}
.stat strong{{display:block;font-size:31px;margin-bottom:4px}}
.stat span{{font-size:13px;color:#667085;font-weight:700}}
.uploads{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:18px}}
.panel{{background:#fff;border:1px solid #e4e7ec;border-radius:18px;padding:20px}}
.panel h2{{margin:0 0 7px;font-size:20px}}
.panel p{{margin:0 0 16px;color:#667085;font-size:13px;line-height:1.5}}
.upload-form{{display:flex;gap:10px;flex-wrap:wrap;align-items:center}}
.single-upload-action{{display:flex;justify-content:center;margin:-2px 0 22px}}
.single-upload-btn{{min-width:260px;padding:15px 28px;font-size:15px}}
.table-panel{{background:#fff;border:1px solid #e4e7ec;border-radius:18px;overflow:hidden}}
table{{width:100%;border-collapse:collapse}}
th,td{{padding:14px 16px;text-align:left;border-bottom:1px solid #eef0f2}}
th{{font-size:12px;color:#667085;background:#fafafa}}
.badge{{display:inline-block;padding:6px 10px;border-radius:999px;font-size:12px;font-weight:900}}
.ok{{background:#e9f8ef;color:#147a42}}
.missing{{background:#fff0f0;color:#b42318}}
.review{{background:#fff7e6;color:#9a6700}}
.confidence{{color:#667085;font-size:11px;font-weight:700}}
.empty{{padding:32px;text-align:center;color:#667085}}
@media(max-width:850px){{.stats{{grid-template-columns:1fr 1fr}}.uploads{{grid-template-columns:1fr}}.topbar{{display:block}}.actions{{margin-top:15px}}}}
@media(max-width:520px){{.stats{{grid-template-columns:1fr}}h1{{font-size:31px}}}}
</style>
</head>
<body>
<main class="wrap">
  <div class="topbar">
    <div>
      <div class="brand">FICO CONTROL</div>
      <h1>Mentor Check</h1>
      <div style="margin-top:7px;color:#667085;font-size:13px">Compară lista Cortex cu Mentor Shift Report</div>
    </div>
    <div class="actions">
      <a class="btn btn-light" href="/admin?d={selected}">FICO Dashboard</a>
      <a class="btn btn-light" href="/admin/hours">Control ore</a>
      <a class="btn btn-light" href="/admin/pod-ccc">POD & CCC</a>
      <a class="btn btn-light" href="/admin/concessions">Concesii</a>
      <a class="btn btn-light" href="/admin/owner">Owner</a>
      <button class="btn btn-dark" type="button" onclick="copyMissing()">Copiază șoferii lipsă</button>
    </div>
  </div>

  <div class="date-row">
    <form class="date-form" method="get" action="/admin/mentor">
      <input type="date" name="d" value="{selected}">
      <button class="btn btn-dark" type="submit">Afișează</button>
    </form>
    <strong>{selected}</strong>
  </div>

  <section class="stats">
    <div class="stat"><strong>{len(required_names)}</strong><span>Trebuie conectați</span></div>
    <div class="stat"><strong>{connected_count}</strong><span>Conectați în Mentor</span></div>
    <div class="stat"><strong>{len(missing)}</strong><span>Nu s-au conectat</span></div>
    <div class="stat"><strong>{review_count}</strong><span>Necesită verificare</span></div>
  </section>

  <form method="post" action="/admin/mentor/upload-both" enctype="multipart/form-data">
    <input type="hidden" name="work_date" value="{selected}">

    <section class="uploads">
      <div class="panel">
        <h2>1. Cortex</h2>
        <p>Lista tuturor șoferilor care lucrează. Pentru „Name des Fahrers” cu mai multe nume separate prin |, se ia numai primul șofer.</p>
        <div class="upload-form">
          <input type="file" name="cortex_file" accept=".xlsx" required>
        </div>
      </div>

      <div class="panel">
        <h2>2. Mentor Shift Report</h2>
        <p>Lista celor care s-au conectat. Dacă același șofer apare de mai multe ori, este considerat o singură persoană, iar „Prima conectare” este cea mai devreme valoare din Begin Route Time.</p>
        <div class="upload-form">
          <input type="file" name="mentor_file" accept=".xlsx" required>
        </div>
      </div>
    </section>

    <div class="single-upload-action">
      <button class="btn btn-dark single-upload-btn" type="submit">Verifică Mentor</button>
    </div>
  </form>

  <section class="table-panel">
    <table>
      <thead><tr><th>ȘOFER CORTEX</th><th>STATUS MENTOR</th><th>PRIMA CONECTARE</th><th>POTRIVIRE MENTOR</th></tr></thead>
      <tbody>{table_rows}</tbody>
    </table>
  </section>
</main>
<script>
async function copyMissing() {{
  const text = `{missing_text}`;
  try {{
    await navigator.clipboard.writeText(text);
    alert(text ? "Lista șoferilor lipsă a fost copiată." : "Nu există șoferi lipsă.");
  }} catch (e) {{
    alert("Nu am putut copia lista.");
  }}
}}
</script>
</body>
</html>
"""
    return HTMLResponse(page)


@app.post("/admin/upload")
async def upload_daily_list(
    work_date: str = Form(...),
    file: UploadFile = File(...)
):
    filename = (file.filename or "").lower()

    if not filename.endswith(".xlsx"):
        return RedirectResponse(
            f"/admin?d={work_date}&error=unsupported",
            status_code=303
        )

    raw = await file.read()

    try:
        names = extract_driver_names_from_xlsx(raw)
    except Exception:
        return RedirectResponse(
            f"/admin?d={work_date}&error=import",
            status_code=303
        )

    if not names:
        return RedirectResponse(
            f"/admin?d={work_date}&error=empty",
            status_code=303
        )

    conn = db()
    conn.execute("DELETE FROM daily_required WHERE work_date=?", (work_date,))

    for name in names:
        conn.execute("INSERT OR IGNORE INTO drivers(name) VALUES(?)", (name,))
        driver = conn.execute("SELECT id FROM drivers WHERE name=?", (name,)).fetchone()
        conn.execute(
            "INSERT OR IGNORE INTO daily_required(work_date, driver_id) VALUES(?,?)",
            (work_date, driver["id"])
        )

    # Persist metadata for this day's Excel list. The list itself is already
    # stored in daily_required by work_date and survives logout/redeploy.
    _daily_count_row = conn.execute(
        "SELECT COUNT(*) AS c FROM daily_required WHERE work_date=?",
        (work_date,)
    ).fetchone()
    _daily_count = int(_daily_count_row["c"] if _daily_count_row else 0)
    _source_filename = getattr(file, "filename", None)

    conn.execute(
        """
        INSERT INTO daily_list_imports(
            work_date, imported_at, source_filename, driver_count
        )
        VALUES(?,?,?,?)
        ON CONFLICT(work_date) DO UPDATE SET
            imported_at=excluded.imported_at,
            source_filename=excluded.source_filename,
            driver_count=excluded.driver_count
        """,
        (
            work_date,
            datetime.now(timezone.utc).isoformat(),
            _source_filename,
            _daily_count
        )
    )


    conn.commit()
    conn.close()

    return RedirectResponse(
        f"/admin?d={work_date}&imported={len(names)}",
        status_code=303
    )



@app.post("/submit")
async def submit_web(
    full_name: str = Form(...),
    fico_score: int = Form(...),
    proof: UploadFile = File(...)
):
    if fico_score < 300 or fico_score > 850:
        return RedirectResponse("/?error=invalid_score", status_code=303)

    today = date.today().isoformat()
    conn = db()

    driver, match_score, ambiguous = find_required_driver(conn, today, full_name)

    if ambiguous:
        conn.close()
        return RedirectResponse("/?error=ambiguous_name", status_code=303)

    if not driver:
        conn.close()
        return RedirectResponse("/?error=name_not_found", status_code=303)

    existing = conn.execute(
        "SELECT 1 FROM submissions WHERE work_date=? AND driver_id=?",
        (today, driver["id"])
    ).fetchone()

    if existing:
        conn.close()
        return RedirectResponse("/?error=already_sent", status_code=303)

    try:
        filename, original_name, raw = await save_proof_image(proof)
    except HTTPException as exc:
        conn.close()
        return RedirectResponse(f"/?error={exc.detail}", status_code=303)

    detected_score, ocr_state = detect_fico_from_image_bytes(
        raw,
        proof.content_type or "image/jpeg"
    )

    if detected_score is None:
        verification_status = "manual_review"
    elif detected_score == fico_score:
        verification_status = "verified"
    else:
        verification_status = "mismatch"

    conn.execute("""
        INSERT INTO submissions(
            work_date,
            driver_id,
            fico_score,
            submitted_at,
            entered_full_name,
            proof_filename,
            proof_original_name,
            detected_fico_score,
            verification_status,
            name_match_score
        ) VALUES(?,?,?,?,?,?,?,?,?,?)
    """, (
        today,
        driver["id"],
        fico_score,
        datetime.now().isoformat(timespec="seconds"),
        " ".join(full_name.strip().split()),
        filename,
        original_name,
        detected_score,
        verification_status,
        match_score
    ))

    conn.commit()
    conn.close()

    return RedirectResponse("/?success=1", status_code=303)


@app.get("/proof/{filename}")
def proof_image(filename: str):
    safe_name = os.path.basename(filename)

    if R2_ENABLED:
        try:
            client = r2_client()
            response = client.get_object(
                Bucket=R2_BUCKET_NAME,
                Key=safe_name
            )

            content_type = (
                response.get("ContentType")
                or "application/octet-stream"
            )
            body = response["Body"]

            return StreamingResponse(
                body.iter_chunks(chunk_size=1024 * 1024),
                media_type=content_type,
                headers={
                    "Cache-Control": "private, max-age=3600"
                }
            )
        except Exception as exc:
            error_code = ""
            try:
                error_code = exc.response.get("Error", {}).get("Code", "")
            except Exception:
                pass

            if error_code in ("NoSuchKey", "404", "NotFound"):
                raise HTTPException(
                    status_code=404,
                    detail="proof_not_found"
                )

            print(
                "R2_READ_ERROR:",
                type(exc).__name__,
                str(exc)[:500],
                flush=True
            )
            raise HTTPException(
                status_code=502,
                detail="proof_storage_error"
            )

    # Local fallback for development only.
    path = os.path.join(UPLOAD_DIR, safe_name)

    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="proof_not_found")

    return FileResponse(path)


@app.get("/", response_class=HTMLResponse)
def driver_page():
    today = date.today().isoformat()

    page = f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FICO Control</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;font-family:Arial,sans-serif;background:#f4f6f8;color:#17212b}}
.card{{width:min(92%,500px);margin:45px auto;background:#fff;padding:30px;border-radius:20px;box-shadow:0 12px 35px rgba(0,0,0,.08)}}
.topbar{{display:flex;align-items:center;justify-content:space-between;gap:16px}}
.brand{{font-size:13px;font-weight:800;letter-spacing:2px;white-space:nowrap}}
.languages{{display:flex;gap:5px;background:#f4f6f8;padding:4px;border-radius:10px}}
.lang-btn{{width:auto;margin:0;padding:7px 9px;border:0;border-radius:7px;background:transparent;color:#667085;font-size:12px;font-weight:800;cursor:pointer}}
.lang-btn.active{{background:#17212b;color:#fff}}
h1{{font-size:31px;margin:22px 0 8px}}
.muted{{color:#667085;line-height:1.45}}
label{{display:block;margin:20px 0 8px;font-weight:700}}
input,button{{width:100%;padding:14px;border-radius:11px;border:1px solid #d8dde3;font-size:16px}}
input[type=file]{{background:#fff}}
.proof-input-native{{position:absolute;left:-9999px;width:1px;height:1px;opacity:0}}
.proof-action{{display:block;width:100%;box-sizing:border-box;padding:14px 12px;border-radius:11px;border:1px solid #d8dde3;background:#fff;color:#17212b;font-size:15px;font-weight:800;cursor:pointer;text-align:center}}
.proof-action.primary{{background:#17212b;color:#fff;border-color:#17212b}}
.camera-label{{margin-top:2px}}
.proof-selected{{display:none;margin-top:10px;padding:10px 12px;border-radius:9px;background:#eaf7ef;color:#146c43;font-size:13px;font-weight:700;word-break:break-word}}
.proof-selected.show{{display:block}}
.submit{{margin-top:22px;background:#17212b;color:#fff;border:0;font-weight:800;cursor:pointer}}
.hint{{font-size:13px;color:#667085;margin-top:7px}}
.notice{{display:none;margin:18px 0 0;padding:12px 14px;border-radius:10px;font-size:14px;font-weight:700;line-height:1.4}}
.notice.success{{display:block;background:#eaf7ef;color:#146c43}}
.notice.error{{display:block;background:#fdeeee;color:#b42318}}
@media(max-width:420px){{
  .card{{padding:23px}}
  .topbar{{align-items:flex-start}}
  .brand{{font-size:12px}}
  .lang-btn{{padding:7px}}
}}
</style>
</head>
<body>
<main class="card">
<div class="topbar">
  <div class="brand">FICO CONTROL</div>
  <div class="languages" aria-label="Language">
    <button type="button" class="lang-btn" data-lang="ro">RO</button>
    <button type="button" class="lang-btn" data-lang="de">DE</button>
    <button type="button" class="lang-btn" data-lang="en">EN</button>
  </div>
</div>

<h1 data-i18n="title">Trimite scorul FICO</h1>
<p class="muted">
  {today}<br>
  <span data-i18n="intro">Încarcă o poză sau un screenshot clar în care scorul FICO este vizibil.</span>
</p>

<div id="notice" class="notice"></div>

<form action="/submit" method="post" enctype="multipart/form-data">

<label data-i18n="photoLabel">1. Poză / Screenshot FICO</label>

<label class="proof-action primary camera-label" for="proof" data-i18n="takePhoto">Fă o poză acum</label>
<input
  id="proof"
  class="proof-input-native"
  type="file"
  name="proof"
  accept="image/*"
  capture="environment"
  required
>
<div id="proofSelected" class="proof-selected"></div>
<div class="hint" data-i18n="photoHint">Fă o poză clară în care scorul FICO este vizibil · maximum 10 MB</div>

<label data-i18n="nameLabel">2. Numele complet</label>
<input id="fullName" type="text" name="full_name" autocomplete="name" placeholder="Prenume și nume complet" required>
<div class="hint" data-i18n="nameHint">Scrie numele exact așa cum apare în lista de lucru.</div>

<label data-i18n="scoreLabel">3. Scor FICO</label>
<input id="ficoScore" type="number" name="fico_score" min="300" max="850" inputmode="numeric" placeholder="ex. 850" required>

<button class="submit" type="submit" data-i18n="submit">Trimite scorul și dovada</button>
</form>
</main>

<script>
const translations = {{
  ro: {{
    title: "Trimite scorul FICO",
    intro: "Încarcă o poză sau un screenshot clar în care scorul FICO este vizibil.",
    photoLabel: "1. Poză / Screenshot FICO",
    takePhoto: "Fă o poză acum",
    photoSelected: "Poză selectată:",
    photoHint: "Fă o poză clară în care scorul FICO este vizibil · maximum 10 MB",
    nameLabel: "2. Numele complet",
    nameHint: "Scrie numele exact așa cum apare în lista de lucru.",
    namePlaceholder: "Prenume și nume complet",
    scoreLabel: "3. Scor FICO",
    scorePlaceholder: "ex. 850",
    submit: "Trimite scorul și dovada",
    success: "Scorul FICO și dovada au fost trimise cu succes.",
    already_sent: "Ai trimis deja scorul FICO pentru astăzi.",
    name_not_found: "Numele introdus nu apare în lista șoferilor programați astăzi.",
    ambiguous_name: "Numele este prea scurt sau seamănă cu mai mulți șoferi. Te rog să scrii mai mult din numele complet.",
    invalid_score: "Scorul FICO introdus nu este valid.",
    invalid_image_type: "Te rog să încarci o imagine JPG, PNG sau WEBP.",
    empty_image: "Imaginea selectată este goală. Te rog să alegi alt fișier.",
    image_too_large: "Imaginea este prea mare. Dimensiunea maximă este 10 MB.",
    generic_error: "Trimiterea nu a putut fi finalizată. Te rog să încerci din nou."
  }},
  de: {{
    title: "FICO-Score senden",
    intro: "Lade ein klares Foto oder einen Screenshot hoch, auf dem der FICO-Score sichtbar ist.",
    photoLabel: "1. FICO Foto / Screenshot",
    takePhoto: "Jetzt Foto aufnehmen",
    photoSelected: "Ausgewähltes Bild:",
    photoHint: "Nimm ein klares Foto auf, auf dem der FICO-Score sichtbar ist · maximal 10 MB",
    nameLabel: "2. Vollständiger Name",
    nameHint: "Schreibe deinen Namen genau so, wie er in der Fahrer-/Arbeitsliste steht.",
    namePlaceholder: "Vor- und Nachname",
    scoreLabel: "3. FICO-Score",
    scorePlaceholder: "z. B. 850",
    submit: "Score und Nachweis senden",
    success: "FICO-Score und Nachweis wurden erfolgreich gesendet.",
    already_sent: "Du hast deinen FICO-Score für heute bereits gesendet.",
    name_not_found: "Der eingegebene Name steht heute nicht auf der Liste der eingeplanten Fahrer.",
    ambiguous_name: "Der Name ist zu kurz oder passt zu mehreren Fahrern. Bitte gib mehr vom vollständigen Namen ein.",
    invalid_score: "Der eingegebene FICO-Score ist ungültig.",
    invalid_image_type: "Bitte lade ein JPG-, PNG- oder WEBP-Bild hoch.",
    empty_image: "Die ausgewählte Bilddatei ist leer. Bitte wähle eine andere Datei.",
    image_too_large: "Das Bild ist zu groß. Die maximale Dateigröße beträgt 10 MB.",
    generic_error: "Die Übermittlung konnte nicht abgeschlossen werden. Bitte versuche es erneut."
  }},
  en: {{
    title: "Submit FICO Score",
    intro: "Upload a clear photo or screenshot where the FICO score is visible.",
    photoLabel: "1. FICO Photo / Screenshot",
    takePhoto: "Take a photo now",
    photoSelected: "Selected image:",
    photoHint: "Take a clear photo where the FICO score is visible · maximum 10 MB",
    nameLabel: "2. Full Name",
    nameHint: "Enter your name exactly as it appears on the driver/work list.",
    namePlaceholder: "First and last name",
    scoreLabel: "3. FICO Score",
    scorePlaceholder: "e.g. 850",
    submit: "Submit score and proof",
    success: "Your FICO score and proof were submitted successfully.",
    already_sent: "You have already submitted your FICO score for today.",
    name_not_found: "The entered name is not on today's scheduled driver list.",
    ambiguous_name: "The name is too short or matches multiple drivers. Please enter more of the full name.",
    invalid_score: "The FICO score you entered is invalid.",
    invalid_image_type: "Please upload a JPG, PNG, or WEBP image.",
    empty_image: "The selected image is empty. Please choose another file.",
    image_too_large: "The image is too large. The maximum file size is 10 MB.",
    generic_error: "The submission could not be completed. Please try again."
  }}
}};


function updateSelectedProof() {{
  const input = document.getElementById("proof");
  const box = document.getElementById("proofSelected");
  const lang = localStorage.getItem("ficoLanguage") || detectInitialLanguage();
  const t = translations[lang] || translations.ro;

  if (input.files && input.files.length > 0) {{
    const name = input.files[0].name || "FICO";
    box.textContent = `${{t.photoSelected}} ${{name}}`;
    box.classList.add("show");
  }} else {{
    box.textContent = "";
    box.classList.remove("show");
  }}
}}

document.getElementById("proof").addEventListener("change", updateSelectedProof);

function detectInitialLanguage() {{
  const saved = localStorage.getItem("ficoLanguage");
  if (saved && translations[saved]) return saved;

  const browser = (navigator.language || "").toLowerCase();
  if (browser.startsWith("de")) return "de";
  if (browser.startsWith("en")) return "en";
  return "ro";
}}

function applyLanguage(lang) {{
  if (!translations[lang]) lang = "ro";
  localStorage.setItem("ficoLanguage", lang);
  document.documentElement.lang = lang;

  const t = translations[lang];
  document.querySelectorAll("[data-i18n]").forEach(el => {{
    const key = el.dataset.i18n;
    if (t[key]) el.textContent = t[key];
  }});

  document.getElementById("fullName").placeholder = t.namePlaceholder;
  document.getElementById("ficoScore").placeholder = t.scorePlaceholder;

  document.querySelectorAll(".lang-btn").forEach(btn => {{
    btn.classList.toggle("active", btn.dataset.lang === lang);
  }});

  if (document.getElementById("proof").files.length > 0) {{
    updateSelectedProof();
  }}

  renderNotice(lang);
}}

function renderNotice(lang) {{
  const params = new URLSearchParams(window.location.search);
  const notice = document.getElementById("notice");
  const t = translations[lang];

  notice.className = "notice";
  notice.textContent = "";

  if (params.get("success") === "1") {{
    notice.textContent = t.success;
    notice.classList.add("success");
    return;
  }}

  const error = params.get("error");
  if (error) {{
    notice.textContent = t[error] || t.generic_error;
    notice.classList.add("error");
  }}
}}

document.querySelectorAll(".lang-btn").forEach(btn => {{
  btn.addEventListener("click", () => applyLanguage(btn.dataset.lang));
}});

applyLanguage(detectInitialLanguage());
</script>
</body>
</html>
"""
    return HTMLResponse(page)



@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request, d: str | None = None, q: str | None = None):
    selected = d or date.today().isoformat()
    search = (q or "").strip()

    conn = db()

    rows = conn.execute("""
        SELECT d.name,
               s.fico_score,
               s.submitted_at,
               s.entered_full_name,
               s.proof_filename,
               s.proof_original_name,
               s.detected_fico_score,
               s.verification_status,
               s.name_match_score,
               CASE WHEN s.id IS NULL THEN 0 ELSE 1 END AS sent
        FROM daily_required r
        JOIN drivers d ON d.id = r.driver_id
        LEFT JOIN submissions s
          ON s.driver_id = d.id
         AND s.work_date = r.work_date
        WHERE r.work_date = ?
        ORDER BY sent ASC, d.name ASC
    """, (selected,)).fetchall()

    unresolved_rows = conn.execute("""
        SELECT id,
               entered_full_name,
               fico_score,
               submitted_at,
               proof_filename,
               proof_original_name,
               detected_fico_score,
               verification_status,
               best_match_name,
               best_match_score,
               match_reason
        FROM unresolved_submissions
        WHERE work_date = ?
        ORDER BY submitted_at DESC
    """, (selected,)).fetchall()

    try:
        selected_date_obj = datetime.strptime(selected, "%Y-%m-%d").date()
    except Exception:
        selected_date_obj = date.today()
        selected = selected_date_obj.isoformat()

    month_start = selected_date_obj.replace(day=1)
    next_month = (
        month_start.replace(year=month_start.year + 1, month=1)
        if month_start.month == 12
        else month_start.replace(month=month_start.month + 1)
    )

    history_rows = conn.execute("""
        SELECT r.work_date,
               COUNT(*) AS total,
               SUM(CASE WHEN s.id IS NULL THEN 0 ELSE 1 END) AS sent
        FROM daily_required r
        LEFT JOIN submissions s
          ON s.driver_id = r.driver_id
         AND s.work_date = r.work_date
        WHERE r.work_date >= ?
          AND r.work_date < ?
        GROUP BY r.work_date
        ORDER BY r.work_date
    """, (month_start.isoformat(), next_month.isoformat())).fetchall()

    conn.close()

    history_by_date = {
        r["work_date"]: {
            "total": int(r["total"] or 0),
            "sent": int(r["sent"] or 0)
        }
        for r in history_rows
    }

    total = len(rows)
    sent = sum(int(r["sent"]) for r in rows)
    missing = total - sent
    low_fico = sum(
        1 for r in rows
        if r["fico_score"] is not None and int(r["fico_score"]) < 800
    )

    needs_review = (
        sum(
            1 for r in rows
            if r["verification_status"] in ("mismatch", "manual_review")
        )
        + len(unresolved_rows)
    )

    missing_names = [r["name"] for r in rows if not r["sent"]]

    filtered_rows = rows
    if search:
        key = normalize_name(search)
        filtered_rows = [
            r for r in rows
            if key in normalize_name(r["name"])
            or key in normalize_name(r["entered_full_name"] or "")
        ]

    table_rows = ""

    for r in filtered_rows:
        sent_bool = bool(r["sent"])
        status = "Trimis" if sent_bool else "Nu a trimis"
        status_class = "sent" if sent_bool else "missing"

        fico = r["fico_score"] if r["fico_score"] is not None else "—"
        hour = r["submitted_at"][11:16] if r["submitted_at"] else "—"

        score_class = ""
        score_badge = ""

        if r["fico_score"] is not None:
            score = int(r["fico_score"])
            if score < 800:
                score_class = "score-low"
                score_badge = '<span class="score-note danger">Scăzut</span>'
            elif score < 850:
                score_class = "score-mid"
                score_badge = '<span class="score-note warning">Sub 850</span>'
            else:
                score_class = "score-good"
                score_badge = '<span class="score-note good">850</span>'

        if r["proof_filename"]:
            proof_url = f'/proof/{html.escape(r["proof_filename"])}'
            proof = f'<a class="proof-btn" href="{proof_url}" target="_blank">Vezi poza</a>'
        else:
            proof = '<span class="dash">—</span>'

        entered_name = html.escape(r["entered_full_name"]) if r["entered_full_name"] else "—"

        detected_fico = (
            r["detected_fico_score"]
            if r["detected_fico_score"] is not None
            else "—"
        )

        verification_status = r["verification_status"] or (
            "manual_review" if sent_bool else ""
        )

        if verification_status == "verified":
            verification = '<span class="verify-pill verified">✓ Verificat</span>'
        elif verification_status == "mismatch":
            verification = '<span class="verify-pill mismatch">⚠ Verifică</span>'
        elif verification_status == "manual_review":
            verification = '<span class="verify-pill manual">? Manual</span>'
        else:
            verification = '<span class="dash">—</span>'

        table_rows += f"""
        <tr class="{'row-missing' if not sent_bool else ''}">
            <td>
                <div class="driver-name">{html.escape(r["name"])}</div>
                <div class="entered-name">Introdus: {entered_name}</div>
            </td>
            <td><span class="status-pill {status_class}">{status}</span></td>
            <td class="{score_class}">
                <div class="score-wrap">
                    <strong>{fico}</strong>
                    {score_badge}
                </div>
            </td>
            <td><strong>{detected_fico}</strong></td>
            <td>{verification}</td>
            <td>{hour}</td>
            <td>{proof}</td>
        </tr>
        """

    unresolved_table = ""
    if unresolved_rows:
        unresolved_items = ""

        for u in unresolved_rows:
            entered = html.escape(u["entered_full_name"] or "—")
            best_match = html.escape(u["best_match_name"] or "Nicio potrivire sigură")
            best_score = (
                f'{round(float(u["best_match_score"]) * 100)}%'
                if u["best_match_score"] is not None
                else "—"
            )
            fico_unresolved = (
                u["fico_score"]
                if u["fico_score"] is not None
                else "—"
            )
            detected_unresolved = (
                u["detected_fico_score"]
                if u["detected_fico_score"] is not None
                else "—"
            )
            unresolved_hour = (
                u["submitted_at"][11:16]
                if u["submitted_at"]
                else "—"
            )

            if u["proof_filename"]:
                unresolved_proof_url = (
                    f'/proof/{html.escape(u["proof_filename"])}'
                )
                unresolved_proof = (
                    f'<a class="proof-btn" href="{unresolved_proof_url}" '
                    f'target="_blank">Vezi poza</a>'
                )
            else:
                unresolved_proof = '<span class="dash">—</span>'

            if u["verification_status"] == "mismatch":
                fico_check = '<span class="verify-pill mismatch">⚠ FICO diferit</span>'
            elif u["verification_status"] == "verified":
                fico_check = '<span class="verify-pill verified">✓ FICO verificat</span>'
            else:
                fico_check = '<span class="verify-pill manual">? FICO manual</span>'

            unresolved_items += f"""
            <tr class="row-name-review">
                <td>
                    <div class="driver-name">⚠ {entered}</div>
                    <div class="entered-name">
                        Posibil: {best_match} · potrivire {best_score}
                    </div>
                </td>
                <td><span class="verify-pill mismatch">⚠ Nume de verificat</span></td>
                <td><strong>{fico_unresolved}</strong></td>
                <td><strong>{detected_unresolved}</strong></td>
                <td>{fico_check}</td>
                <td>{unresolved_hour}</td>
                <td>{unresolved_proof}</td>
            </tr>
            """

        unresolved_table = f"""
        <section class="name-review-section">
            <div class="review-title">
                ⚠ Trimiteri cu nume neidentificat ({len(unresolved_rows)})
            </div>
            <div class="review-subtitle">
                Poza și scorul au fost salvate. Verifică manual cui aparține trimiterea.
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Nume introdus</th>
                        <th>Atenție</th>
                        <th>FICO introdus</th>
                        <th>FICO detectat</th>
                        <th>Verificare</th>
                        <th>Ora</th>
                        <th>Dovadă</th>
                    </tr>
                </thead>
                <tbody>{unresolved_items}</tbody>
            </table>
        </section>
        """

    missing_js = "\\n".join(missing_names).replace("\\\\", "\\\\\\\\").replace("`", "\\\\`")

    if filtered_rows:
        table_content = (
            '<table><thead><tr>'
            '<th>Șofer</th><th>Status</th><th>FICO introdus</th><th>FICO detectat</th><th>Verificare</th><th>Ora</th><th>Dovadă</th>'
            '</tr></thead><tbody>'
            + table_rows +
            '</tbody></table>'
        )
    else:
        table_content = '<div class="empty">Nu există șoferi pentru această selecție.</div>'

    month_names_ro = [
        "",
        "Ianuarie", "Februarie", "Martie", "Aprilie",
        "Mai", "Iunie", "Iulie", "August",
        "Septembrie", "Octombrie", "Noiembrie", "Decembrie"
    ]

    prev_month_date = (
        month_start.replace(year=month_start.year - 1, month=12)
        if month_start.month == 1
        else month_start.replace(month=month_start.month - 1)
    )
    next_month_date = next_month

    cal = pycalendar.Calendar(firstweekday=0)
    calendar_cells = ""
    today_iso = date.today().isoformat()

    for day_obj in cal.itermonthdates(month_start.year, month_start.month):
        iso = day_obj.isoformat()
        day_info = history_by_date.get(iso)

        classes = ["calendar-day"]
        if day_obj.month != month_start.month:
            classes.append("outside")
        if iso == selected:
            classes.append("selected")
        if iso == today_iso:
            classes.append("today")
        if day_info:
            classes.append("has-data")

        if day_info:
            total_d = day_info["total"]
            sent_d = day_info["sent"]
            missing_d = max(total_d - sent_d, 0)
            detail = (
                f'<span class="calendar-dot"></span>'
                f'<span class="calendar-count">{sent_d}/{total_d}</span>'
            )
            title = (
                f"{day_obj.strftime('%d.%m.%Y')} · "
                f"{sent_d} au trimis · {missing_d} lipsesc"
            )
        else:
            detail = ""
            title = day_obj.strftime("%d.%m.%Y")

        calendar_cells += f"""
        <a class="{' '.join(classes)}"
           href="/admin?d={iso}"
           title="{html.escape(title)}">
            <span class="calendar-number">{day_obj.day}</span>
            {detail}
        </a>
        """

    calendar_html = f"""
    <section class="panel calendar-panel">
        <div class="calendar-header">
            <div>
                <div class="import-title">Istoric FICO</div>
                <div class="calendar-month">
                    {month_names_ro[month_start.month]} {month_start.year}
                </div>
            </div>
            <div class="calendar-nav">
                <a class="btn-light calendar-nav-btn"
                   href="/admin?d={prev_month_date.isoformat()}">‹</a>
                <a class="btn-light calendar-today"
                   href="/admin?d={date.today().isoformat()}">Astăzi</a>
                <a class="btn-light calendar-nav-btn"
                   href="/admin?d={next_month_date.isoformat()}">›</a>
            </div>
        </div>

        <div class="calendar-weekdays">
            <span>Lu</span><span>Ma</span><span>Mi</span>
            <span>Jo</span><span>Vi</span><span>Sâ</span><span>Du</span>
        </div>

        <div class="calendar-grid">
            {calendar_cells}
        </div>

        <div class="calendar-legend">
            <span><i class="legend-dot saved"></i> Zi cu listă salvată</span>
            <span>Zi afișată: <strong>{selected_date_obj.strftime("%d.%m.%Y")}</strong></span>
        </div>
    </section>
    """

    page = f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FICO Control Admin</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;font-family:Arial,Helvetica,sans-serif;background:#f4f6f8;color:#17212b}}
.app-shell{{width:min(96%,1280px);margin:28px auto 60px;display:block}}
.admin{{width:100%;min-width:0;margin:0}}
.side-nav{{position:fixed;left:18px;top:18px;z-index:20;width:170px;background:transparent;color:#17212b;padding:0}}
.side-brand{{font-size:10px;font-weight:800;letter-spacing:1.5px;color:#98a2b3;margin:0 3px 9px}}
.side-nav-links{{display:flex;flex-direction:column;gap:9px}}
.side-link{{display:flex;align-items:center;color:#17212b;text-decoration:none;padding:11px 12px;border-radius:9px;font-size:13px;font-weight:700;border:1px solid #d8dde3;background:#fff;box-shadow:0 2px 7px rgba(23,33,43,.04)}}
.side-link:hover{{background:#f8fafb;border-color:#b8c0ca}}
.side-link i{{display:none}}
.topbar{{display:flex;align-items:flex-start;justify-content:space-between;gap:20px;margin-bottom:28px}}
.brand{{font-size:13px;font-weight:800;letter-spacing:2px}}
h1{{font-size:38px;margin:14px 0 0}}
.top-actions{{display:flex;gap:10px;flex-wrap:wrap;align-items:center}}
.btn,.btn-light{{display:inline-flex;align-items:center;justify-content:center;text-decoration:none;padding:12px 15px;border-radius:10px;font-size:14px;font-weight:800;cursor:pointer}}
.btn{{border:0;background:#17212b;color:#fff}}
.btn-light{{border:1px solid #d8dde3;background:#fff;color:#17212b}}
.stats{{display:grid;grid-template-columns:repeat(5,1fr);gap:15px;margin-bottom:20px}}
.card,.panel{{background:#fff;border-radius:16px;box-shadow:0 5px 18px rgba(0,0,0,.05)}}
.card{{padding:20px}}
.card strong{{display:block;font-size:32px;line-height:1;margin-bottom:8px}}
.card span{{color:#667085}}
.card.alert strong{{color:#d13b2e}}
.panel{{padding:20px;margin-bottom:15px}}
.controls{{display:flex;justify-content:space-between;gap:18px;flex-wrap:wrap}}
.control-group{{display:flex;gap:10px;flex-wrap:wrap;align-items:center}}
input,button{{padding:12px 14px;border-radius:10px;border:1px solid #d8dde3;font-size:15px}}
button{{cursor:pointer}}
button.primary{{border:0;background:#17212b;color:#fff;font-weight:800}}
.search{{min-width:260px}}
.import-title{{font-size:13px;color:#667085;font-weight:700;margin-bottom:10px}}
.table-wrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;min-width:850px}}
th,td{{text-align:left;padding:14px 13px;border-bottom:1px solid #eceff2;vertical-align:middle}}
th{{font-size:13px;color:#667085;text-transform:uppercase;letter-spacing:.4px}}
.driver-name{{font-weight:800}}
.entered-name{{color:#667085;font-size:12px;margin-top:4px}}
.status-pill{{display:inline-block;border-radius:999px;padding:6px 9px;font-size:12px;font-weight:800}}
.status-pill.sent{{background:#e9f8ef;color:#14804a}}
.status-pill.missing{{background:#fdeeee;color:#c9362b}}
.row-missing{{background:#fffafa}}
.score-wrap{{display:flex;align-items:center;gap:8px}}
.score-wrap strong{{font-size:17px}}
.score-note{{font-size:11px;font-weight:800;border-radius:999px;padding:4px 7px}}
.score-note.danger{{background:#fde2e1;color:#b42318}}
.score-note.warning{{background:#fff4d6;color:#8a5a00}}
.name-review-section{{margin:0 0 20px;border:2px solid #f59e0b;border-radius:16px;overflow:hidden;background:#fffaf0}}
.review-title{{font-size:18px;font-weight:900;color:#9a5a00;padding:16px 18px 4px}}
.review-subtitle{{font-size:13px;color:#8a6500;padding:0 18px 14px}}
.row-name-review{{background:#fff8e6}}
.score-note.good{{background:#e9f8ef;color:#14804a}}
.proof-btn{{display:inline-block;text-decoration:none;color:#17212b;border:1px solid #d8dde3;background:#fff;padding:8px 10px;border-radius:8px;font-size:13px;font-weight:800}}
.proof-btn:hover{{background:#f4f6f8}}
.verify-pill{{display:inline-block;border-radius:999px;padding:6px 9px;font-size:12px;font-weight:800}}
.verify-pill.verified{{background:#e9f8ef;color:#14804a}}
.verify-pill.mismatch{{background:#fde2e1;color:#b42318}}
.verify-pill.manual{{background:#fff4d6;color:#8a5a00}}
.dash{{color:#98a2b3}}
.helper{{font-size:12px;color:#667085;margin-top:10px}}
.copy-ok{{display:none;margin-left:4px;color:#14804a;font-size:13px;font-weight:800}}
.copy-ok.show{{display:inline}}
.empty{{padding:35px;text-align:center;color:#667085}}
.calendar-panel{{padding:22px}}
.calendar-header{{display:flex;justify-content:space-between;align-items:center;gap:15px;margin-bottom:16px}}
.calendar-month{{font-size:24px;font-weight:900;margin-top:3px}}
.calendar-nav{{display:flex;gap:8px;align-items:center}}
.calendar-nav-btn{{width:42px;height:42px;padding:0;font-size:24px}}
.calendar-today{{height:42px}}
.calendar-weekdays,.calendar-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:7px}}
.calendar-weekdays{{margin-bottom:7px}}
.calendar-weekdays span{{text-align:center;color:#667085;font-size:12px;font-weight:800;padding:4px}}
.calendar-day{{position:relative;min-height:70px;border:1px solid #e4e7ec;border-radius:12px;padding:9px;text-decoration:none;color:#17212b;background:#fff;transition:.15s ease}}
.calendar-day:hover{{border-color:#98a2b3;transform:translateY(-1px)}}
.calendar-day.outside{{opacity:.34}}
.calendar-day.has-data{{background:#f7fbf8;border-color:#b7e1c8}}
.calendar-day.selected{{border:2px solid #17212b;background:#f3f5f7}}
.calendar-day.today .calendar-number{{background:#17212b;color:#fff;border-radius:999px;min-width:27px;height:27px;display:inline-flex;align-items:center;justify-content:center}}
.calendar-number{{font-weight:900;font-size:14px}}
.calendar-dot{{position:absolute;left:10px;bottom:11px;width:8px;height:8px;border-radius:50%;background:#14804a}}
.calendar-count{{position:absolute;right:9px;bottom:9px;font-size:11px;font-weight:800;color:#14804a}}
.calendar-legend{{display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-top:14px;color:#667085;font-size:12px}}
.legend-dot{{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:5px}}
.legend-dot.saved{{background:#14804a}}
@media(max-width:1500px){{.side-nav{{position:static;width:100%;margin-bottom:16px}}.side-brand{{display:none}}.side-nav-links{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr))}}.side-link{{justify-content:center}}}}
@media(max-width:850px){{.stats{{grid-template-columns:repeat(2,1fr)}}.topbar{{display:block}}.top-actions{{margin-top:18px}}.side-nav-links{{grid-template-columns:repeat(2,1fr)}}}}
@media(max-width:520px){{.stats{{grid-template-columns:1fr}}h1{{font-size:30px}}.search{{min-width:100%;width:100%}}.control-group{{width:100%}}.calendar-panel{{padding:14px}}.calendar-header{{align-items:flex-start}}.calendar-month{{font-size:20px}}.calendar-day{{min-height:54px;padding:7px}}.calendar-count{{display:none}}}}
</style>
</head>
<body>

<div class="app-shell">
<aside class="side-nav">
    <div class="side-brand">INSTRUMENTE FICO</div>
    <nav class="side-nav-links">
        <a class="side-link" href="/admin/mentor?d={selected}"><i></i>Mentor Check</a>
        <a class="side-link" href="/admin/score-check"><i></i>Verificare Scor</a>
        <a class="side-link" href="/admin/hours"><i></i>Control ore</a>
        <a class="side-link" href="/admin/pod-ccc"><i></i>POD & CCC</a>
        <a class="side-link" href="/admin/concessions"><i></i>Concesii</a>
        <a class="side-link" href="/admin/atlas-paket"><i></i>Atlas Paket</a>
        <a class="side-link" href="https://clinquant-sawine-dba777.netlify.app/" target="_blank" rel="noopener noreferrer"><i></i>A doua reîncercare</a>
    </nav>
</aside>
<main class="admin">

<div class="topbar">
    <div>
        <div class="brand">FICO CONTROL</div>
        <h1>Admin Dashboard</h1>
        <div style="margin-top:8px;color:#667085;font-size:13px">
            Conectat ca: <strong>{html.escape(getattr(request.state, "admin_session", {}).get("display_name", "Admin"))}</strong>
            · Bază date: <strong>{html.escape(DB_BACKEND.upper())}</strong>
        </div>
    </div>

    <div class="top-actions">
        <a class="btn-light" href="/admin/owner">Owner</a>
        <form method="post" action="/admin/logout" style="margin:0">
            <button class="btn-light" type="submit">Ieșire</button>
        </form>
        <a class="btn-light" href="/admin/export.xlsx?d={selected}">Export Excel</a>
        <button class="btn" type="button" onclick="copyMissing()">Copiază șoferii lipsă</button>
        <span id="copyOk" class="copy-ok">Copiat</span>
    </div>
</div>

<section class="stats">
    <div class="card"><strong>{total}</strong><span>Programați</span></div>
    <div class="card"><strong>{sent}</strong><span>Au trimis</span></div>
    <div class="card alert"><strong>{missing}</strong><span>Lipsesc</span></div>
    <div class="card alert"><strong>{low_fico}</strong><span>FICO sub 800</span></div>
    <div class="card alert"><strong>{needs_review}</strong><span>Necesită verificare</span></div>
</section>

{calendar_html}

<section class="panel">
    <div class="controls">
        <form class="control-group" action="/admin" method="get">
            <input type="date" name="d" value="{selected}">
            <input class="search" type="text" name="q" value="{html.escape(search)}" placeholder="Caută șofer...">
            <button class="primary" type="submit">Afișează</button>
            <a class="btn-light" href="/admin?d={selected}">Reset</a>
        </form>
    </div>
</section>

<section class="panel">
    <div class="import-title">Lista zilnică din Cortex</div>
    <form class="control-group" action="/admin/upload" method="post" enctype="multipart/form-data">
        <input type="date" name="work_date" value="{selected}" required>
        <input type="file" name="file" accept=".xlsx" required>
        <button class="primary" type="submit">Încarcă Excel Cortex</button>
    </form>
    <div class="helper">Aplicația extrage automat șoferii din coloana „Name des Fahrers”.</div>
</section>

<section class="panel table-wrap">
    {unresolved_table}
    {table_content}
</section>

</main>
</div>

<script>
const missingNames = `{missing_js}`;

async function copyMissing() {{
    const message = missingNames.trim();

    if (!message) {{
        alert("Toți șoferii au trimis scorul FICO.");
        return;
    }}

    try {{
        await navigator.clipboard.writeText(message);
        const el = document.getElementById("copyOk");
        el.classList.add("show");
        setTimeout(() => el.classList.remove("show"), 1800);
    }} catch (e) {{
        window.prompt("Copiază lista:", message);
    }}
}}
</script>

</body>
</html>
"""

    return HTMLResponse(page)


@app.get("/admin/export.xlsx")
def export_day_excel(d: str | None = None):
    selected = d or date.today().isoformat()
    conn = db()

    rows = conn.execute("""
        SELECT d.name,
               CASE WHEN s.id IS NULL THEN 'Nu a trimis' ELSE 'Trimis' END AS status,
               s.fico_score,
               s.submitted_at,
               COALESCE(s.entered_full_name, '') AS entered_full_name,
               COALESCE(s.proof_filename, '') AS proof_filename,
               s.detected_fico_score,
               COALESCE(s.verification_status, '') AS verification_status,
               s.name_match_score
        FROM daily_required r
        JOIN drivers d ON d.id = r.driver_id
        LEFT JOIN submissions s
          ON s.driver_id = d.id
         AND s.work_date = r.work_date
        WHERE r.work_date = ?
        ORDER BY
            CASE WHEN s.id IS NULL THEN 0 ELSE 1 END ASC,
            d.name ASC
    """, (selected,)).fetchall()

    conn.close()

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "FICO"

    headers = [
        "Șofer",
        "Status",
        "FICO introdus",
        "FICO detectat",
        "Verificare",
        "Ora trimiterii",
        "Nume introdus",
        "Potrivire nume",
        "Dovadă"
    ]
    ws.append(headers)

    for row in rows:
        proof = f"/proof/{row['proof_filename']}" if row["proof_filename"] else ""
        ws.append([
            row["name"],
            row["status"],
            row["fico_score"] if row["fico_score"] is not None else "",
            row["detected_fico_score"] if row["detected_fico_score"] is not None else "",
            row["verification_status"],
            row["submitted_at"] or "",
            row["entered_full_name"],
            round(float(row["name_match_score"]), 3) if row["name_match_score"] is not None else "",
            proof
        ])

    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill("solid", fgColor="17212B")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 30,
        "B": 16,
        "C": 15,
        "D": 15,
        "E": 20,
        "F": 23,
        "G": 30,
        "H": 15,
        "I": 30
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    missing_fill = PatternFill("solid", fgColor="FDEEEE")
    low_fill = PatternFill("solid", fgColor="F4CCCC")

    for row_idx in range(2, ws.max_row + 1):
        status = ws[f"B{row_idx}"].value
        score = ws[f"C{row_idx}"].value

        if status == "Nu a trimis":
            for col_idx in range(1, 10):
                ws.cell(row=row_idx, column=col_idx).fill = missing_fill

        if isinstance(score, int) and score < 800:
            ws[f"C{row_idx}"].fill = low_fill
            ws[f"C{row_idx}"].font = Font(bold=True)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"FICO_{selected}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="{filename}"'
        }
    )

# ============================================================================
# ATLAS PAKET V2
# Multiple routes per Transfer Sheet image + multi-page continuation support
# ============================================================================

ATLAS_MAX_IMAGES = 12
ATLAS_MAX_IMAGE_BYTES = 12 * 1024 * 1024
ATLAS_ALLOWED_TYPES = {"image/jpeg", "image/png", "image/webp"}

_ATLAS_OCR_DIGITS = {
    "0": "0",
    "1": "1",
    "2": "2",
    "3": "3",
    "4": "4",
    "5": "5",
    "6": "6",
    "7": "7",
    "8": "8",
    "9": "9",
    "O": "0",
    "Q": "0",
    "I": "1",
    "L": "1",
    "S": "5",
    "B": "8",
    "Z": "2",
    "G": "6",
}


def atlas_clean_route(value: str) -> str:
    value = str(value or "").upper().strip()
    value = re.sub(r"\s+", "", value)
    value = value.replace("-", "_")
    return value


def atlas_extract_cortex_routes(raw: bytes) -> dict:
    workbook = openpyxl.load_workbook(
        io.BytesIO(raw),
        read_only=True,
        data_only=True
    )

    try:
        ws = (
            workbook["Strecken"]
            if "Strecken" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )

        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)

        if not headers:
            raise ValueError("Cortex-ul nu conține antet.")

        normalized = [
            re.sub(r"\s+", " ", str(header or "")).strip().casefold()
            for header in headers
        ]

        route_candidates = [
            "routencode",
            "route code",
            "route",
            "route-code"
        ]
        driver_candidates = [
            "name des fahrers",
            "fahrername",
            "driver name",
            "fahrer"
        ]

        route_index = next(
            (
                normalized.index(candidate)
                for candidate in route_candidates
                if candidate in normalized
            ),
            None
        )
        driver_index = next(
            (
                normalized.index(candidate)
                for candidate in driver_candidates
                if candidate in normalized
            ),
            None
        )

        if route_index is None:
            raise ValueError("Nu am găsit coloana Routencode în Cortex.")

        if driver_index is None:
            raise ValueError(
                "Nu am găsit coloana Name des Fahrers în Cortex."
            )

        route_to_driver = {}

        for row in rows:
            if route_index >= len(row) or driver_index >= len(row):
                continue

            route = atlas_clean_route(row[route_index])
            raw_driver = str(row[driver_index] or "").strip()

            if not route or not raw_driver:
                continue

            # Only the first driver counts. Everything after | / newline is rescue.
            driver_parts = re.split(r"[|\n\r]+", raw_driver)
            driver = next(
                (part.strip() for part in driver_parts if part.strip()),
                ""
            )

            if route and driver:
                route_to_driver[route] = driver

        if not route_to_driver:
            raise ValueError("Nu am găsit rute valide în Cortex.")

        return route_to_driver

    finally:
        workbook.close()


def atlas_ocr_page(
    raw: bytes,
    content_type: str,
    filename: str,
    upload_index: int
) -> dict:
    """
    Read one Amazon Transfer Sheet image.

    V2 requests:
    - table mode, so rows are preserved;
    - overlay coordinates, so route and tracking columns can be reunited;
    - orientation detection and scaling.
    """
    api_key = os.getenv("OCRSPACE_API_KEY", "").strip()

    if not api_key:
        raise ValueError(
            "OCRSPACE_API_KEY nu este configurat în Render. "
            "Atlas Paket are nevoie de OCR pentru a citi pozele Amazon."
        )

    boundary = "----AtlasPaketV2Boundary7MA4YWxkTrZu0gW"

    extension = "jpg"
    if content_type == "image/png":
        extension = "png"
    elif content_type == "image/webp":
        extension = "webp"

    parts = []

    def add_field(name, value):
        parts.append(
            (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n"
            ).encode("utf-8")
        )

    add_field("apikey", api_key)
    add_field("language", "eng")
    add_field("isOverlayRequired", "true")
    add_field("isTable", "true")
    add_field("OCREngine", "2")
    add_field("scale", "true")
    add_field("detectOrientation", "true")

    parts.append(
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; '
            f'filename="atlas.{extension}"\r\n'
            f"Content-Type: {content_type}\r\n\r\n"
        ).encode("utf-8")
        + raw
        + b"\r\n"
    )

    parts.append(f"--{boundary}--\r\n".encode("utf-8"))

    request = urllib.request.Request(
        "https://api.ocr.space/parse/image",
        data=b"".join(parts),
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "User-Agent": "FICO-Control-Atlas/2.0"
        },
        method="POST"
    )

    try:
        with urllib.request.urlopen(request, timeout=40) as response:
            payload = json.loads(
                response.read().decode("utf-8", errors="replace")
            )
    except Exception as exc:
        raise ValueError(
            f"OCR-ul nu a putut citi poza {filename}."
        ) from exc

    if payload.get("IsErroredOnProcessing"):
        message = payload.get("ErrorMessage") or payload.get("ErrorDetails")

        if isinstance(message, list):
            message = " ".join(str(item) for item in message)

        raise ValueError(
            f"OCR Atlas pentru {filename}: "
            + str(message or "eroare necunoscută")
        )

    parsed_results = payload.get("ParsedResults") or []

    if not parsed_results:
        raise ValueError(
            f"Poza {filename} nu a produs niciun rezultat OCR."
        )

    parsed_text_parts = []
    line_records = []
    synthetic_top = 0.0

    for result in parsed_results:
        parsed_text = str(result.get("ParsedText") or "")
        if parsed_text:
            parsed_text_parts.append(parsed_text)

        overlay = result.get("TextOverlay") or {}
        overlay_lines = overlay.get("Lines") or []

        for line in overlay_lines:
            words = line.get("Words") or []
            line_text = str(line.get("LineText") or "").strip()

            if not line_text and words:
                line_text = " ".join(
                    str(word.get("WordText") or "").strip()
                    for word in words
                    if str(word.get("WordText") or "").strip()
                )

            if not line_text:
                continue

            top = line.get("MinTop")
            if top is None and words:
                top = min(
                    float(word.get("Top") or 0)
                    for word in words
                )
            if top is None:
                top = synthetic_top

            left = 0.0
            if words:
                left = min(
                    float(word.get("Left") or 0)
                    for word in words
                )

            height = line.get("MaxHeight")
            if height is None and words:
                height = max(
                    float(word.get("Height") or 0)
                    for word in words
                )
            if not height:
                height = 12.0

            line_records.append({
                "text": line_text,
                "top": float(top),
                "left": float(left),
                "height": float(height),
            })

            synthetic_top = max(
                synthetic_top + 18.0,
                float(top) + float(height) + 3.0
            )

    combined_text = "\n".join(parsed_text_parts).strip()

    # Overlay fallback: ParsedText still gives usable line order.
    if not line_records:
        for line_number, line_text in enumerate(
            combined_text.splitlines()
        ):
            line_text = line_text.strip()
            if not line_text:
                continue

            line_records.append({
                "text": line_text,
                "top": float(line_number * 22),
                "left": 0.0,
                "height": 14.0,
            })

    page_number = None
    page_total = None

    page_match = re.search(
        r"\bPAGE\s*(\d+)\s*(?:OF|/)\s*(\d+)\b",
        combined_text,
        flags=re.IGNORECASE
    )

    if page_match:
        page_number = int(page_match.group(1))
        page_total = int(page_match.group(2))

    return {
        "filename": filename,
        "upload_index": upload_index,
        "page_number": page_number,
        "page_total": page_total,
        "text": combined_text,
        "lines": line_records,
    }


def atlas_route_codes_from_text(text: str) -> list[str]:
    upper = str(text or "").upper()

    route_codes = []

    for match in re.finditer(
        r"\b(CA|SA|SB|SC)\s*[_\-\s]*A\s*[_\-\s]*"
        r"([0-9OQILSBZG]{1,3})\b",
        upper
    ):
        prefix = match.group(1)
        number_raw = match.group(2)

        number_digits = "".join(
            _ATLAS_OCR_DIGITS.get(character, "")
            for character in number_raw
        )

        if not number_digits:
            continue

        route_codes.append(
            atlas_clean_route(
                f"{prefix}_A{int(number_digits)}"
            )
        )

    return list(dict.fromkeys(route_codes))


def atlas_tracking_ids_from_text(text: str) -> list[str]:
    """
    Read DE + ten digits while tolerating spaces and common OCR confusions.
    The scanner intentionally starts at every DE-like prefix and stops after
    exactly ten normalized digits.
    """
    upper = str(text or "").upper()
    results = []

    for prefix_match in re.finditer(r"D\s*[E3]", upper):
        tail = upper[prefix_match.end():prefix_match.end() + 45]

        digits = []
        started = False

        for character in tail:
            if character in _ATLAS_OCR_DIGITS:
                digits.append(_ATLAS_OCR_DIGITS[character])
                started = True

                if len(digits) == 10:
                    break

                continue

            if character in " \t\r\n-_:./|":
                continue

            if started:
                break

        if len(digits) == 10:
            tracking_id = "DE" + "".join(digits)
            if tracking_id not in results:
                results.append(tracking_id)

    return results


def atlas_group_overlay_rows(lines: list[dict]) -> list[dict]:
    """
    OCR sometimes returns the Route Code cell and Tracking ID cell as separate
    overlay lines. Group nearby vertical positions into one visual table row.
    """
    ordered = sorted(
        lines,
        key=lambda item: (
            float(item.get("top") or 0),
            float(item.get("left") or 0)
        )
    )

    grouped_rows = []

    for item in ordered:
        top = float(item.get("top") or 0)
        height = max(float(item.get("height") or 12), 8.0)

        if not grouped_rows:
            grouped_rows.append({
                "top": top,
                "height": height,
                "items": [item],
            })
            continue

        current = grouped_rows[-1]
        tolerance = max(
            9.0,
            min(
                18.0,
                max(current["height"], height) * 0.85
            )
        )

        if abs(top - current["top"]) <= tolerance:
            current["items"].append(item)
            current["top"] = min(current["top"], top)
            current["height"] = max(current["height"], height)
        else:
            grouped_rows.append({
                "top": top,
                "height": height,
                "items": [item],
            })

    visual_rows = []

    for row in grouped_rows:
        row_items = sorted(
            row["items"],
            key=lambda item: float(item.get("left") or 0)
        )

        texts = [
            str(item.get("text") or "").strip()
            for item in row_items
            if str(item.get("text") or "").strip()
        ]

        if texts:
            visual_rows.append({
                "top": row["top"],
                "text": " | ".join(texts),
            })

    return visual_rows


def atlas_build_assignments(
    cortex_routes: dict,
    pages: list[dict]
):
    """
    Parse every Transfer Sheet row in sequence.

    Important:
    - one image can contain many routes;
    - a route can have several tracking rows;
    - the first rows on Page 2 can continue the last route from Page 1.
    """
    pages_sorted = sorted(
        pages,
        key=lambda page: (
            page["page_number"] is None,
            (
                page["page_number"]
                if page["page_number"] is not None
                else page["upload_index"]
            ),
            page["upload_index"]
        )
    )

    assignments = []
    review = []
    seen_tracking = set()
    current_route = None

    for page in pages_sorted:
        visual_rows = atlas_group_overlay_rows(page["lines"])

        for visual_row in visual_rows:
            row_text = visual_row["text"]
            row_routes = atlas_route_codes_from_text(row_text)
            row_tracking_ids = atlas_tracking_ids_from_text(row_text)

            if row_routes:
                valid_row_routes = list(dict.fromkeys(row_routes))

                if len(valid_row_routes) == 1:
                    current_route = valid_row_routes[0]
                else:
                    cortex_candidates = [
                        route
                        for route in valid_row_routes
                        if route in cortex_routes
                    ]

                    if len(cortex_candidates) == 1:
                        current_route = cortex_candidates[0]
                    else:
                        for tracking_id in row_tracking_ids or ["—"]:
                            review.append({
                                "file": page["filename"],
                                "route": ", ".join(valid_row_routes),
                                "tracking": tracking_id,
                                "reason": (
                                    "Mai multe rute au fost citite "
                                    "pe același rând"
                                )
                            })

                        current_route = None
                        continue

            if not row_tracking_ids:
                continue

            for tracking_id in row_tracking_ids:
                if tracking_id in seen_tracking:
                    continue

                seen_tracking.add(tracking_id)

                if not current_route:
                    review.append({
                        "file": page["filename"],
                        "route": "—",
                        "tracking": tracking_id,
                        "reason": (
                            "Tracking ID fără o rută anterioară "
                            "identificată"
                        )
                    })
                    continue

                if current_route not in cortex_routes:
                    review.append({
                        "file": page["filename"],
                        "route": current_route,
                        "tracking": tracking_id,
                        "reason": "Ruta nu există în Excelul Cortex"
                    })
                    continue

                assignments.append({
                    "driver": cortex_routes[current_route],
                    "route": current_route,
                    "tracking": tracking_id,
                    "file": page["filename"],
                })

    assignments.sort(
        key=lambda item: (
            item["driver"].casefold(),
            item["route"],
            item["tracking"]
        )
    )

    return assignments, review

def atlas_package_count_label(count: int) -> str:
    return f"{count} pachet" if count == 1 else f"{count} pachete"


def atlas_build_excel(assignments, review):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Atlas Paket"
    worksheet.sheet_view.showGridLines = False

    dark = "17212B"
    white = "FFFFFF"
    green = "E9F8EF"
    red = "FDEEEE"
    line = "D8DDE3"

    worksheet.merge_cells("A1:E1")
    title = worksheet["A1"]
    title.value = "ATLAS PAKET"
    title.font = openpyxl.styles.Font(
        name="Arial",
        size=17,
        bold=True,
        color=white
    )
    title.fill = openpyxl.styles.PatternFill(
        "solid",
        fgColor=dark
    )
    title.alignment = openpyxl.styles.Alignment(
        horizontal="center",
        vertical="center"
    )
    worksheet.row_dimensions[1].height = 32

    headers = [
        "Șofer",
        "Route Code",
        "Nr. pachete",
        "Tracking ID",
        "Status"
    ]

    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(2, column, value)
        cell.font = openpyxl.styles.Font(
            name="Arial",
            bold=True,
            color=white
        )
        cell.fill = openpyxl.styles.PatternFill(
            "solid",
            fgColor=dark
        )
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center",
            vertical="center"
        )

    row_index = 3
    grouped = {}

    for item in assignments:
        grouped.setdefault(
            (item["driver"], item["route"]),
            []
        ).append(item["tracking"])

    for (driver, route), tracking_ids in grouped.items():
        package_label = atlas_package_count_label(
            len(tracking_ids)
        )

        for position, tracking_id in enumerate(tracking_ids):
            values = [
                driver if position == 0 else "",
                route if position == 0 else "",
                package_label if position == 0 else "",
                tracking_id,
                "OK"
            ]

            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row_index, column, value)
                cell.fill = openpyxl.styles.PatternFill(
                    "solid",
                    fgColor=green
                )
                cell.border = openpyxl.styles.Border(
                    bottom=openpyxl.styles.Side(
                        style="thin",
                        color=line
                    )
                )
                cell.alignment = openpyxl.styles.Alignment(
                    vertical="center",
                    horizontal=(
                        "center"
                        if column in {2, 3, 5}
                        else "left"
                    )
                )

            if position == 0:
                worksheet.cell(
                    row_index,
                    1
                ).font = openpyxl.styles.Font(
                    name="Arial",
                    bold=True
                )
                worksheet.cell(
                    row_index,
                    3
                ).font = openpyxl.styles.Font(
                    name="Arial",
                    bold=True,
                    color="14804A"
                )

            row_index += 1

    if review:
        row_index += 1

        worksheet.cell(
            row_index,
            1,
            "NECESITĂ VERIFICARE"
        )
        worksheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=5
        )
        worksheet.cell(
            row_index,
            1
        ).font = openpyxl.styles.Font(
            bold=True,
            color="B42318"
        )

        row_index += 1

        for item in review:
            values = [
                item.get("file", ""),
                item.get("route", ""),
                "",
                item.get("tracking", ""),
                item.get("reason", "")
            ]

            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row_index, column, value)
                cell.fill = openpyxl.styles.PatternFill(
                    "solid",
                    fgColor=red
                )
                cell.alignment = openpyxl.styles.Alignment(
                    wrap_text=True,
                    vertical="center"
                )

            row_index += 1

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 23
    worksheet.column_dimensions["E"].width = 42
    worksheet.freeze_panes = "A3"

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()

def atlas_page_html(assignments=None, review=None, error=""):
    assignments = assignments or []
    review = review or []

    error_html = (
        f'<div class="atlas-error"><strong>Eroare:</strong> '
        f'{html.escape(error)}</div>'
        if error else ""
    )

    result_html = ""

    if assignments or review:
        grouped = {}

        for item in assignments:
            key = (item["driver"], item["route"])
            grouped.setdefault(key, []).append(item["tracking"])

        assignment_rows = ""
        copy_sections = []

        for (driver, route), tracking_ids in grouped.items():
            package_count = len(tracking_ids)
            package_label = atlas_package_count_label(
                package_count
            )

            copy_sections.append(
                "\n".join([
                    f"{driver} — {route} — {package_label}",
                    *tracking_ids
                ])
            )

            for position, tracking_id in enumerate(tracking_ids):
                group_cells = ""

                if position == 0:
                    group_cells = f"""
                    <td
                      class="atlas-group-cell"
                      rowspan="{package_count}"
                    >
                      <strong>{html.escape(driver)}</strong>
                    </td>
                    <td
                      class="atlas-group-cell atlas-route"
                      rowspan="{package_count}"
                    >
                      {html.escape(route)}
                    </td>
                    <td
                      class="atlas-group-cell"
                      rowspan="{package_count}"
                    >
                      <span class="atlas-count">
                        {html.escape(package_label)}
                      </span>
                    </td>
                    """

                assignment_rows += f"""
                <tr class="{'atlas-group-start' if position == 0 else ''}">
                  {group_cells}
                  <td>
                    <code>{html.escape(tracking_id)}</code>
                  </td>
                  <td>
                    <span class="atlas-ok">OK</span>
                  </td>
                </tr>
                """

        review_rows = ""

        for item in review:
            review_rows += f"""
            <tr class="atlas-review-row">
              <td>{html.escape(item.get("file", "—"))}</td>
              <td>{html.escape(item.get("route", "—"))}</td>
              <td>
                <code>{html.escape(item.get("tracking", "—"))}</code>
              </td>
              <td>{html.escape(item.get("reason", ""))}</td>
            </tr>
            """

        excel_bytes = atlas_build_excel(
            assignments,
            review
        )
        excel_base64 = base64.b64encode(
            excel_bytes
        ).decode("ascii")

        copy_text = "\n\n".join(copy_sections)

        review_block = ""

        if review:
            review_block = f"""
            <section class="atlas-table atlas-review">
              <div class="atlas-review-title">
                Necesită verificare ({len(review)})
              </div>
              <table>
                <thead>
                  <tr>
                    <th>Poză</th>
                    <th>Rută citită</th>
                    <th>Tracking ID</th>
                    <th>Motiv</th>
                  </tr>
                </thead>
                <tbody>{review_rows}</tbody>
              </table>
            </section>
            """

        result_html = f"""
        <section class="atlas-stats">
          <div class="atlas-stat">
            <span>Pachete identificate</span>
            <strong>{len(assignments)}</strong>
          </div>
          <div class="atlas-stat">
            <span>Șoferi cu Atlas</span>
            <strong>{len(grouped)}</strong>
          </div>
          <div class="atlas-stat atlas-warn">
            <span>Necesită verificare</span>
            <strong>{len(review)}</strong>
          </div>
        </section>

        <div class="atlas-result-actions">
          <a
            class="atlas-btn atlas-primary"
            download="Atlas_Paket.xlsx"
            href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}"
          >Descarcă Excel</a>

          <button
            class="atlas-btn"
            type="button"
            onclick="copyAtlas()"
          >Copiază lista</button>
        </div>

        <section class="atlas-table">
          <table>
            <thead>
              <tr>
                <th>Șofer</th>
                <th>Route Code</th>
                <th>Pachete</th>
                <th>Tracking ID</th>
                <th>Status</th>
              </tr>
            </thead>
            <tbody>
              {
                assignment_rows
                or '<tr><td colspan="5">'
                   'Niciun pachet atribuit automat.'
                   '</td></tr>'
              }
            </tbody>
          </table>
        </section>

        {review_block}

        <script>
        const atlasCopyText = {
            json.dumps(copy_text, ensure_ascii=False)
        };

        async function copyAtlas() {{
          if (!atlasCopyText) {{
            alert("Nu există pachete de copiat.");
            return;
          }}

          try {{
            await navigator.clipboard.writeText(
              atlasCopyText
            );
            alert("Lista Atlas a fost copiată.");
          }} catch (error) {{
            window.prompt(
              "Copiază lista:",
              atlasCopyText
            );
          }}
        }}
        </script>
        """

    return f"""
<!doctype html>
<html lang="ro">
<head>
<meta charset="utf-8">
<meta
  name="viewport"
  content="width=device-width,initial-scale=1"
>
<title>Atlas Paket · FICO Control</title>
<style>
*{{box-sizing:border-box}}
body{{
  margin:0;
  background:#f4f6f8;
  color:#17212b;
  font-family:Arial,sans-serif
}}
.atlas-wrap{{
  width:min(96%,1200px);
  margin:28px auto 60px
}}
.atlas-top{{
  display:flex;
  justify-content:space-between;
  align-items:flex-start;
  gap:18px
}}
.atlas-brand{{
  font-size:12px;
  font-weight:900;
  letter-spacing:2px
}}
h1{{
  font-size:38px;
  margin:9px 0 5px
}}
.atlas-sub{{color:#667085}}
.atlas-actions{{
  display:flex;
  gap:8px;
  flex-wrap:wrap
}}
.atlas-btn{{
  display:inline-flex;
  align-items:center;
  justify-content:center;
  text-decoration:none;
  border:1px solid #d8dde3;
  border-radius:10px;
  padding:12px 15px;
  background:#fff;
  color:#17212b;
  font-weight:800;
  cursor:pointer
}}
.atlas-primary{{
  background:#17212b;
  color:#fff;
  border-color:#17212b
}}
.atlas-hero{{
  margin-top:22px;
  border-radius:18px;
  padding:26px;
  background:linear-gradient(
    120deg,
    #0d4f6b,
    #177e9c
  );
  color:#fff
}}
.atlas-hero h2{{margin:0 0 7px}}
.atlas-hero p{{
  margin:0 0 20px;
  color:#d5edf3;
  line-height:1.5
}}
.atlas-form{{
  display:grid;
  grid-template-columns:1fr 1fr auto;
  gap:12px;
  align-items:end
}}
.atlas-upload{{
  background:#fff;
  color:#17212b;
  padding:14px;
  border-radius:12px
}}
.atlas-upload label{{
  display:block;
  font-size:12px;
  font-weight:900;
  margin-bottom:8px
}}
.atlas-upload small{{
  display:block;
  color:#667085;
  margin-top:7px;
  line-height:1.4
}}
.atlas-upload input{{max-width:100%}}
.atlas-error{{
  margin-top:16px;
  padding:14px 16px;
  background:#fff0f0;
  border-left:4px solid #d92d20;
  color:#9f2f27
}}
.atlas-stats{{
  display:grid;
  grid-template-columns:repeat(3,1fr);
  gap:13px;
  margin-top:18px
}}
.atlas-stat{{
  background:#fff;
  border:1px solid #e4e7ec;
  border-radius:15px;
  padding:18px;
  display:flex;
  justify-content:space-between;
  align-items:center
}}
.atlas-stat span{{
  color:#667085;
  font-weight:800;
  font-size:13px
}}
.atlas-stat strong{{font-size:32px}}
.atlas-warn strong{{color:#b42318}}
.atlas-result-actions{{
  display:flex;
  gap:9px;
  margin:15px 0
}}
.atlas-table{{
  overflow:hidden;
  background:#fff;
  border:1px solid #e4e7ec;
  border-radius:15px;
  margin-top:14px
}}
.atlas-table table{{
  width:100%;
  border-collapse:collapse
}}
.atlas-table th,
.atlas-table td{{
  padding:13px 15px;
  border-bottom:1px solid #edf0f2;
  text-align:left;
  vertical-align:middle
}}
.atlas-table th{{
  background:#fafafa;
  color:#667085;
  font-size:11px;
  text-transform:uppercase
}}
.atlas-table code{{
  font-weight:800;
  color:#17212b
}}
.atlas-group-start td{{
  border-top:2px solid #d8dde3
}}
.atlas-group-cell{{
  background:#fbfcfd
}}
.atlas-route{{
  font-weight:800
}}
.atlas-count{{
  display:inline-flex;
  align-items:center;
  justify-content:center;
  min-width:88px;
  padding:7px 10px;
  border-radius:999px;
  background:#e7f8ef;
  color:#147a42;
  font-size:12px;
  font-weight:900;
  white-space:nowrap
}}
.atlas-ok{{
  background:#e9f8ef;
  color:#14804a;
  border-radius:999px;
  padding:6px 9px;
  font-size:11px;
  font-weight:900
}}
.atlas-review{{border-color:#f0c2bd}}
.atlas-review-title{{
  padding:15px;
  background:#fff7f6;
  color:#b42318;
  font-weight:900
}}
.atlas-review-row{{background:#fffafa}}
@media(max-width:850px){{
  .atlas-top{{display:block}}
  .atlas-actions{{margin-top:14px}}
  .atlas-form{{grid-template-columns:1fr}}
  .atlas-stats{{grid-template-columns:1fr}}
}}
</style>
</head>
<body>
<main class="atlas-wrap">
  <div class="atlas-top">
    <div>
      <div class="atlas-brand">
        FICO CONTROL
      </div>
      <h1>Atlas Paket</h1>
      <div class="atlas-sub">
        Cortex + pozele Amazon Atlas →
        șoferul și pachetele corecte
      </div>
    </div>

    <div class="atlas-actions">
      <a
        class="atlas-btn"
        href="/admin"
      >FICO Dashboard</a>
      <a
        class="atlas-btn"
        href="/admin/mentor"
      >Mentor Check</a>
      <a
        class="atlas-btn"
        href="/admin/hours"
      >Control ore</a>
      <a
        class="atlas-btn"
        href="/admin/pod-ccc"
      >POD & CCC</a>
      <a
        class="atlas-btn"
        href="/admin/concessions"
      >Concesii</a>
    </div>
  </div>

  {error_html}

  <section class="atlas-hero">
    <h2>Generează distribuirea Atlas</h2>

    <p>
      Încarcă Excelul Cortex cu rutele normale
      și toate pozele Transfer Sheet primite de
      la Amazon. Sistemul afișează numărul de
      pachete pentru fiecare rută, inclusiv când
      există un singur pachet.
    </p>

    <form
      class="atlas-form"
      method="post"
      action="/admin/atlas-paket"
      enctype="multipart/form-data"
    >
      <div class="atlas-upload">
        <label>1. Excel Cortex</label>

        <input
          type="file"
          name="cortex_file"
          accept=".xlsx"
          required
        >

        <small>
          Routencode + primul nume din
          Name des Fahrers. Rescue-ul după |
          este ignorat.
        </small>
      </div>

      <div class="atlas-upload">
        <label>2. Poze Atlas Amazon</label>

        <input
          type="file"
          name="atlas_images"
          accept="image/jpeg,image/png,image/webp"
          multiple
          required
        >

        <small>
          Selectează toate paginile.
          Page 2 poate continua ruta începută
          la finalul Page 1.
        </small>
      </div>

      <button
        class="atlas-btn atlas-primary"
        type="submit"
      >Generează Atlas</button>
    </form>
  </section>

  {result_html}
</main>
</body>
</html>
"""

@app.get("/admin/atlas-paket", response_class=HTMLResponse)
def atlas_paket_page():
    return HTMLResponse(atlas_page_html())


@app.post("/admin/atlas-paket", response_class=HTMLResponse)
async def atlas_paket_generate(
    cortex_file: UploadFile = File(...),
    atlas_images: list[UploadFile] = File(...)
):
    try:
        if not (cortex_file.filename or "").lower().endswith(".xlsx"):
            raise ValueError(
                "Cortex trebuie să fie un fișier XLSX."
            )

        cortex_raw = await cortex_file.read(
            15 * 1024 * 1024 + 1
        )

        if len(cortex_raw) > 15 * 1024 * 1024:
            raise ValueError(
                "Excelul Cortex este prea mare."
            )

        cortex_routes = atlas_extract_cortex_routes(
            cortex_raw
        )

        images = list(atlas_images or [])

        if not images:
            raise ValueError(
                "Selectează cel puțin o poză Atlas."
            )

        if len(images) > ATLAS_MAX_IMAGES:
            raise ValueError(
                f"Poți încărca maximum "
                f"{ATLAS_MAX_IMAGES} poze Atlas odată."
            )

        pages = []

        for upload_index, image in enumerate(images):
            content_type = (
                image.content_type or ""
            ).lower()

            if content_type not in ATLAS_ALLOWED_TYPES:
                raise ValueError(
                    f"{image.filename or 'O poză'} "
                    f"nu este JPG, PNG sau WEBP."
                )

            raw = await image.read(
                ATLAS_MAX_IMAGE_BYTES + 1
            )

            if not raw:
                raise ValueError(
                    f"{image.filename or 'O poză'} "
                    f"este goală."
                )

            if len(raw) > ATLAS_MAX_IMAGE_BYTES:
                raise ValueError(
                    f"{image.filename or 'O poză'} "
                    f"depășește 12 MB."
                )

            pages.append(
                atlas_ocr_page(
                    raw,
                    content_type,
                    image.filename or "Atlas",
                    upload_index
                )
            )

        assignments, review = atlas_build_assignments(
            cortex_routes,
            pages
        )

        if not assignments and not review:
            raise ValueError(
                "Nu am identificat niciun pachet Atlas "
                "în pozele încărcate."
            )

        return HTMLResponse(
            atlas_page_html(assignments, review)
        )

    except ValueError as exc:
        return HTMLResponse(
            atlas_page_html(error=str(exc)),
            status_code=400
        )

    finally:
        await cortex_file.close()

        for image in atlas_images or []:
            await image.close()
# ============================================================================
# ATLAS PAKET V4 - MULTI-PAGE + TOTAL PACKAGES VALIDATION
# PASTE THIS BLOCK AT THE VERY END OF backend/app_mobile_api.py
# ============================================================================

_ATLAS_V4_ORIGINAL_BUILD_ASSIGNMENTS = atlas_build_assignments
_ATLAS_V4_ORIGINAL_PAGE_HTML = atlas_page_html


def _atlas_v4_expected_total(text):
    """
    Detect Amazon totals such as:
      Total: 28 packages
      Total Packages: 28
      Total 28 Packages
    """
    upper = str(text or "").upper()

    patterns = (
        r"\bTOTAL\s*PACKAGES?\s*[:\-]?\s*(\d{1,4})\b",
        r"\bTOTAL\s*[:\-]?\s*(\d{1,4})\s*PACKAGES?\b",
        r"\bTOTAL\s+(\d{1,4})\s*PACKAGES?\b",
    )

    for pattern in patterns:
        match = re.search(pattern, upper)
        if match:
            try:
                return int(match.group(1))
            except (TypeError, ValueError):
                pass

    return None


def _atlas_v4_page_info(text):
    """
    Detect:
      Page 1 of 2
      Page 1/2
    """
    match = re.search(
        r"\bPAGE\s*(\d+)\s*(?:OF|/)\s*(\d+)\b",
        str(text or ""),
        flags=re.IGNORECASE,
    )

    if not match:
        return None, None

    try:
        return int(match.group(1)), int(match.group(2))
    except (TypeError, ValueError):
        return None, None


def _atlas_v4_build_assignments(cortex_routes, pages):
    """
    V4:
    1. keeps the existing Atlas logic;
    2. additionally scans ParsedText line-by-line so tracking IDs missed by
       OCR overlay can still be recovered;
    3. preserves current_route between Page 1 and Page 2;
    4. validates Page X of Y;
    5. validates Amazon Total packages against all detected tracking IDs.
    """
    assignments, review = _ATLAS_V4_ORIGINAL_BUILD_ASSIGNMENTS(
        cortex_routes,
        pages,
    )

    pages_sorted = sorted(
        pages,
        key=lambda page: (
            page.get("page_number") is None,
            (
                page.get("page_number")
                if page.get("page_number") is not None
                else page.get("upload_index", 0)
            ),
            page.get("upload_index", 0),
        ),
    )

    # Everything already recognized by V2/V3.
    seen_tracking = {
        item.get("tracking")
        for item in assignments
        if item.get("tracking") and item.get("tracking") != "—"
    }

    # IMPORTANT:
    # Do not put review tracking IDs in seen_tracking here.
    # ParsedText fallback may be able to recover a route for an item that the
    # overlay parser could not safely assign.
    current_route = None

    for page in pages_sorted:
        filename = page.get("filename") or "Atlas"
        page_text = str(page.get("text") or "")

        # ParsedText has a stable top-to-bottom order and often contains rows
        # that OCR overlay failed to expose correctly.
        for raw_line in page_text.splitlines():
            line = raw_line.strip()
            if not line:
                continue

            row_routes = atlas_route_codes_from_text(line)

            if row_routes:
                unique_routes = list(dict.fromkeys(row_routes))
                cortex_candidates = [
                    route
                    for route in unique_routes
                    if route in cortex_routes
                ]

                if len(cortex_candidates) == 1:
                    current_route = cortex_candidates[0]
                elif len(unique_routes) == 1:
                    current_route = unique_routes[0]
                else:
                    # Ambiguous row: do not guess.
                    current_route = None

            row_tracking_ids = atlas_tracking_ids_from_text(line)

            for tracking_id in row_tracking_ids:
                if tracking_id in seen_tracking:
                    continue

                if not current_route:
                    # Keep it for manual review rather than silently losing it.
                    if not any(
                        item.get("tracking") == tracking_id
                        for item in review
                    ):
                        review.append({
                            "file": filename,
                            "route": "—",
                            "tracking": tracking_id,
                            "reason": (
                                "Tracking ID citit, dar ruta nu a putut fi "
                                "identificată sigur"
                            ),
                        })
                    continue

                if current_route not in cortex_routes:
                    if not any(
                        item.get("tracking") == tracking_id
                        for item in review
                    ):
                        review.append({
                            "file": filename,
                            "route": current_route,
                            "tracking": tracking_id,
                            "reason": "Ruta nu există în Excelul Cortex",
                        })
                    continue

                # If the original parser placed this tracking in review,
                # remove that review row now because V4 recovered it.
                review = [
                    item
                    for item in review
                    if item.get("tracking") != tracking_id
                ]

                assignments.append({
                    "driver": cortex_routes[current_route],
                    "route": current_route,
                    "tracking": tracking_id,
                    "file": filename,
                })
                seen_tracking.add(tracking_id)

    # De-duplicate assignments safely by Tracking ID.
    unique_assignments = {}
    for item in assignments:
        tracking_id = item.get("tracking")
        if not tracking_id:
            continue
        unique_assignments.setdefault(tracking_id, item)

    assignments = list(unique_assignments.values())

    assignments.sort(
        key=lambda item: (
            item["driver"].casefold(),
            item["route"],
            item["tracking"],
        )
    )

    # ------------------------------------------------------------------
    # PAGE COVERAGE CHECK
    # ------------------------------------------------------------------
    page_info = []

    for page in pages_sorted:
        page_number = page.get("page_number")
        page_total = page.get("page_total")

        if page_number is None or page_total is None:
            page_number, page_total = _atlas_v4_page_info(
                page.get("text")
            )

        if page_number is not None and page_total is not None:
            page_info.append((page_number, page_total))

    if page_info:
        expected_page_total = max(total for _, total in page_info)
        received_pages = {
            number
            for number, _ in page_info
        }

        missing_pages = [
            number
            for number in range(1, expected_page_total + 1)
            if number not in received_pages
        ]

        if missing_pages:
            review.append({
                "file": "Atlas",
                "route": "—",
                "tracking": "—",
                "reason": (
                    "LIPSESC PAGINI ATLAS: "
                    + ", ".join(str(number) for number in missing_pages)
                    + f" din {expected_page_total}. "
                    "Selectează toate paginile în aceeași încărcare."
                ),
            })

    # ------------------------------------------------------------------
    # AMAZON TOTAL PACKAGES CHECK
    # ------------------------------------------------------------------
    expected_totals = []

    for page in pages_sorted:
        value = _atlas_v4_expected_total(page.get("text"))
        if value:
            expected_totals.append(value)

    expected_total = (
        max(expected_totals)
        if expected_totals
        else None
    )

    all_detected_ids = {
        item.get("tracking")
        for item in assignments
        if item.get("tracking") and item.get("tracking") != "—"
    }

    all_detected_ids.update(
        item.get("tracking")
        for item in review
        if item.get("tracking") and item.get("tracking") != "—"
    )

    detected_total = len(all_detected_ids)

    if expected_total is not None and detected_total != expected_total:
        difference = expected_total - detected_total

        if difference > 0:
            message = (
                f"TOTAL INCOMPLET: Amazon arată {expected_total} pachete, "
                f"dar sistemul a citit {detected_total}. "
                f"Lipsesc {difference} pachete."
            )
        else:
            message = (
                f"TOTAL DIFERIT: Amazon arată {expected_total} pachete, "
                f"dar sistemul a citit {detected_total}. "
                "Verifică duplicatele OCR."
            )

        review.append({
            "file": "Atlas",
            "route": "—",
            "tracking": "—",
            "reason": message,
        })

    return assignments, review


def _atlas_v4_page_html(assignments=None, review=None, error=""):
    """
    Adds:
    - selected-image counter before submit;
    - strong red warning for missing pages / total mismatch.
    """
    assignments = assignments or []
    review = review or []

    page = _ATLAS_V4_ORIGINAL_PAGE_HTML(
        assignments,
        review,
        error,
    )

    # Add an ID to the multi-file input.
    page = page.replace(
        'name="atlas_images"\n          accept=',
        'id="atlasImagesInput"\n'
        '          name="atlas_images"\n'
        '          accept=',
        1,
    )

    # Strong warning above the result cards.
    important_warnings = [
        str(item.get("reason") or "")
        for item in review
        if (
            str(item.get("reason") or "").startswith("TOTAL ")
            or str(item.get("reason") or "").startswith("LIPSESC PAGINI")
        )
    ]

    if important_warnings:
        warning_html = (
            '<div style="margin:18px 0;padding:17px 19px;'
            'border-radius:14px;background:#fff0f0;'
            'border:2px solid #d92d20;color:#b42318;'
            'font-weight:900;font-size:15px;line-height:1.55">'
            '⚠ '
            + "<br>⚠ ".join(
                html.escape(message)
                for message in important_warnings
            )
            + "</div>"
        )

        page = page.replace(
            '<section class="atlas-stats">',
            warning_html + '<section class="atlas-stats">',
            1,
        )

    selection_script = r"""
<script>
(() => {
  const input = document.getElementById('atlasImagesInput');
  if (!input) return;

  const box = document.createElement('div');
  box.id = 'atlasSelectedFiles';
  box.style.cssText =
    'margin-top:8px;padding:7px 9px;border-radius:8px;'
    + 'background:#e9f8ef;color:#147a42;'
    + 'font-size:12px;font-weight:900;line-height:1.45';

  input.insertAdjacentElement('afterend', box);

  function refreshAtlasFiles() {
    const files = Array.from(input.files || []);

    if (!files.length) {
      box.textContent = '0 poze Atlas selectate';
      return;
    }

    box.textContent =
      files.length
      + (files.length === 1 ? ' poză Atlas selectată: ' : ' poze Atlas selectate: ')
      + files.map(file => file.name).join(' · ');
  }

  input.addEventListener('change', refreshAtlasFiles);
  refreshAtlasFiles();
})();
</script>
"""

    page = page.replace(
        "</body>",
        selection_script + "</body>",
        1,
    )

    return page


# Replace the live functions used by the already-registered FastAPI routes.
atlas_build_assignments = _atlas_v4_build_assignments
atlas_page_html = _atlas_v4_page_html

print(
    "ATLAS_PAKET_V4_DIRECT_PATCH_LOADED",
    flush=True,
)


# ============================================================================
# ATLAS PAKET V5 - PER-IMAGE OCR DIAGNOSTICS + RETRY
# PASTE THIS BLOCK AT THE VERY END OF backend/app_mobile_api.py
# ============================================================================

_ATLAS_V5_ORIGINAL_OCR_PAGE = atlas_ocr_page
_ATLAS_V5_ORIGINAL_BUILD_ASSIGNMENTS = atlas_build_assignments
_ATLAS_V5_ORIGINAL_PAGE_HTML = atlas_page_html


def _atlas_v5_count_tracking_from_page(page):
    ids = set()

    for line in page.get("lines") or []:
        text = str(line.get("text") or "")
        for tracking in atlas_tracking_ids_from_text(text):
            ids.add(tracking)

    for raw_line in str(page.get("text") or "").splitlines():
        for tracking in atlas_tracking_ids_from_text(raw_line):
            ids.add(tracking)

    return sorted(ids)


def _atlas_v5_ocr_request(raw, content_type, filename, upload_index, *, engine="2", table=True):
    api_key = os.getenv("OCRSPACE_API_KEY", "").strip()

    if not api_key:
        raise ValueError(
            "OCRSPACE_API_KEY nu este configurat în Render. "
            "Atlas Paket are nevoie de OCR pentru a citi pozele Amazon."
        )

    boundary = "----AtlasPaketV5Boundary7MA4YWxkTrZu0gW"

    extension = "jpg"
    if content_type == "image/png":
        extension = "png"
    elif content_type == "image/webp":
        extension = "webp"

    parts = []

    def add_field(name, value):
        parts.append(
            (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n"
            ).encode("utf-8")
        )

    add_field("apikey", api_key)
    add_field("language", "eng")
    add_field("isOverlayRequired", "true")
    add_field("isTable", "true" if table else "false")
    add_field("OCREngine", str(engine))
    add_field("scale", "true")
    add_field("detectOrientation", "true")

    parts.append(
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; '
            f'filename="atlas.{extension}"\r\n'
            f"Content-Type: {content_type}\r\n\r\n"
        ).encode("utf-8")
        + raw
        + b"\r\n"
    )

    parts.append(
        f"--{boundary}--\r\n".encode("utf-8")
    )

    request = urllib.request.Request(
        "https://api.ocr.space/parse/image",
        data=b"".join(parts),
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "User-Agent": "FICO-Control-Atlas/5.0",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = json.loads(
                response.read().decode("utf-8", errors="replace")
            )
    except Exception as exc:
        raise ValueError(
            f"OCR-ul nu a putut citi poza {filename}."
        ) from exc

    if payload.get("IsErroredOnProcessing"):
        message = (
            payload.get("ErrorMessage")
            or payload.get("ErrorDetails")
        )

        if isinstance(message, list):
            message = " ".join(str(item) for item in message)

        raise ValueError(
            f"OCR Atlas pentru {filename}: "
            + str(message or "eroare necunoscută")
        )

    parsed_results = payload.get("ParsedResults") or []

    if not parsed_results:
        raise ValueError(
            f"Poza {filename} nu a produs niciun rezultat OCR."
        )

    parsed_text_parts = []
    line_records = []
    synthetic_top = 0.0

    for result in parsed_results:
        parsed_text = str(result.get("ParsedText") or "")
        if parsed_text:
            parsed_text_parts.append(parsed_text)

        overlay = result.get("TextOverlay") or {}
        overlay_lines = overlay.get("Lines") or []

        for line in overlay_lines:
            words = line.get("Words") or []
            line_text = str(
                line.get("LineText") or ""
            ).strip()

            if not line_text and words:
                line_text = " ".join(
                    str(
                        word.get("WordText") or ""
                    ).strip()
                    for word in words
                    if str(
                        word.get("WordText") or ""
                    ).strip()
                )

            if not line_text:
                continue

            top = line.get("MinTop")
            if top is None and words:
                top = min(
                    float(word.get("Top") or 0)
                    for word in words
                )

            if top is None:
                top = synthetic_top

            left = 0.0
            if words:
                left = min(
                    float(word.get("Left") or 0)
                    for word in words
                )

            height = line.get("MaxHeight")

            if height is None and words:
                height = max(
                    float(word.get("Height") or 0)
                    for word in words
                )

            if not height:
                height = 12.0

            line_records.append({
                "text": line_text,
                "top": float(top),
                "left": float(left),
                "height": float(height),
            })

            synthetic_top = max(
                synthetic_top + 18.0,
                float(top)
                + float(height)
                + 3.0,
            )

    combined_text = "\n".join(
        parsed_text_parts
    ).strip()

    if not line_records:
        for line_number, line_text in enumerate(
            combined_text.splitlines()
        ):
            line_text = line_text.strip()

            if not line_text:
                continue

            line_records.append({
                "text": line_text,
                "top": float(line_number * 22),
                "left": 0.0,
                "height": 14.0,
            })

    page_number = None
    page_total = None

    page_match = re.search(
        r"\bPAGE\s*(\d+)\s*(?:OF|/)\s*(\d+)\b",
        combined_text,
        flags=re.IGNORECASE,
    )

    if page_match:
        page_number = int(page_match.group(1))
        page_total = int(page_match.group(2))

    return {
        "filename": filename,
        "upload_index": upload_index,
        "page_number": page_number,
        "page_total": page_total,
        "text": combined_text,
        "lines": line_records,
        "ocr_engine": str(engine),
        "ocr_table": bool(table),
    }


def _atlas_v5_ocr_page(raw, content_type, filename, upload_index):
    """
    First pass: existing OCR logic.
    Retry automatically if the first pass reads too few tracking IDs.
    """
    first = _ATLAS_V5_ORIGINAL_OCR_PAGE(
        raw,
        content_type,
        filename,
        upload_index,
    )

    first_ids = _atlas_v5_count_tracking_from_page(first)

    # If page looks healthy, keep it.
    if len(first_ids) >= 2:
        first["tracking_count"] = len(first_ids)
        first["ocr_attempts"] = 1
        first["ocr_retry_used"] = False
        return first

    # Retry with alternate OCR mode.
    try:
        second = _atlas_v5_ocr_request(
            raw,
            content_type,
            filename,
            upload_index,
            engine="1",
            table=False,
        )
        second_ids = _atlas_v5_count_tracking_from_page(second)

        if len(second_ids) > len(first_ids):
            second["tracking_count"] = len(second_ids)
            second["ocr_attempts"] = 2
            second["ocr_retry_used"] = True
            return second
    except Exception as exc:
        print(
            "ATLAS_V5_RETRY_ERROR:",
            filename,
            type(exc).__name__,
            str(exc)[:300],
            flush=True,
        )

    first["tracking_count"] = len(first_ids)
    first["ocr_attempts"] = 2
    first["ocr_retry_used"] = True
    return first


def _atlas_v5_build_assignments(cortex_routes, pages):
    assignments, review = _ATLAS_V5_ORIGINAL_BUILD_ASSIGNMENTS(
        cortex_routes,
        pages,
    )

    # Per-image diagnostics.
    for page in sorted(
        pages,
        key=lambda item: item.get("upload_index", 0),
    ):
        count = page.get("tracking_count")
        if count is None:
            count = len(
                _atlas_v5_count_tracking_from_page(page)
            )

        if count == 0:
            review.append({
                "file": page.get("filename") or "Atlas",
                "route": "—",
                "tracking": "—",
                "reason": (
                    "OCR PAGINĂ: această poză a produs 0 Tracking ID-uri. "
                    "Sistemul a încercat automat și a doua metodă OCR."
                ),
            })

    return assignments, review


def _atlas_v5_page_html(assignments=None, review=None, error=""):
    page = _ATLAS_V5_ORIGINAL_PAGE_HTML(
        assignments,
        review,
        error,
    )

    # Add V5 marker so we can visually confirm this patch is active.
    marker = (
        '<div style="margin-top:8px;font-size:11px;'
        'font-weight:900;color:#d5edf3">'
        'Atlas OCR V5 activ · verificare pe fiecare poză + retry automat'
        '</div>'
    )

    page = page.replace(
        "</p>",
        "</p>" + marker,
        1,
    )

    return page


# Replace live functions.
atlas_ocr_page = _atlas_v5_ocr_page
atlas_build_assignments = _atlas_v5_build_assignments
atlas_page_html = _atlas_v5_page_html

print(
    "ATLAS_PAKET_V5_OCR_RETRY_LOADED",
    flush=True,
)
# ============================================================================
# ATLAS PAKET V6 - TWO SEPARATE PAGE INPUTS
# PASTE THIS BLOCK AT THE VERY END OF backend/app_mobile_api.py
# ============================================================================

_ATLAS_V6_ORIGINAL_PAGE_HTML = atlas_page_html
_ATLAS_V6_ORIGINAL_BUILD_ASSIGNMENTS = atlas_build_assignments


def _atlas_v6_page_html(assignments=None, review=None, error=""):
    page = _ATLAS_V6_ORIGINAL_PAGE_HTML(assignments, review, error)

    old_block_1 = '''<input
          id="atlasImagesInput"
          type="file"
          name="atlas_images"
          accept="image/jpeg,image/png,image/webp"
          multiple
          required
        >'''

    old_block_2 = '''<input
          type="file"
          name="atlas_images"
          accept="image/jpeg,image/png,image/webp"
          multiple
          required
        >'''

    new_block = '''<div style="display:grid;gap:9px">
          <div>
            <div style="font-size:12px;font-weight:900;margin-bottom:5px">Pagina 1 Atlas</div>
            <input id="atlasPage1" type="file" name="atlas_images"
                   accept="image/jpeg,image/png,image/webp" required>
          </div>
          <div>
            <div style="font-size:12px;font-weight:900;margin-bottom:5px">Pagina 2 Atlas</div>
            <input id="atlasPage2" type="file" name="atlas_images"
                   accept="image/jpeg,image/png,image/webp" required>
          </div>
          <div id="atlasPagesSelected"
               style="padding:8px 10px;border-radius:8px;background:#fff4d6;
                      color:#8a5a00;font-size:12px;font-weight:900">
            Selectează separat Page 1 și Page 2
          </div>
        </div>'''

    if old_block_1 in page:
        page = page.replace(old_block_1, new_block, 1)
    elif old_block_2 in page:
        page = page.replace(old_block_2, new_block, 1)

    script = r'''
<script>
(() => {
  const old = document.getElementById('atlasSelectedFiles');
  if (old) old.remove();

  const page1 = document.getElementById('atlasPage1');
  const page2 = document.getElementById('atlasPage2');
  const box = document.getElementById('atlasPagesSelected');
  if (!page1 || !page2 || !box) return;

  function refresh() {
    const f1 = page1.files && page1.files[0];
    const f2 = page2.files && page2.files[0];

    if (f1 && f2) {
      box.textContent = '2/2 pagini selectate: ' + f1.name + ' · ' + f2.name;
      box.style.background = '#e9f8ef';
      box.style.color = '#147a42';
    } else if (f1 || f2) {
      box.textContent = '1/2 pagini selectate · mai selectează încă o pagină';
      box.style.background = '#fff4d6';
      box.style.color = '#8a5a00';
    } else {
      box.textContent = 'Selectează separat Page 1 și Page 2';
      box.style.background = '#fff4d6';
      box.style.color = '#8a5a00';
    }
  }

  page1.addEventListener('change', refresh);
  page2.addEventListener('change', refresh);
  refresh();
})();
</script>
'''
    page = page.replace("</body>", script + "</body>", 1)

    marker = (
        '<div style="margin-top:6px;font-size:11px;font-weight:900;color:#d5edf3">'
        'Atlas V6 activ · Page 1 și Page 2 se încarcă separat'
        '</div>'
    )

    if "Atlas V6 activ" not in page:
        page = page.replace(
            "Atlas OCR V5 activ · verificare pe fiecare poză + retry automat",
            "Atlas OCR V5 activ · verificare pe fiecare poză + retry automat" + marker,
            1,
        )

    return page


def _atlas_v6_build_assignments(cortex_routes, pages):
    assignments, review = _ATLAS_V6_ORIGINAL_BUILD_ASSIGNMENTS(cortex_routes, pages)

    if len(pages) < 2:
        review.append({
            "file": "Atlas",
            "route": "—",
            "tracking": "—",
            "reason": (
                f"UPLOAD INCOMPLET: serverul a primit doar {len(pages)} "
                "pagină/pagini Atlas din 2."
            ),
        })

    return assignments, review


atlas_page_html = _atlas_v6_page_html
atlas_build_assignments = _atlas_v6_build_assignments

print("ATLAS_PAKET_V6_SEPARATE_UPLOADS_LOADED", flush=True)
